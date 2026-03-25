"""
Centralized export views for Sales and Cheque Collections reports.
Generates multi-sheet Excel exports with openpyxl.

Supports both synchronous (GET → blob) and background (POST → WebSocket
notification with download token) export modes.
"""
import logging
import os
import threading
import uuid
from collections import defaultdict
from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.conf import settings
from django.db.models import Count, Sum, Value
from django.db.models.functions import Coalesce, TruncDate
from django.http import FileResponse, HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import status as drf_status
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from rest_framework.views import APIView

logger = logging.getLogger(__name__)

# ── Temp export directory ──────────────────────────────────────────
EXPORT_DIR = os.path.join(settings.BASE_DIR, "media", "exports")
EXPORT_MAX_AGE_SECONDS = 300  # auto-delete after 5 min


def _ensure_export_dir():
    os.makedirs(EXPORT_DIR, exist_ok=True)


def _cleanup_old_exports():
    """Remove export files older than EXPORT_MAX_AGE_SECONDS."""
    try:
        now = timezone.now().timestamp()
        for fname in os.listdir(EXPORT_DIR):
            fpath = os.path.join(EXPORT_DIR, fname)
            if os.path.isfile(fpath):
                age = now - os.path.getmtime(fpath)
                if age > EXPORT_MAX_AGE_SECONDS:
                    os.remove(fpath)
    except OSError:
        pass


# ── Shared styling ─────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
GREEN_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
AMBER_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
BLUE_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, col_count, row=1):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def _write_rows(ws, headers, rows, start_row=1):
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col_idx, value=header)
    _style_header(ws, len(headers), start_row)
    for row_idx, row_data in enumerate(rows, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
    _auto_width(ws)
    return start_row + len(rows)


def _parse_date(val, default=None):
    """Parse a date string (YYYY-MM-DD) into a date object."""
    from datetime import datetime
    if not val:
        return default
    try:
        return datetime.strptime(val, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return default


def _workbook_response(wb, prefix):
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    timestamp = timezone.now().strftime("%Y%m%d_%H%M")
    filename = f"{prefix}_{timestamp}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _save_workbook(wb, prefix):
    """Save workbook to temp dir and return (token, filename)."""
    _ensure_export_dir()
    _cleanup_old_exports()
    token = uuid.uuid4().hex
    timestamp = timezone.now().strftime("%Y%m%d_%H%M")
    filename = f"{prefix}_{timestamp}.xlsx"
    safe_name = f"{token}_{filename}"
    fpath = os.path.join(EXPORT_DIR, safe_name)
    wb.save(fpath)
    return token, filename


# ═══════════════════════════════════════════════════════════════════
#  SALES WORKBOOK BUILDER
# ═══════════════════════════════════════════════════════════════════

SALES_VALID_SHEETS = {
    "all_transactions", "daily_summary", "monthly_summary",
    "quarterly_summary", "payment_breakdown", "top_items",
}


def _effective_date(txn):
    """Return transaction_date if set, otherwise created_at.date()."""
    return txn.transaction_date or txn.created_at.date()


def build_sales_workbook(start, end, requested):
    """Build and return a sales report Workbook."""
    from sales.models import SalesItem, SalesTransaction

    txns = (
        SalesTransaction.objects
        .annotate(effective_date=Coalesce("transaction_date", TruncDate("created_at")))
        .filter(
            is_deleted=False, voided=False,
            effective_date__gte=start,
            effective_date__lte=end,
        )
        .select_related("stall", "client", "sales_clerk")
        .prefetch_related("items__item", "payments")
        .order_by("effective_date", "created_at")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="RVDC Sales Report").font = Font(bold=True, size=16, color="2563EB")
    ws.cell(row=2, column=1, value=f"Period: {start.strftime('%b %d, %Y')} — {end.strftime('%b %d, %Y')}")
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666")
    ws.cell(row=3, column=1, value=f"Generated: {timezone.now().strftime('%B %d, %Y %I:%M %p')}")
    ws.cell(row=3, column=1).font = Font(italic=True, color="666666")

    total_count = txns.count()
    total_sales = sum(float(t.computed_total) for t in txns)
    total_paid = sum(float(t.total_paid) for t in txns)

    stats = [
        ("Total Transactions", total_count),
        ("Total Sales", f"₱{total_sales:,.2f}"),
        ("Total Collected", f"₱{total_paid:,.2f}"),
        ("Outstanding Balance", f"₱{max(total_sales - total_paid, 0):,.2f}"),
    ]
    row = 5
    ws.cell(row=row, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=2).border = THIN_BORDER
    for label, value in stats:
        row += 1
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=value).border = THIN_BORDER
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25

    if "all_transactions" in requested:
        ws_t = wb.create_sheet("All Transactions")
        headers = [
            "ID", "Date", "Client", "Stall", "Clerk", "Type",
            "Items", "Subtotal", "Discount", "Total",
            "Paid", "Balance", "Status",
        ]
        rows_data = []
        for t in txns:
            subtotal = float(t.subtotal)
            total = float(t.computed_total)
            paid = float(t.total_paid)
            rows_data.append([
                t.id,
                _effective_date(t).strftime("%Y-%m-%d"),
                t.client.full_name if t.client else "Walk-in",
                t.stall.name if t.stall else "—",
                t.sales_clerk.get_full_name() if t.sales_clerk else "—",
                t.get_transaction_type_display(),
                t.total_items,
                subtotal, float(t.order_discount or 0),
                total, paid, max(total - paid, 0),
                t.get_payment_status_display(),
            ])
        last = _write_rows(ws_t, headers, rows_data)
        stat_col = headers.index("Status") + 1
        for ri in range(2, last + 1):
            cell = ws_t.cell(row=ri, column=stat_col)
            if cell.value == "Paid":
                for c in range(1, len(headers) + 1):
                    ws_t.cell(row=ri, column=c).fill = GREEN_FILL
            elif cell.value == "Unpaid":
                for c in range(1, len(headers) + 1):
                    ws_t.cell(row=ri, column=c).fill = RED_FILL
            elif cell.value == "Partial":
                for c in range(1, len(headers) + 1):
                    ws_t.cell(row=ri, column=c).fill = AMBER_FILL

    if "daily_summary" in requested:
        ws_d = wb.create_sheet("Daily Summary")
        daily_map = defaultdict(lambda: {"count": 0, "total": Decimal("0"), "paid": Decimal("0")})
        for t in txns:
            day = _effective_date(t)
            daily_map[day]["count"] += 1
            daily_map[day]["total"] += t.computed_total
            daily_map[day]["paid"] += t.total_paid
        headers = ["Date", "Transactions", "Total Sales", "Total Collected", "Outstanding"]
        rows_data = []
        for day in sorted(daily_map.keys()):
            d = daily_map[day]
            rows_data.append([
                day.strftime("%Y-%m-%d"), d["count"],
                float(d["total"]), float(d["paid"]),
                float(max(d["total"] - d["paid"], 0)),
            ])
        _write_rows(ws_d, headers, rows_data)

    if "monthly_summary" in requested:
        ws_m = wb.create_sheet("Monthly Summary")
        monthly_map = defaultdict(lambda: {"count": 0, "total": Decimal("0"), "paid": Decimal("0")})
        for t in txns:
            d = _effective_date(t)
            key = d.strftime("%Y-%m")
            monthly_map[key]["count"] += 1
            monthly_map[key]["total"] += t.computed_total
            monthly_map[key]["paid"] += t.total_paid
        headers = ["Month", "Transactions", "Total Sales", "Total Collected", "Outstanding"]
        rows_data = []
        for month in sorted(monthly_map.keys()):
            d = monthly_map[month]
            rows_data.append([
                month, d["count"], float(d["total"]),
                float(d["paid"]), float(max(d["total"] - d["paid"], 0)),
            ])
        _write_rows(ws_m, headers, rows_data)

    if "quarterly_summary" in requested:
        ws_q = wb.create_sheet("Quarterly Summary")
        quarterly_map = defaultdict(lambda: {"count": 0, "total": Decimal("0"), "paid": Decimal("0")})
        for t in txns:
            d = _effective_date(t)
            q = (d.month - 1) // 3 + 1
            key = f"{d.year} Q{q}"
            quarterly_map[key]["count"] += 1
            quarterly_map[key]["total"] += t.computed_total
            quarterly_map[key]["paid"] += t.total_paid
        headers = ["Quarter", "Transactions", "Total Sales", "Total Collected", "Outstanding"]
        rows_data = []
        for quarter in sorted(quarterly_map.keys()):
            d = quarterly_map[quarter]
            rows_data.append([
                quarter, d["count"], float(d["total"]),
                float(d["paid"]), float(max(d["total"] - d["paid"], 0)),
            ])
        _write_rows(ws_q, headers, rows_data)

    if "payment_breakdown" in requested:
        ws_p = wb.create_sheet("Payment Breakdown")
        payment_map = defaultdict(lambda: {"count": 0, "amount": Decimal("0")})
        for t in txns:
            for p in t.payments.all():
                payment_map[p.get_payment_type_display()]["count"] += 1
                payment_map[p.get_payment_type_display()]["amount"] += p.amount
        headers = ["Payment Type", "Count", "Total Amount"]
        rows_data = []
        for ptype in sorted(payment_map.keys()):
            d = payment_map[ptype]
            rows_data.append([ptype, d["count"], float(d["amount"])])
        _write_rows(ws_p, headers, rows_data)

    if "top_items" in requested:
        ws_i = wb.create_sheet("Top Items Sold")
        items_agg = (
            SalesItem.objects.filter(transaction__in=txns, item__isnull=False)
            .values("item__name", "item__sku", "item__category__name")
            .annotate(
                total_qty=Coalesce(Sum("quantity"), Decimal("0")),
                total_revenue=Coalesce(Sum("quantity") * Sum("final_price_per_unit") / Count("id"), Decimal("0")),
            )
            .order_by("-total_qty")[:50]
        )
        headers = ["Rank", "Item Name", "SKU", "Category", "Qty Sold", "Revenue"]
        rows_data = []
        for rank, r in enumerate(items_agg, 1):
            item_sales = SalesItem.objects.filter(
                transaction__in=txns, item__name=r["item__name"], item__isnull=False,
            )
            revenue = sum(float(si.line_total) for si in item_sales)
            rows_data.append([
                rank, r["item__name"], r["item__sku"] or "—",
                r["item__category__name"] or "—", float(r["total_qty"]), revenue,
            ])
        _write_rows(ws_i, headers, rows_data)

    return wb


# ═══════════════════════════════════════════════════════════════════
#  DAILY SALES WORKBOOK BUILDER (1 tab per day, Qty / Item / Amount)
# ═══════════════════════════════════════════════════════════════════


def build_daily_sales_workbook(start, end, stall_type):
    """
    Build a workbook with one sheet per day.
    Each sheet has columns: Qty, Item, Amount.
    Filtered to a specific stall_type ("main" or "sub").
    """
    from inventory.models import Stall
    from sales.models import SalesTransaction

    stall_ids = list(
        Stall.objects.filter(stall_type=stall_type, is_deleted=False)
        .values_list("id", flat=True)
    )

    txns = (
        SalesTransaction.objects
        .annotate(effective_date=Coalesce("transaction_date", TruncDate("created_at")))
        .filter(
            is_deleted=False, voided=False,
            effective_date__gte=start,
            effective_date__lte=end,
            stall_id__in=stall_ids,
        )
        .prefetch_related("items__item")
        .order_by("effective_date", "created_at")
    )

    # Group transactions by date
    daily_txns = defaultdict(list)
    for t in txns:
        day = _effective_date(t)
        daily_txns[day].append(t)

    wb = Workbook()
    # Remove default sheet — we'll create per-day sheets
    wb.remove(wb.active)

    sorted_days = sorted(daily_txns.keys())

    if not sorted_days:
        # Create a single empty sheet so the workbook is valid
        ws = wb.create_sheet("No Data")
        ws.cell(row=1, column=1, value="No sales found for the selected period.")
        return wb

    for day in sorted_days:
        sheet_title = day.strftime("%b %d, %Y")  # e.g. "Mar 23, 2026"
        # Excel sheet names max 31 chars — this format is safe
        ws = wb.create_sheet(sheet_title)

        headers = ["Qty", "Item", "Amount"]
        rows_data = []
        day_total = Decimal("0")

        for t in daily_txns[day]:
            for si in t.items.all():
                qty = float(si.quantity)
                description = si.description or (si.item.name if si.item else "—")
                amount = float(si.line_total)
                day_total += si.line_total
                rows_data.append([qty, description, amount])

        last_row = _write_rows(ws, headers, rows_data)

        # Add a total row
        total_row = last_row + 1
        ws.cell(row=total_row, column=1, value="").border = THIN_BORDER
        ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=2).border = THIN_BORDER
        ws.cell(row=total_row, column=3, value=float(day_total)).font = Font(bold=True)
        ws.cell(row=total_row, column=3).border = THIN_BORDER
        ws.cell(row=total_row, column=3).number_format = "#,##0.00"

        # Format Amount column as numbers
        for row_idx in range(2, total_row):
            ws.cell(row=row_idx, column=3).number_format = "#,##0.00"

    return wb


# ═══════════════════════════════════════════════════════════════════
#  CHEQUE WORKBOOK BUILDER
# ═══════════════════════════════════════════════════════════════════

CHEQUE_VALID_SHEETS = {"all_cheques", "by_status", "by_bank", "monthly_summary"}

STATUS_FILLS = {
    "pending": AMBER_FILL,
    "deposited": BLUE_FILL,
    "encashed": GREEN_FILL,
    "bounced": RED_FILL,
    "returned": RED_FILL,
    "cancelled": RED_FILL,
}


def build_cheque_workbook(start, end, requested):
    """Build and return a cheque collections report Workbook."""
    from receivables.models import ChequeCollection

    cheques = (
        ChequeCollection.objects.filter(
            date_collected__date__gte=start,
            date_collected__date__lte=end,
        )
        .select_related("client", "collected_by", "sales_transaction")
        .order_by("date_collected")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="RVDC Cheque Collections Report").font = Font(bold=True, size=16, color="2563EB")
    ws.cell(row=2, column=1, value=f"Period: {start.strftime('%b %d, %Y')} — {end.strftime('%b %d, %Y')}")
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666")
    ws.cell(row=3, column=1, value=f"Generated: {timezone.now().strftime('%B %d, %Y %I:%M %p')}")
    ws.cell(row=3, column=1).font = Font(italic=True, color="666666")

    total_count = cheques.count()
    total_billing = sum(float(c.billing_amount) for c in cheques)
    total_cheque = sum(float(c.cheque_amount) for c in cheques)
    pending = cheques.filter(status="pending").count()
    deposited = cheques.filter(status="deposited").count()
    encashed = cheques.filter(status="encashed").count()
    bounced = cheques.filter(status="bounced").count()

    stats = [
        ("Total Cheques", total_count),
        ("Total Billing Amount", f"₱{total_billing:,.2f}"),
        ("Total Cheque Amount", f"₱{total_cheque:,.2f}"),
        ("", ""),
        ("Pending", pending), ("Deposited", deposited),
        ("Encashed", encashed), ("Bounced", bounced),
    ]
    row = 5
    ws.cell(row=row, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=2).border = THIN_BORDER
    for label, value in stats:
        row += 1
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=value).border = THIN_BORDER
        if label == "Bounced" and bounced > 0:
            ws.cell(row=row, column=2).font = Font(bold=True, color="DC2626")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25

    if "all_cheques" in requested:
        ws_c = wb.create_sheet("All Cheques")
        headers = [
            "ID", "Date Collected", "Client", "Issued By",
            "Cheque #", "Cheque Date", "Bank", "Deposit Bank",
            "Billing Amount", "Cheque Amount", "OR #",
            "Status", "Collection Type", "Collected By", "Notes",
        ]
        rows_data = []
        for c in cheques:
            rows_data.append([
                c.id, c.date_collected.strftime("%Y-%m-%d"),
                c.client.full_name if c.client else "—",
                c.issued_by or "—", c.cheque_number,
                c.cheque_date.strftime("%Y-%m-%d") if c.cheque_date else "—",
                c.get_bank_name_display() if c.bank_name else "—",
                c.get_deposit_bank_display() if c.deposit_bank else "—",
                float(c.billing_amount), float(c.cheque_amount),
                c.or_number or "—", c.get_status_display(),
                c.get_collection_type_display(),
                c.collected_by.get_full_name() if c.collected_by else "—",
                c.notes or "",
            ])
        last = _write_rows(ws_c, headers, rows_data)
        stat_col = headers.index("Status") + 1
        for ri in range(2, last + 1):
            cell = ws_c.cell(row=ri, column=stat_col)
            fill = STATUS_FILLS.get(cell.value.lower() if cell.value else "", None)
            if fill:
                for ci in range(1, len(headers) + 1):
                    ws_c.cell(row=ri, column=ci).fill = fill

    if "by_status" in requested:
        ws_s = wb.create_sheet("By Status")
        status_agg = defaultdict(lambda: {"count": 0, "billing": Decimal("0"), "cheque": Decimal("0")})
        for c in cheques:
            key = c.get_status_display()
            status_agg[key]["count"] += 1
            status_agg[key]["billing"] += c.billing_amount
            status_agg[key]["cheque"] += c.cheque_amount
        headers = ["Status", "Count", "Total Billing", "Total Cheque Amount"]
        rows_data = [[s, d["count"], float(d["billing"]), float(d["cheque"])] for s, d in sorted(status_agg.items())]
        _write_rows(ws_s, headers, rows_data)

    if "by_bank" in requested:
        ws_b = wb.create_sheet("By Bank")
        bank_agg = defaultdict(lambda: {"count": 0, "amount": Decimal("0")})
        for c in cheques:
            key = c.get_bank_name_display() if c.bank_name else "Unknown"
            bank_agg[key]["count"] += 1
            bank_agg[key]["amount"] += c.cheque_amount
        headers = ["Bank", "Cheques", "Total Amount"]
        rows_data = [[bank, d["count"], float(d["amount"])] for bank, d in sorted(bank_agg.items())]
        _write_rows(ws_b, headers, rows_data)

    if "monthly_summary" in requested:
        ws_m = wb.create_sheet("Monthly Summary")
        monthly_map = defaultdict(lambda: {"count": 0, "billing": Decimal("0"), "cheque": Decimal("0")})
        for c in cheques:
            key = c.date_collected.strftime("%Y-%m")
            monthly_map[key]["count"] += 1
            monthly_map[key]["billing"] += c.billing_amount
            monthly_map[key]["cheque"] += c.cheque_amount
        headers = ["Month", "Cheques", "Total Billing", "Total Cheque Amount"]
        rows_data = [[m, d["count"], float(d["billing"]), float(d["cheque"])] for m, d in sorted(monthly_map.items())]
        _write_rows(ws_m, headers, rows_data)

    return wb


# ═══════════════════════════════════════════════════════════════════
#  INVENTORY WORKBOOK BUILDER (delegates to inventory app)
# ═══════════════════════════════════════════════════════════════════

def build_inventory_workbook(requested):
    """Build inventory report workbook using existing builder functions."""
    from inventory.api.export_views import InventoryExportView, _build_summary_sheet
    wb = Workbook()
    _build_summary_sheet(wb)
    for sheet_key in requested:
        if sheet_key in InventoryExportView.SHEET_BUILDERS:
            InventoryExportView.SHEET_BUILDERS[sheet_key](wb)
    return wb


# ═══════════════════════════════════════════════════════════════════
#  BIR 2307 WORKBOOK BUILDER
# ═══════════════════════════════════════════════════════════════════

BIR_2307_VALID_SHEETS = {"main_stall", "sub_stall"}

ORANGE_FILL = PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid")

# Default tax rates
DEFAULT_EWT_RATE = Decimal("0.02")   # 2% EWT for services/contractors (Main Stall)
DEFAULT_TAX_RATE = Decimal("0.03")   # 3% Percentage Tax for sales (Sub Stall, non-VAT)


def _extract_numeric_receipt(receipt_str):
    """
    Extract a numeric value from a receipt string for gap detection.
    Returns int if the string is purely numeric (or numeric with leading zeros),
    otherwise returns None.
    """
    import re
    cleaned = re.sub(r'[^0-9]', '', str(receipt_str).strip())
    if cleaned:
        return int(cleaned)
    return None


def _fill_receipt_gaps(rows):
    """
    Given a list of {receipt_number, receipt_book, total_amount, date, source} dicts,
    detect gaps in the numeric receipt sequence *per receipt book* and insert
    placeholder rows with total_amount=0 and source="MISSING" for each skipped number.
    The receipt numbers are zero-padded to match the longest existing receipt.
    """
    if not rows:
        return rows

    # Group rows by receipt_book
    from collections import defaultdict as _dd
    book_groups = _dd(list)
    for r in rows:
        book_groups[r.get("receipt_book") or ""].append(r)

    filled = []
    for book_key in sorted(book_groups.keys(), key=lambda x: (x == "", x)):
        book_rows = book_groups[book_key]
        # Build map of numeric receipt -> row
        numeric_map = {}
        max_len = 0
        for r in book_rows:
            num = _extract_numeric_receipt(r["receipt_number"])
            if num is not None:
                numeric_map[num] = r
                max_len = max(max_len, len(str(r["receipt_number"]).strip()))

        if len(numeric_map) < 2:
            filled.extend(book_rows)
            continue

        min_num = min(numeric_map.keys())
        max_num = max(numeric_map.keys())

        for n in range(min_num, max_num + 1):
            if n in numeric_map:
                filled.append(numeric_map[n])
            else:
                padded = str(n).zfill(max_len) if max_len > 0 else str(n)
                filled.append({
                    "date": None,
                    "receipt_number": padded,
                    "receipt_book": book_key or None,
                    "total_amount": Decimal("0"),
                    "source": "MISSING",
                    "with_2307": False,
                })

    return filled


def _get_bir_2307_main_stall_rows(start, end):
    """Services with manual_receipt_number (main stall OR)."""
    from services.models import Service

    services = Service.objects.filter(
        manual_receipt_number__isnull=False,
        is_deleted=False,
        created_at__date__gte=start,
        created_at__date__lte=end,
    ).exclude(manual_receipt_number="").order_by("manual_receipt_number")

    rows = []
    for svc in services:
        rows.append({
            "date": svc.created_at.date(),
            "receipt_number": svc.manual_receipt_number,
            "receipt_book": getattr(svc, 'receipt_book', None) or None,
            "total_amount": svc.total_revenue or Decimal("0"),
            "source": "Service",
            "with_2307": getattr(svc, 'with_2307', False),
        })

    return _fill_receipt_gaps(rows)


def _get_bir_2307_sub_stall_rows(start, end):
    """Sub stall direct sales with manual_receipt_number (not linked to any service)."""
    from inventory.models import Stall
    from sales.models import SalesTransaction
    from services.models import Service

    service_txn_ids = set(
        Service.objects.filter(
            related_transaction__isnull=False,
        ).values_list("related_transaction_id", flat=True)
    ) | set(
        Service.objects.filter(
            related_sub_transaction__isnull=False,
        ).values_list("related_sub_transaction_id", flat=True)
    )

    sub_stall_ids = list(
        Stall.objects.filter(stall_type="sub", is_deleted=False)
        .values_list("id", flat=True)
    )

    direct_sales = (
        SalesTransaction.objects
        .annotate(effective_date=Coalesce("transaction_date", TruncDate("created_at")))
        .filter(
            manual_receipt_number__isnull=False,
            is_deleted=False,
            voided=False,
            stall_id__in=sub_stall_ids,
            effective_date__gte=start,
            effective_date__lte=end,
        )
        .exclude(manual_receipt_number="")
        .exclude(id__in=service_txn_ids)
        .prefetch_related("items")
        .order_by("manual_receipt_number")
    )

    rows = []
    for txn in direct_sales:
        rows.append({
            "date": _effective_date(txn),
            "receipt_number": txn.manual_receipt_number,
            "receipt_book": getattr(txn, 'receipt_book', None) or None,
            "total_amount": txn.computed_total,
            "source": "Direct Sale",
        })

    return _fill_receipt_gaps(rows)


def _write_bir_sheet(wb, title, rows, period_key_fn, period_label, include_2307_column=False, tax_rate=None, tax_label=None):
    """
    Write a BIR 2307 sheet with Receipt # | Book # | Total Amount, grouped by period.
    Inserts subtotals per period. Highlights MISSING receipt rows in orange.
    If include_2307_column is True (main stall), adds 'With 2307' and tax columns,
    applying tax only to 2307-flagged rows.
    If tax_rate is provided without include_2307_column (sub stall), adds a tax column
    applied to all non-MISSING rows.
    """
    ws = wb.create_sheet(title)
    headers = [period_label, "Official Receipt #", "Book #", "Total Amount"]
    has_tax_col = False
    if include_2307_column:
        rate = tax_rate or DEFAULT_EWT_RATE
        rate_pct = f"{float(rate * 100):g}%"
        headers.append("With 2307")
        headers.append(tax_label or f"Tax Withheld ({rate_pct})")
        has_tax_col = True
    elif tax_rate:
        rate = tax_rate
        rate_pct = f"{float(rate * 100):g}%"
        headers.append(tax_label or f"Tax ({rate_pct})")
        has_tax_col = True
    else:
        rate = Decimal("0")

    amount_col = 4  # column index for Total Amount
    rows_data = []
    missing_row_indices = []  # track which data rows are MISSING
    with_2307_row_indices = []  # track which data rows have with_2307

    # Group by period
    period_map = defaultdict(list)
    for r in rows:
        key = period_key_fn(r)
        period_map[key].append(r)

    data_row_idx = 0  # 0-based index in rows_data
    for period in sorted(period_map.keys()):
        period_rows = period_map[period]
        for pr in period_rows:
            amount = float(pr["total_amount"])
            row_data = [
                period,
                pr["receipt_number"],
                pr.get("receipt_book") or "",
                amount,
            ]
            is_missing = pr["source"] == "MISSING"
            if include_2307_column:
                has_2307 = pr.get("with_2307", False)
                row_data.append("Yes" if has_2307 else "No")
                tax = float(Decimal(str(amount)) * rate) if has_2307 and not is_missing else 0.0
                row_data.append(tax)
                if has_2307 and not is_missing:
                    with_2307_row_indices.append(data_row_idx)
            elif has_tax_col:
                tax = float(Decimal(str(amount)) * rate) if not is_missing else 0.0
                row_data.append(tax)
            rows_data.append(row_data)
            if is_missing:
                missing_row_indices.append(data_row_idx)
            data_row_idx += 1
        # Subtotal
        period_total = sum(float(pr["total_amount"]) for pr in period_rows)
        subtotal_row = ["", "", "SUBTOTAL", period_total]
        if include_2307_column:
            subtotal_row.append("")
            period_tax = sum(
                float(Decimal(str(pr["total_amount"])) * rate)
                for pr in period_rows
                if pr.get("with_2307") and pr["source"] != "MISSING"
            )
            subtotal_row.append(period_tax)
        elif has_tax_col:
            period_tax = sum(
                float(Decimal(str(pr["total_amount"])) * rate)
                for pr in period_rows
                if pr["source"] != "MISSING"
            )
            subtotal_row.append(period_tax)
        rows_data.append(subtotal_row)
        data_row_idx += 1

    col_count = len(headers)
    last = _write_rows(ws, headers, rows_data)

    # Format and highlight
    tax_col_idx = col_count if has_tax_col else None
    for ri in range(2, last + 1):
        ws.cell(row=ri, column=amount_col).number_format = "#,##0.00"
        if tax_col_idx:
            ws.cell(row=ri, column=tax_col_idx).number_format = "#,##0.00"
        # Bold subtotal rows
        if ws.cell(row=ri, column=3).value == "SUBTOTAL":
            for ci in range(1, col_count + 1):
                ws.cell(row=ri, column=ci).font = Font(bold=True)
                ws.cell(row=ri, column=ci).fill = BLUE_FILL

    # Highlight MISSING rows in orange
    for idx in missing_row_indices:
        excel_row = idx + 2
        for ci in range(1, col_count + 1):
            ws.cell(row=excel_row, column=ci).fill = ORANGE_FILL
            ws.cell(row=excel_row, column=ci).font = Font(italic=True, color="9A3412")

    # Highlight With 2307 rows in light blue
    for idx in with_2307_row_indices:
        excel_row = idx + 2
        for ci in range(1, col_count + 1):
            cell = ws.cell(row=excel_row, column=ci)
            # Don't override MISSING highlight
            if idx not in missing_row_indices:
                cell.fill = BLUE_FILL


def build_bir_2307_workbook(start, end, requested_sheets, stall_type, ewt_rate=None, tax_rate=None, pre_system_amount=None, pre_system_receipts=None):
    """
    Build BIR 2307 report for a specific stall type.
    stall_type: "main" or "sub"
    ewt_rate: custom EWT rate for main stall (default 2%)
    tax_rate: custom tax rate for sub stall (default 3% percentage tax)
    pre_system_amount: total revenue from sales not recorded in the system
    pre_system_receipts: receipt number range string (e.g. "0001-0045")
    """
    if stall_type == "main":
        all_rows = _get_bir_2307_main_stall_rows(start, end)
        stall_label = "Main Stall (Services)"
        rate = ewt_rate or DEFAULT_EWT_RATE
    else:
        all_rows = _get_bir_2307_sub_stall_rows(start, end)
        stall_label = "Sub Stall (Direct Sales)"
        rate = tax_rate or DEFAULT_TAX_RATE

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value=f"RVDC BIR 2307 — {stall_label}").font = Font(
        bold=True, size=16, color="2563EB"
    )
    ws.cell(
        row=2, column=1,
        value=f"Period: {start.strftime('%b %d, %Y')} — {end.strftime('%b %d, %Y')}",
    )
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666")
    ws.cell(
        row=3, column=1,
        value=f"Generated: {timezone.now().strftime('%B %d, %Y %I:%M %p')}",
    )
    ws.cell(row=3, column=1).font = Font(italic=True, color="666666")

    actual_rows = [r for r in all_rows if r["source"] != "MISSING"]
    missing_count = sum(1 for r in all_rows if r["source"] == "MISSING")
    total_amount = sum(float(r["total_amount"]) for r in actual_rows)
    is_main = stall_type == "main"
    rate_pct = f"{float(rate * 100):g}%"

    stats = [
        ("Stall", stall_label),
        ("Total Receipted Transactions", len(actual_rows)),
        ("Total Amount", f"₱{total_amount:,.2f}"),
        ("Missing Receipt Numbers", missing_count),
    ]
    # Pre-system adjustment
    pre_amount = float(pre_system_amount) if pre_system_amount else 0
    combined_amount = total_amount + pre_amount

    if is_main:
        with_2307_count = sum(1 for r in actual_rows if r.get("with_2307"))
        without_2307_count = len(actual_rows) - with_2307_count
        with_2307_amount = sum(float(r["total_amount"]) for r in actual_rows if r.get("with_2307"))
        total_tax_withheld = float(Decimal(str(with_2307_amount)) * rate)
        stats.extend([
            ("With 2307", f"{with_2307_count} (₱{with_2307_amount:,.2f})"),
            ("Without 2307", without_2307_count),
            ("Tax Rate (EWT)", rate_pct),
            (f"Estimated Tax Withheld ({rate_pct})", f"₱{total_tax_withheld:,.2f}"),
        ])
    else:
        total_tax = float(Decimal(str(total_amount)) * rate)
        stats.extend([
            ("Tax Rate", rate_pct),
            (f"Estimated Tax ({rate_pct})", f"₱{total_tax:,.2f}"),
        ])

    if pre_amount > 0:
        pre_tax = float(Decimal(str(pre_amount)) * rate)
        combined_tax = float(Decimal(str(combined_amount)) * rate)
        stats.append(("", ""))  # blank separator row
        stats.append(("── Pre-System Data ──", ""))
        stats.append(("Pre-System Amount", f"₱{pre_amount:,.2f}"))
        if pre_system_receipts:
            stats.append(("Pre-System Receipts", pre_system_receipts))
        stats.append((f"Pre-System Tax ({rate_pct})", f"₱{pre_tax:,.2f}"))
        stats.append(("", ""))  # blank separator row
        stats.append(("── Combined Totals ──", ""))
        stats.append(("Combined Amount (System + Pre-System)", f"₱{combined_amount:,.2f}"))
        stats.append((f"Combined Tax ({rate_pct})", f"₱{combined_tax:,.2f}"))
    row = 5
    ws.cell(row=row, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=2).border = THIN_BORDER
    for label, value in stats:
        row += 1
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=value).border = THIN_BORDER
        if label == "Missing Receipt Numbers" and missing_count > 0:
            ws.cell(row=row, column=2).font = Font(bold=True, color="C2410C")
            ws.cell(row=row, column=2).fill = ORANGE_FILL
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25

    # Legend
    row += 2
    ws.cell(row=row, column=1, value="Legend:").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Orange rows = missing receipt numbers (gap in sequence)")
    ws.cell(row=row, column=1).font = Font(italic=True, color="9A3412")
    ws.cell(row=row, column=1).fill = ORANGE_FILL
    if is_main:
        row += 1
        ws.cell(row=row, column=1, value="Blue rows = transactions with BIR Form 2307")
        ws.cell(row=row, column=1).font = Font(italic=True, color="1E40AF")
        ws.cell(row=row, column=1).fill = BLUE_FILL

    # Helper for period keys (MISSING rows have date=None, use receipt_number sorting only)
    def _daily_key(r):
        if r["date"]:
            return r["date"].strftime("%Y-%m-%d")
        return "Unknown Date"

    def _monthly_key(r):
        if r["date"]:
            return r["date"].strftime("%Y-%m")
        return "Unknown Month"

    def _quarterly_key(r):
        if r["date"]:
            q = (r["date"].month - 1) // 3 + 1
            return f"{r['date'].year} Q{q}"
        return "Unknown Quarter"

    def _yearly_key(r):
        if r["date"]:
            return str(r["date"].year)
        return "Unknown Year"

    if is_main:
        tax_label = f"Tax Withheld ({rate_pct})"
    else:
        tax_label = f"Tax ({rate_pct})"

    sheet_kwargs = dict(
        include_2307_column=is_main,
        tax_rate=rate if not is_main else None,
        tax_label=tax_label,
    )
    if is_main:
        sheet_kwargs["tax_rate"] = rate

    _write_bir_sheet(wb, "Daily Summary", all_rows, _daily_key, "Date", **sheet_kwargs)
    _write_bir_sheet(wb, "Monthly Summary", all_rows, _monthly_key, "Month", **sheet_kwargs)
    _write_bir_sheet(wb, "Quarterly Summary", all_rows, _quarterly_key, "Quarter", **sheet_kwargs)
    _write_bir_sheet(wb, "Yearly Summary", all_rows, _yearly_key, "Year", **sheet_kwargs)

    return wb


# ═══════════════════════════════════════════════════════════════════
#  SYNCHRONOUS EXPORT VIEWS (kept for backward compatibility)
# ═══════════════════════════════════════════════════════════════════

class SalesExportView(APIView):
    """GET /api/sales/export-report/ — synchronous blob download."""
    permission_classes = [IsAdminUser]

    def get(self, request):
        today = timezone.now().date()
        start = _parse_date(request.query_params.get("start_date"), today - timedelta(days=30))
        end = _parse_date(request.query_params.get("end_date"), today)
        sheets_param = request.query_params.get("sheets", "")
        if sheets_param:
            requested = [s.strip() for s in sheets_param.split(",") if s.strip()]
            invalid = [s for s in requested if s not in SALES_VALID_SHEETS]
            if invalid:
                return Response(
                    {"detail": f"Invalid sheet(s): {', '.join(invalid)}"},
                    status=drf_status.HTTP_400_BAD_REQUEST,
                )
        else:
            requested = list(SALES_VALID_SHEETS)
        wb = build_sales_workbook(start, end, requested)
        return _workbook_response(wb, "sales_report")


class ChequeExportView(APIView):
    """GET /api/receivables/export-report/ — synchronous blob download."""
    permission_classes = [IsAdminUser]

    def get(self, request):
        today = timezone.now().date()
        start = _parse_date(request.query_params.get("start_date"), today - timedelta(days=90))
        end = _parse_date(request.query_params.get("end_date"), today)
        sheets_param = request.query_params.get("sheets", "")
        if sheets_param:
            requested = [s.strip() for s in sheets_param.split(",") if s.strip()]
            invalid = [s for s in requested if s not in CHEQUE_VALID_SHEETS]
            if invalid:
                return Response(
                    {"detail": f"Invalid sheet(s): {', '.join(invalid)}"},
                    status=drf_status.HTTP_400_BAD_REQUEST,
                )
        else:
            requested = list(CHEQUE_VALID_SHEETS)
        wb = build_cheque_workbook(start, end, requested)
        return _workbook_response(wb, "cheque_report")


# ═══════════════════════════════════════════════════════════════════
#  BACKGROUND EXPORT (Thread + WebSocket)
# ═══════════════════════════════════════════════════════════════════

EXPORT_TYPES = {
    "inventory": {
        "label": "Inventory Report",
        "valid_sheets": {"no_stock", "low_stock", "most_bought", "least_bought",
                         "custom_items", "duplicates", "by_category"},
        "prefix": "inventory_report",
    },
    "sales": {
        "label": "Sales Report",
        "valid_sheets": SALES_VALID_SHEETS,
        "prefix": "sales_report",
    },
    "cheques": {
        "label": "Cheque Collections Report",
        "valid_sheets": CHEQUE_VALID_SHEETS,
        "prefix": "cheque_report",
    },
    "daily_sales": {
        "label": "Daily Sales Report",
        "valid_sheets": {"main_stall", "sub_stall"},
        "prefix": "daily_sales",
    },
    "bir_2307": {
        "label": "BIR 2307 Sales Report",
        "valid_sheets": BIR_2307_VALID_SHEETS,
        "prefix": "bir_2307_report",
    },
}


class BackgroundExportView(APIView):
    """
    POST /api/analytics/export/
    Start a background export. Notifies via WebSocket when file is ready.

    Body:
        export_type  — "inventory" | "sales" | "cheques"
        sheets       — comma-separated sheet keys (default: all)
        start_date   — YYYY-MM-DD (sales/cheques only)
        end_date     — YYYY-MM-DD (sales/cheques only)
    """
    permission_classes = [IsAdminUser]

    def post(self, request):
        export_type = request.data.get("export_type", "")
        if export_type not in EXPORT_TYPES:
            return Response(
                {"error": f"Invalid export_type. Options: {list(EXPORT_TYPES.keys())}"},
                status=drf_status.HTTP_400_BAD_REQUEST,
            )

        config = EXPORT_TYPES[export_type]
        sheets_param = request.data.get("sheets", "")
        if sheets_param:
            requested = [s.strip() for s in sheets_param.split(",") if s.strip()]
            invalid = [s for s in requested if s not in config["valid_sheets"]]
            if invalid:
                return Response(
                    {"error": f"Invalid sheet(s): {', '.join(invalid)}"},
                    status=drf_status.HTTP_400_BAD_REQUEST,
                )
        else:
            requested = list(config["valid_sheets"])

        today = timezone.now().date()
        start_date = _parse_date(request.data.get("start_date"), today - timedelta(days=30))
        end_date = _parse_date(request.data.get("end_date"), today)
        user_id = request.user.id
        user_name = request.user.get_full_name()

        # Parse optional tax rates and pre-system data for BIR 2307 exports
        ewt_rate = None
        tax_rate = None
        pre_system_amount = None
        pre_system_receipts = None
        if export_type == "bir_2307":
            try:
                raw_ewt = request.data.get("ewt_rate")
                if raw_ewt is not None:
                    ewt_rate = Decimal(str(raw_ewt)) / Decimal("100")
            except Exception:
                pass
            try:
                raw_tax = request.data.get("tax_rate")
                if raw_tax is not None:
                    tax_rate = Decimal(str(raw_tax)) / Decimal("100")
            except Exception:
                pass
            try:
                raw_pre = request.data.get("pre_system_amount")
                if raw_pre is not None and str(raw_pre).strip():
                    pre_system_amount = Decimal(str(raw_pre))
            except Exception:
                pass
            pre_system_receipts = request.data.get("pre_system_receipts", "") or None

        def _generate():
            try:
                if export_type == "inventory":
                    wb = build_inventory_workbook(requested)
                    token, filename = _save_workbook(wb, config["prefix"])
                    logger.info("Export %s generated by %s: %s", export_type, user_name, filename)
                    _notify_export_ready(user_id, export_type, config["label"], token, filename)
                elif export_type == "sales":
                    wb = build_sales_workbook(start_date, end_date, requested)
                    token, filename = _save_workbook(wb, config["prefix"])
                    logger.info("Export %s generated by %s: %s", export_type, user_name, filename)
                    _notify_export_ready(user_id, export_type, config["label"], token, filename)
                elif export_type == "cheques":
                    wb = build_cheque_workbook(start_date, end_date, requested)
                    token, filename = _save_workbook(wb, config["prefix"])
                    logger.info("Export %s generated by %s: %s", export_type, user_name, filename)
                    _notify_export_ready(user_id, export_type, config["label"], token, filename)
                elif export_type == "daily_sales":
                    # Generate separate file per stall type
                    stall_map = {"main_stall": "main", "sub_stall": "sub"}
                    generated = []
                    for sheet_key in requested:
                        stall_type = stall_map.get(sheet_key)
                        if not stall_type:
                            continue
                        wb = build_daily_sales_workbook(start_date, end_date, stall_type)
                        prefix = f"daily_sales_{stall_type}_stall"
                        token, filename = _save_workbook(wb, prefix)
                        generated.append((token, filename))
                        logger.info("Export %s (%s) generated by %s: %s", export_type, sheet_key, user_name, filename)

                    if generated:
                        # Notify for each file
                        for token, filename in generated:
                            _notify_export_ready(user_id, export_type, config["label"], token, filename)
                    else:
                        _notify_export_failed(user_id, export_type, config["label"])
                elif export_type == "bir_2307":
                    # Generate separate file per stall type (like daily_sales)
                    stall_map = {"main_stall": "main", "sub_stall": "sub"}
                    generated = []
                    for sheet_key in requested:
                        stall_type = stall_map.get(sheet_key)
                        if not stall_type:
                            continue
                        wb = build_bir_2307_workbook(
                            start_date, end_date, requested, stall_type,
                            ewt_rate=ewt_rate, tax_rate=tax_rate,
                            pre_system_amount=pre_system_amount,
                            pre_system_receipts=pre_system_receipts,
                        )
                        prefix = f"bir_2307_{stall_type}_stall"
                        token, filename = _save_workbook(wb, prefix)
                        generated.append((token, filename))
                        logger.info("Export %s (%s) generated by %s: %s", export_type, sheet_key, user_name, filename)

                    if generated:
                        for token, filename in generated:
                            _notify_export_ready(user_id, export_type, config["label"], token, filename)
                    else:
                        _notify_export_failed(user_id, export_type, config["label"])
                else:
                    return
            except Exception:
                logger.exception("Background export %s failed for %s", export_type, user_name)
                _notify_export_failed(user_id, export_type, config["label"])

        thread = threading.Thread(target=_generate, daemon=True)
        thread.start()

        return Response({
            "accepted": True,
            "export_type": export_type,
            "message": f"{config['label']} is being generated. You'll be notified when it's ready to download.",
        }, status=drf_status.HTTP_202_ACCEPTED)


def _notify_export_ready(user_id, export_type, label, token, filename):
    """Push export_ready event via WebSocket."""
    try:
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer

        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f"notifications_{user_id}",
                {
                    "type": "send_notification",
                    "data": {
                        "event": "export_ready",
                        "export_type": export_type,
                        "title": f"{label} Ready",
                        "message": f"Your {label.lower()} is ready to download.",
                        "token": token,
                        "filename": filename,
                    },
                },
            )
    except Exception:
        logger.exception("Failed to send export_ready via WebSocket")


def _notify_export_failed(user_id, export_type, label):
    """Push export_failed event via WebSocket."""
    try:
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer

        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f"notifications_{user_id}",
                {
                    "type": "send_notification",
                    "data": {
                        "event": "export_failed",
                        "export_type": export_type,
                        "title": f"{label} Failed",
                        "message": f"Failed to generate {label.lower()}. Please try again.",
                    },
                },
            )
    except Exception:
        logger.exception("Failed to send export_failed via WebSocket")


class ExportDownloadView(APIView):
    """
    GET /api/analytics/export/download/<token>/
    Serve a previously generated export file and delete it after serving.
    """
    permission_classes = [IsAdminUser]

    def get(self, request, token):
        # Validate token format (hex UUID)
        try:
            uuid.UUID(token)
        except ValueError:
            return Response({"error": "Invalid token."}, status=drf_status.HTTP_400_BAD_REQUEST)

        _ensure_export_dir()
        # Find the file matching this token
        for fname in os.listdir(EXPORT_DIR):
            if fname.startswith(f"{token}_"):
                fpath = os.path.join(EXPORT_DIR, fname)
                # Extract original filename (after token_)
                original_name = fname[len(token) + 1:]
                response = FileResponse(
                    open(fpath, "rb"),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                response["Content-Disposition"] = f'attachment; filename="{original_name}"'
                # Schedule cleanup after response is sent
                def _cleanup(f=fpath):
                    try:
                        os.remove(f)
                    except OSError:
                        pass
                response.close = lambda orig=response.close, cb=_cleanup: (orig(), cb())
                return response

        return Response({"error": "Export not found or expired."}, status=drf_status.HTTP_404_NOT_FOUND)
