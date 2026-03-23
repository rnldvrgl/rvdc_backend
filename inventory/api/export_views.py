import re
from collections import defaultdict
from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.db.models import F, Sum
from django.db.models.functions import Coalesce, TruncDate, TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from inventory.models import Item, Stock, StockRoomStock, ProductCategory
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework.permissions import IsAdminUser
from rest_framework.views import APIView


# ── Styling constants ──────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
NO_STOCK_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
LOW_STOCK_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
DUPLICATE_FILL = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, col_count):
    """Apply header styling to the first row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def _write_rows(ws, headers, rows, start_row=1):
    """Write headers + data rows with borders."""
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col_idx, value=header)
    _style_header_row(ws, len(headers))

    for row_idx, row_data in enumerate(rows, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    _auto_width(ws)
    return start_row + len(rows)


# ── Helper queries ─────────────────────────────────────────────────

def _get_stock_items(status_filter):
    """Get stock items filtered by status (no_stock or low_stock)."""
    stocks = Stock.objects.filter(
        is_deleted=False,
        item__is_deleted=False,
    ).select_related("item__category", "stall").annotate(
        avail=F("quantity") - F("reserved_quantity"),
    )

    if status_filter == "no_stock":
        stocks = stocks.filter(avail__lte=0)
    elif status_filter == "low_stock":
        stocks = stocks.filter(avail__gt=0, avail__lte=F("low_stock_threshold"))

    return stocks.order_by("item__name")


def _get_stockroom_items(status_filter):
    """Get stockroom items filtered by status."""
    qs = StockRoomStock.objects.filter(
        is_deleted=False,
        item__is_deleted=False,
    ).select_related("item__category")

    if status_filter == "no_stock":
        qs = qs.filter(quantity=0)
    elif status_filter == "low_stock":
        qs = qs.filter(quantity__gt=0, quantity__lte=F("low_stock_threshold"))

    return qs.order_by("item__name")


def _get_most_bought_items(limit=50):
    """
    Aggregate total quantity consumed across sales, appliance services,
    and service-level items. Returns list of dicts.
    """
    from sales.models import SalesItem
    from services.models import ApplianceItemUsed, ServiceItemUsed

    # Sales items (exclude voided transactions)
    sales_agg = (
        SalesItem.objects.filter(
            item__isnull=False,
            item__is_deleted=False,
            transaction__voided=False,
            transaction__is_deleted=False,
        )
        .values("item_id")
        .annotate(total_qty=Coalesce(Sum("quantity"), Decimal("0")))
    )

    # Appliance items used (exclude cancelled)
    appliance_agg = (
        ApplianceItemUsed.objects.filter(
            item__isnull=False,
            item__is_deleted=False,
            is_cancelled=False,
        )
        .values("item_id")
        .annotate(total_qty=Coalesce(Sum("quantity"), Decimal("0")))
    )

    # Service-level items used
    service_agg = (
        ServiceItemUsed.objects.filter(
            item__isnull=False,
            item__is_deleted=False,
        )
        .values("item_id")
        .annotate(total_qty=Coalesce(Sum("quantity"), Decimal("0")))
    )

    # Merge all aggregations
    combined = defaultdict(Decimal)
    source_breakdown = defaultdict(lambda: {"sales": Decimal("0"), "service": Decimal("0")})

    for row in sales_agg:
        combined[row["item_id"]] += row["total_qty"]
        source_breakdown[row["item_id"]]["sales"] += row["total_qty"]

    for row in appliance_agg:
        combined[row["item_id"]] += row["total_qty"]
        source_breakdown[row["item_id"]]["service"] += row["total_qty"]

    for row in service_agg:
        combined[row["item_id"]] += row["total_qty"]
        source_breakdown[row["item_id"]]["service"] += row["total_qty"]

    # Sort by total and apply limit
    sorted_items = sorted(combined.items(), key=lambda x: x[1], reverse=True)[:limit]

    # Fetch item details
    item_ids = [item_id for item_id, _ in sorted_items]
    items_map = {
        item.id: item
        for item in Item.objects.filter(id__in=item_ids).select_related("category")
    }

    results = []
    for item_id, total_qty in sorted_items:
        item = items_map.get(item_id)
        if not item:
            continue
        results.append({
            "item": item,
            "total_qty": total_qty,
            "sales_qty": source_breakdown[item_id]["sales"],
            "service_qty": source_breakdown[item_id]["service"],
        })

    return results


def _get_least_bought_items(limit=50):
    """Items with the lowest quantity consumed (but at least one use)."""
    results = _get_most_bought_items(limit=9999)
    # Reverse to get least bought, then take top N
    results.reverse()
    return results[:limit]


def _get_custom_items_usage():
    """Aggregate custom items (non-inventory) used across services."""
    from services.models import ApplianceItemUsed, ServiceItemUsed

    combined = defaultdict(lambda: {"qty": Decimal("0"), "price": Decimal("0")})

    for model in [ApplianceItemUsed, ServiceItemUsed]:
        qs = model.objects.filter(
            item__isnull=True,
            custom_description__gt="",
        )
        if hasattr(model, "is_cancelled"):
            qs = qs.filter(is_cancelled=False)

        for obj in qs.values("custom_description", "custom_price").annotate(
            total_qty=Coalesce(Sum("quantity"), Decimal("0")),
        ):
            key = obj["custom_description"].strip().upper()
            combined[key]["qty"] += obj["total_qty"]
            if obj["custom_price"]:
                combined[key]["price"] = obj["custom_price"]

    sorted_items = sorted(combined.items(), key=lambda x: x[1]["qty"], reverse=True)
    return sorted_items


def _find_duplicates():
    """Find potential duplicate items using fuzzy matching."""
    items = list(
        Item.objects.filter(is_deleted=False)
        .select_related("category")
        .order_by("name")
    )

    # Normalize names for comparison
    def normalize(name):
        return re.sub(r"\s+", " ", name.strip().lower())

    def tokenize(name):
        return set(normalize(name).split())

    duplicates = []
    seen_pairs = set()

    for i, item_a in enumerate(items):
        norm_a = normalize(item_a.name)
        tokens_a = tokenize(item_a.name)

        for j in range(i + 1, len(items)):
            item_b = items[j]
            norm_b = normalize(item_b.name)

            # Skip if already paired
            pair_key = tuple(sorted([item_a.id, item_b.id]))
            if pair_key in seen_pairs:
                continue

            match_type = None

            # Exact match (after normalization)
            if norm_a == norm_b:
                match_type = "Exact Match"
            else:
                # Token overlap
                tokens_b = tokenize(item_b.name)
                overlap = len(tokens_a & tokens_b)
                total = max(len(tokens_a | tokens_b), 1)
                ratio = overlap / total

                if ratio >= 0.6:
                    match_type = f"Similar ({int(ratio * 100)}% overlap)"
                # Substring check
                elif norm_a in norm_b or norm_b in norm_a:
                    match_type = "Substring Match"

            if match_type:
                seen_pairs.add(pair_key)
                duplicates.append({
                    "item_a": item_a,
                    "item_b": item_b,
                    "match_type": match_type,
                })

    return duplicates


# ── Sheet builders ─────────────────────────────────────────────────

def _build_no_stock_sheet(wb):
    # Stall no-stock
    ws = wb.create_sheet("No Stock - Stall")
    headers = [
        "Item Name", "SKU", "Category", "Unit",
        "Total Qty", "Reserved", "Available", "Threshold",
        "Cost Price", "Retail Price",
    ]
    rows = []
    for s in _get_stock_items("no_stock"):
        rows.append([
            s.item.name, s.item.sku,
            s.item.category.name if s.item.category else "—",
            s.item.unit_of_measure,
            float(s.quantity), float(s.reserved_quantity),
            float(s.avail), float(s.low_stock_threshold),
            float(s.item.cost_price or 0),
            float(s.item.retail_price),
        ])
    last_row = _write_rows(ws, headers, rows)
    for row_idx in range(2, last_row + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).fill = NO_STOCK_FILL

    # Stockroom no-stock
    ws2 = wb.create_sheet("No Stock - Stockroom")
    headers2 = [
        "Item Name", "SKU", "Category", "Unit",
        "Quantity", "Threshold",
        "Cost Price", "Retail Price",
    ]
    rows2 = []
    for sr in _get_stockroom_items("no_stock"):
        rows2.append([
            sr.item.name, sr.item.sku,
            sr.item.category.name if sr.item.category else "—",
            sr.item.unit_of_measure,
            float(sr.quantity), float(sr.low_stock_threshold),
            float(sr.item.cost_price or 0),
            float(sr.item.retail_price),
        ])
    last_row2 = _write_rows(ws2, headers2, rows2)
    for row_idx in range(2, last_row2 + 1):
        for col in range(1, len(headers2) + 1):
            ws2.cell(row=row_idx, column=col).fill = NO_STOCK_FILL

    return ws


def _build_low_stock_sheet(wb):
    # Stall low-stock
    ws = wb.create_sheet("Low Stock - Stall")
    headers = [
        "Item Name", "SKU", "Category", "Unit",
        "Total Qty", "Reserved", "Available", "Threshold",
        "Suggested Order Qty", "Cost Price", "Retail Price",
    ]
    rows = []
    for s in _get_stock_items("low_stock"):
        avail = float(s.avail)
        threshold = float(s.low_stock_threshold)
        suggested = max(threshold * 2 - avail, 0)
        rows.append([
            s.item.name, s.item.sku,
            s.item.category.name if s.item.category else "—",
            s.item.unit_of_measure,
            float(s.quantity), float(s.reserved_quantity),
            avail, threshold, suggested,
            float(s.item.cost_price or 0),
            float(s.item.retail_price),
        ])
    last_row = _write_rows(ws, headers, rows)
    for row_idx in range(2, last_row + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).fill = LOW_STOCK_FILL

    # Stockroom low-stock
    ws2 = wb.create_sheet("Low Stock - Stockroom")
    headers2 = [
        "Item Name", "SKU", "Category", "Unit",
        "Quantity", "Threshold",
        "Suggested Order Qty", "Cost Price", "Retail Price",
    ]
    rows2 = []
    for sr in _get_stockroom_items("low_stock"):
        qty = float(sr.quantity)
        threshold = float(sr.low_stock_threshold)
        suggested = max(threshold * 2 - qty, 0)
        rows2.append([
            sr.item.name, sr.item.sku,
            sr.item.category.name if sr.item.category else "—",
            sr.item.unit_of_measure,
            qty, threshold, suggested,
            float(sr.item.cost_price or 0),
            float(sr.item.retail_price),
        ])
    last_row2 = _write_rows(ws2, headers2, rows2)
    for row_idx in range(2, last_row2 + 1):
        for col in range(1, len(headers2) + 1):
            ws2.cell(row=row_idx, column=col).fill = LOW_STOCK_FILL

    return ws


def _build_most_bought_sheet(wb):
    ws = wb.create_sheet("Most Bought")
    headers = [
        "Rank", "Item Name", "SKU", "Category", "Unit",
        "Total Qty Used", "From Sales", "From Services",
        "Current Stall Stock", "Current Stockroom Stock",
        "Cost Price", "Retail Price",
    ]

    rows = []
    for rank, entry in enumerate(_get_most_bought_items(50), 1):
        item = entry["item"]
        # Get current stock levels
        stall_stock = Stock.objects.filter(
            item=item, is_deleted=False
        ).aggregate(total=Coalesce(Sum(F("quantity") - F("reserved_quantity")), Decimal("0")))
        stockroom_qty = Decimal("0")
        try:
            stockroom_qty = item.stockroom_stock.quantity
        except StockRoomStock.DoesNotExist:
            pass

        rows.append([
            rank, item.name, item.sku,
            item.category.name if item.category else "—",
            item.unit_of_measure,
            float(entry["total_qty"]),
            float(entry["sales_qty"]),
            float(entry["service_qty"]),
            float(stall_stock["total"]),
            float(stockroom_qty),
            float(item.cost_price or 0),
            float(item.retail_price),
        ])

    _write_rows(ws, headers, rows)
    return ws


def _build_least_bought_sheet(wb):
    ws = wb.create_sheet("Least Bought")
    headers = [
        "Rank", "Item Name", "SKU", "Category", "Unit",
        "Total Qty Used", "From Sales", "From Services",
        "Current Stall Stock", "Current Stockroom Stock",
        "Cost Price", "Retail Price",
    ]

    rows = []
    for rank, entry in enumerate(_get_least_bought_items(50), 1):
        item = entry["item"]
        stall_stock = Stock.objects.filter(
            item=item, is_deleted=False
        ).aggregate(total=Coalesce(Sum(F("quantity") - F("reserved_quantity")), Decimal("0")))
        stockroom_qty = Decimal("0")
        try:
            stockroom_qty = item.stockroom_stock.quantity
        except StockRoomStock.DoesNotExist:
            pass

        rows.append([
            rank, item.name, item.sku,
            item.category.name if item.category else "—",
            item.unit_of_measure,
            float(entry["total_qty"]),
            float(entry["sales_qty"]),
            float(entry["service_qty"]),
            float(stall_stock["total"]),
            float(stockroom_qty),
            float(item.cost_price or 0),
            float(item.retail_price),
        ])

    _write_rows(ws, headers, rows)
    return ws


def _build_custom_items_sheet(wb):
    ws = wb.create_sheet("Custom Items")
    headers = [
        "Rank", "Description", "Total Qty Used", "Last Price",
    ]

    rows = []
    for rank, (desc, data) in enumerate(_get_custom_items_usage(), 1):
        rows.append([
            rank, desc, float(data["qty"]), float(data["price"]),
        ])

    _write_rows(ws, headers, rows)
    return ws


def _build_duplicates_sheet(wb):
    ws = wb.create_sheet("Potential Duplicates")
    headers = [
        "Item A", "SKU A", "Category A",
        "Item B", "SKU B", "Category B",
        "Match Type",
    ]

    dupes = _find_duplicates()
    rows = []
    for d in dupes:
        a, b = d["item_a"], d["item_b"]
        rows.append([
            a.name, a.sku, a.category.name if a.category else "—",
            b.name, b.sku, b.category.name if b.category else "—",
            d["match_type"],
        ])

    last_row = _write_rows(ws, headers, rows)

    for row_idx in range(2, last_row + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).fill = DUPLICATE_FILL

    return ws


def _build_by_category_sheet(wb):
    ws = wb.create_sheet("By Category")
    headers = [
        "Category", "Item Name", "SKU", "Unit",
        "Stall Qty", "Stall Available", "Stockroom Qty",
        "Status", "Cost Price", "Retail Price",
    ]

    # Get all active items grouped by category
    items = (
        Item.objects.filter(is_deleted=False)
        .select_related("category")
        .prefetch_related("stocks", "stockroom_stock")
        .order_by("category__name", "name")
    )

    rows = []
    for item in items:
        cat_name = item.category.name if item.category else "Uncategorized"

        # Stall stock info
        stall_stocks = Stock.objects.filter(item=item, is_deleted=False)
        stall_qty = sum(float(s.quantity) for s in stall_stocks)
        stall_avail = sum(float(s.available_quantity) for s in stall_stocks)

        # Stockroom stock info
        stockroom_qty = 0.0
        try:
            stockroom_qty = float(item.stockroom_stock.quantity)
        except StockRoomStock.DoesNotExist:
            pass

        # Determine status
        if stall_avail == 0 and stockroom_qty == 0:
            item_status = "NO STOCK"
        elif any(s.status() == "low_stock" for s in stall_stocks):
            item_status = "LOW STOCK"
        else:
            item_status = "OK"

        rows.append([
            cat_name, item.name, item.sku, item.unit_of_measure,
            stall_qty, stall_avail, stockroom_qty,
            item_status,
            float(item.cost_price or 0),
            float(item.retail_price),
        ])

    _write_rows(ws, headers, rows)

    # Color-code status column
    status_col = headers.index("Status") + 1
    for row_idx in range(2, len(rows) + 2):
        cell = ws.cell(row=row_idx, column=status_col)
        if cell.value == "NO STOCK":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = NO_STOCK_FILL
        elif cell.value == "LOW STOCK":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = LOW_STOCK_FILL

    return ws


def _build_summary_sheet(wb):
    """Summary / overview sheet (first sheet)."""
    ws = wb.active
    ws.title = "Summary"

    # Title
    ws.cell(row=1, column=1, value="RVDC Inventory Report").font = Font(
        bold=True, size=16, color="2563EB"
    )
    ws.cell(row=2, column=1, value=f"Generated: {timezone.now().strftime('%B %d, %Y %I:%M %p')}")
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666")

    # Stats
    total_items = Item.objects.filter(is_deleted=False).count()
    no_stock_stall = _get_stock_items("no_stock").count()
    low_stock_stall = _get_stock_items("low_stock").count()
    no_stock_room = _get_stockroom_items("no_stock").count()
    low_stock_room = _get_stockroom_items("low_stock").count()
    categories = ProductCategory.objects.filter(is_deleted=False).count()

    stats = [
        ("Total Active Items", total_items),
        ("Total Categories", categories),
        ("", ""),
        ("Stall - No Stock Items", no_stock_stall),
        ("Stall - Low Stock Items", low_stock_stall),
        ("Stockroom - No Stock Items", no_stock_room),
        ("Stockroom - Low Stock Items", low_stock_room),
        ("", ""),
        ("Total Items Needing Attention", no_stock_stall + low_stock_stall + no_stock_room + low_stock_room),
    ]

    row = 4
    ws.cell(row=row, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=2).border = THIN_BORDER

    for label, value in stats:
        row += 1
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=value).border = THIN_BORDER
        if label == "Total Items Needing Attention":
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).font = Font(bold=True, color="DC2626")

    row += 2
    ws.cell(row=row, column=1, value="Sheets Included:").font = Font(bold=True)
    sheets_desc = [
        ("No Stock - Stall", "Stall items with zero available quantity"),
        ("No Stock - Stockroom", "Stockroom items with zero quantity"),
        ("Low Stock - Stall", "Stall items below their threshold"),
        ("Low Stock - Stockroom", "Stockroom items below their threshold"),
        ("Most Bought", "Top 50 most consumed items (sales + services)"),
        ("Least Bought", "Bottom 50 least consumed items"),
        ("Custom Items", "Non-inventory items used in services"),
        ("Potential Duplicates", "Items with similar names that may be duplicates"),
        ("By Category", "All items grouped by category with stock status"),
    ]
    for name, desc in sheets_desc:
        row += 1
        ws.cell(row=row, column=1, value=f"  • {name}").font = Font(bold=True)
        ws.cell(row=row, column=2, value=desc)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50

    return ws


# ── Main API View ──────────────────────────────────────────────────

class InventoryExportView(APIView):
    """
    Generate and download an Excel inventory report.

    GET /api/inventory/export-report/
    Query params:
        sheets - comma-separated list of sheets to include
                 (default: all). Options: no_stock, low_stock,
                 most_bought, least_bought, custom_items,
                 duplicates, by_category
    """

    permission_classes = [IsAdminUser]

    SHEET_BUILDERS = {
        "no_stock": _build_no_stock_sheet,
        "low_stock": _build_low_stock_sheet,
        "most_bought": _build_most_bought_sheet,
        "least_bought": _build_least_bought_sheet,
        "custom_items": _build_custom_items_sheet,
        "duplicates": _build_duplicates_sheet,
        "by_category": _build_by_category_sheet,
    }

    def get(self, request):
        sheets_param = request.query_params.get("sheets", "")
        if sheets_param:
            requested = [s.strip() for s in sheets_param.split(",") if s.strip()]
            # Validate requested sheets
            invalid = [s for s in requested if s not in self.SHEET_BUILDERS]
            if invalid:
                from rest_framework.response import Response
                from rest_framework import status
                return Response(
                    {"detail": f"Invalid sheet(s): {', '.join(invalid)}. "
                     f"Valid options: {', '.join(self.SHEET_BUILDERS.keys())}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            requested = list(self.SHEET_BUILDERS.keys())

        wb = Workbook()
        _build_summary_sheet(wb)

        for sheet_key in requested:
            self.SHEET_BUILDERS[sheet_key](wb)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = timezone.now().strftime("%Y%m%d_%H%M")
        filename = f"inventory_report_{timestamp}.xlsx"

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
