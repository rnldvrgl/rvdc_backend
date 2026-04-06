import io
from decimal import Decimal, InvalidOperation

from django.contrib.auth import get_user_model
from django.db import models, transaction
from django.db.models import Q
from django.db.models.functions import Lower, Replace
from django.http import HttpResponse
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from inventory.api.filters import (
    ItemFilter,
    StockFilter,
    StockRoomFilter,
)
from inventory.api.serializers import (
    CustomItemTemplateSerializer,
    ItemSerializer,
    ProductCategorySerializer,
    StallSerializer,
    StockAuditSerializer,
    StockPatchSerializer,
    StockReadSerializer,
    StockRequestSerializer,
    StockRestockSerializer,
    StockRoomStockSerializer,
    StockWriteSerializer,
)
from inventory.models import (
    CustomItemTemplate,
    Item,
    ProductCategory,
    Stall,
    Stock,
    StockRequest,
    StockRoomStock,
)
from notifications.models import Notification
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from rest_framework.response import Response
from utils.filters.options import (
    get_product_category_options,
    get_stall_options,
    get_status_options,
    get_unit_of_measure_options,
)
from utils.filters.role_filters import get_role_based_filter_response
from utils.inventory import (
    create_item_with_initial_stock,
    user_can_manage_stall,
)
from utils.query import (
    filter_by_date_range,
    get_role_filtered_queryset,
)
from utils.soft_delete import SoftDeleteViewSetMixin


class ItemViewSet(SoftDeleteViewSetMixin, viewsets.ModelViewSet):
    queryset = Item.objects.select_related('category').prefetch_related('price_history').all()
    serializer_class = ItemSerializer
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = ItemFilter
    search_fields = ["name"]
    ordering_fields = "__all__"

    def get_queryset(self):
        qs = super().get_queryset().filter(is_deleted=False)
        return filter_by_date_range(self.request, qs)

    # (moved) explicit disallowed-method handlers belong to StallViewSet, not ItemViewSet.

    @action(detail=False, methods=["get"], url_path="filters")
    def get_filters(self, request):
        filters_config = {
            "category": {
                "options": get_product_category_options,
            },
            "unit_of_measure": {
                "options": get_unit_of_measure_options,
            },
            "retail_price": {
                "options": lambda: [
                    {"label": "Is Zero", "value": "true"},
                    {"label": "Is Not Zero", "value": "false"},
                ],
            },
            "wholesale_price": {
                "options": lambda: [
                    {"label": "Is Zero", "value": "true"},
                    {"label": "Is Not Zero", "value": "false"},
                ],
            },
            "technician_price": {
                "options": lambda: [
                    {"label": "Is Zero", "value": "true"},
                    {"label": "Is Not Zero", "value": "false"},
                ],
            },
            "cost_price": {
                "options": lambda: [
                    {"label": "Is Zero", "value": "true"},
                    {"label": "Is Not Zero", "value": "false"},
                ],
            },
            "is_tracked": {
                "options": lambda: [
                    {"label": "Tracked", "value": "true"},
                    {"label": "Untracked", "value": "false"},
                ],
            },
        }

        ordering_config = [
            {"label": "Name", "value": "name"},
            {"label": "Category", "value": "category__name"},
            {"label": "Unit", "value": "unit_of_measure"},
        ]

        return get_role_based_filter_response(
            request,
            filters_config,
            ordering_config,
        )

    @action(
        detail=True,
        methods=["post"],
        url_path="toggle-tracked",
        permission_classes=[IsAdminUser],
    )
    def toggle_tracked(self, request, pk=None):
        """Toggle the is_tracked flag on an item."""
        item = self.get_object()
        item.is_tracked = not item.is_tracked
        item.save(update_fields=["is_tracked", "updated_at"])
        serializer = self.get_serializer(item)
        return Response(serializer.data)

    @action(
        detail=True,
        methods=["post"],
        url_path="merge",
        permission_classes=[IsAdminUser],
    )
    def merge(self, request, pk=None):
        """
        Merge a source item into this (target) item.

        Body:
            source_item_id  – ID of the item to be absorbed and soft-deleted.

        All transaction references (SalesItem, ApplianceItemUsed,
        ServiceItemUsed, ExpenseItem, StockRequest) are re-pointed to the
        target item.  Stall stock and stockroom stock quantities are summed.
        The source item is then soft-deleted.
        """
        from django.db import transaction as db_transaction
        from django.utils import timezone as tz
        from sales.models import SalesItem
        from services.models import ApplianceItemUsed, ServiceItemUsed
        from expenses.models import ExpenseItem

        target = self.get_object()

        source_id = request.data.get("source_item_id")
        if not source_id:
            return Response(
                {"detail": "source_item_id is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            source_id = int(source_id)
        except (TypeError, ValueError):
            return Response(
                {"detail": "source_item_id must be an integer."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if source_id == target.pk:
            return Response(
                {"detail": "Source and target items must be different."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            source = Item.objects.get(pk=source_id, is_deleted=False)
        except Item.DoesNotExist:
            return Response(
                {"detail": "Source item not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        with db_transaction.atomic():
            # ── FK re-assignments ────────────────────────────────────────
            sales_count = SalesItem.objects.filter(item=source).update(item=target)
            appliance_count = ApplianceItemUsed.objects.filter(item=source).update(item=target)
            service_count = ServiceItemUsed.objects.filter(item=source).update(item=target)
            expense_count = ExpenseItem.objects.filter(item=source).update(item=target)
            request_count = StockRequest.objects.filter(item=source).update(item=target)

            # ── Merge stall stock ────────────────────────────────────────
            source_stocks = Stock.objects.filter(item=source, is_deleted=False)
            stocks_merged = 0
            for src_stock in source_stocks:
                target_stock, _ = Stock.objects.get_or_create(
                    item=target,
                    stall=src_stock.stall,
                    defaults={
                        "quantity": 0,
                        "reserved_quantity": 0,
                        "low_stock_threshold": src_stock.low_stock_threshold,
                        "track_stock": src_stock.track_stock,
                    },
                )
                target_stock.quantity += src_stock.quantity
                target_stock.reserved_quantity += src_stock.reserved_quantity
                target_stock.save(update_fields=["quantity", "reserved_quantity", "updated_at"])
                src_stock.is_deleted = True
                src_stock.deleted_at = tz.now()
                src_stock.save(update_fields=["is_deleted", "deleted_at"])
                stocks_merged += 1

            # ── Merge stockroom stock ────────────────────────────────────
            stockroom_merged = False
            try:
                src_room = StockRoomStock.objects.get(item=source)
                try:
                    target_room = StockRoomStock.objects.get(item=target)
                    target_room.quantity += src_room.quantity
                    target_room.save(update_fields=["quantity", "updated_at"])
                except StockRoomStock.DoesNotExist:
                    src_room.item = target
                    src_room.save(update_fields=["item"])
                stockroom_merged = True
            except StockRoomStock.DoesNotExist:
                pass

            # ── Soft-delete source ───────────────────────────────────────
            source.is_deleted = True
            source.deleted_at = tz.now()
            # Prefix SKU so the unique constraint does not block future items
            source.sku = f"MERGED-{source.sku}"
            source.save(update_fields=["is_deleted", "deleted_at", "sku"])

        serializer = self.get_serializer(target)
        return Response(
            {
                "target_item": serializer.data,
                "merged": {
                    "sales_items_updated": sales_count,
                    "appliance_items_updated": appliance_count,
                    "service_items_updated": service_count,
                    "expense_items_updated": expense_count,
                    "stock_requests_updated": request_count,
                    "stall_stocks_merged": stocks_merged,
                    "stockroom_merged": stockroom_merged,
                },
            },
            status=status.HTTP_200_OK,
        )

    @action(
        detail=False,
        methods=["get"],
        url_path="bulk-template",
        permission_classes=[IsAdminUser],
    )
    def bulk_template(self, request):
        """Download an XLSX file pre-filled with all active items and their prices."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        items = Item.objects.filter(is_deleted=False).select_related("category").order_by("name")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Items"

        headers = ["SKU", "Name", "Category", "Cost Price", "Retail Price", "Wholesale Price", "Technician Price", "Action"]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, item in enumerate(items, 2):
            ws.cell(row=row_idx, column=1, value=item.sku).border = thin_border
            ws.cell(row=row_idx, column=2, value=item.name).border = thin_border
            ws.cell(row=row_idx, column=3, value=item.category.name if item.category else "").border = thin_border
            ws.cell(row=row_idx, column=4, value=float(item.cost_price or 0)).border = thin_border
            ws.cell(row=row_idx, column=5, value=float(item.retail_price or 0)).border = thin_border
            ws.cell(row=row_idx, column=6, value=float(item.wholesale_price or 0)).border = thin_border
            ws.cell(row=row_idx, column=7, value=float(item.technician_price or 0)).border = thin_border
            ws.cell(row=row_idx, column=8, value="").border = thin_border

        ws.sheet_properties.tabColor = "1F4E79"

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="item_pricing_template.xlsx"'
        return response

    @action(
        detail=False,
        methods=["post"],
        url_path="bulk-preview",
        permission_classes=[IsAdminUser],
    )
    def bulk_preview(self, request):
        """
        Upload an XLSX file to preview changes before applying them.
        Returns a list of changes (old vs new) without saving anything.
        Supports:
          - Update:  SKU present, Action blank → update prices/name
          - Delete:  SKU present, Action = "DELETE" → soft-delete item
          - Create:  SKU blank, Name present → create new item
        """
        import openpyxl

        xlsx_file = request.FILES.get("file")
        if not xlsx_file:
            return Response(
                {"detail": "No file uploaded. Send as 'file' in multipart form data."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not xlsx_file.name.endswith((".xlsx", ".xlsm")):
            return Response(
                {"detail": "Only .xlsx files are supported."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if xlsx_file.size > 5 * 1024 * 1024:
            return Response(
                {"detail": "File too large. Maximum 5 MB."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        except Exception:
            return Response(
                {"detail": "Could not parse the uploaded file. Ensure it is a valid .xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        if not rows:
            return Response(
                {"detail": "The file contains no data rows."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        all_items = {
            item.sku: item
            for item in Item.objects.filter(is_deleted=False)
        }

        changes = []
        skipped = 0
        errors = []

        for row_num, row in enumerate(rows, 2):
            if len(row) < 7:
                errors.append({"row": row_num, "error": "Row has fewer than 7 columns."})
                continue

            sku = str(row[0] or "").strip()
            action_val = str(row[7] or "").strip().upper() if len(row) > 7 else ""

            # ── CREATE: blank SKU means new item ──────────────────────────────
            if not sku:
                new_name = str(row[1] or "").strip()
                if not new_name:
                    continue  # fully blank row — skip quietly

                try:
                    new_retail = Decimal(str(row[4])) if row[4] is not None and str(row[4]).strip() != "" else None
                except (InvalidOperation, ValueError):
                    errors.append({"row": row_num, "error": "Invalid retail price for new item."})
                    continue

                if new_retail is None or new_retail <= 0:
                    errors.append({"row": row_num, "error": "New items require a positive retail price."})
                    continue

                changes.append({
                    "row": row_num,
                    "sku": "",
                    "name": new_name,
                    "action": "create",
                    "changes": [{"field": "Name", "old": "", "new": new_name}],
                })
                continue

            item = all_items.get(sku)
            if not item:
                errors.append({"row": row_num, "sku": sku, "error": "SKU not found."})
                continue

            # ── DELETE ────────────────────────────────────────────────────────
            if action_val == "DELETE":
                changes.append({
                    "row": row_num,
                    "sku": sku,
                    "name": item.name,
                    "action": "delete",
                    "changes": [],
                })
                continue

            # ── UPDATE ────────────────────────────────────────────────────────
            new_name = str(row[1] or "").strip()

            try:
                new_cost = Decimal(str(row[3])) if row[3] is not None and str(row[3]).strip() != "" else None
                new_retail = Decimal(str(row[4])) if row[4] is not None and str(row[4]).strip() != "" else None
                new_wholesale = Decimal(str(row[5])) if row[5] is not None and str(row[5]).strip() != "" else None
                new_tech = Decimal(str(row[6])) if row[6] is not None and str(row[6]).strip() != "" else None
            except (InvalidOperation, ValueError) as e:
                errors.append({"row": row_num, "sku": sku, "error": f"Invalid price value: {e}"})
                continue

            for label, val in [("cost", new_cost), ("retail", new_retail), ("wholesale", new_wholesale), ("technician", new_tech)]:
                if val is not None and val < 0:
                    errors.append({"row": row_num, "sku": sku, "error": f"{label} price cannot be negative."})
                    break
            else:
                item_changes = []
                if new_name and new_name != item.name:
                    item_changes.append({"field": "Name", "old": item.name, "new": new_name})
                if new_cost is not None and new_cost != item.cost_price:
                    item_changes.append({"field": "Cost Price", "old": str(item.cost_price), "new": str(new_cost)})
                if new_retail is not None and new_retail != item.retail_price:
                    item_changes.append({"field": "Retail Price", "old": str(item.retail_price), "new": str(new_retail)})
                if new_wholesale is not None and new_wholesale != (item.wholesale_price or Decimal("0")):
                    item_changes.append({"field": "Wholesale Price", "old": str(item.wholesale_price or 0), "new": str(new_wholesale)})
                if new_tech is not None and new_tech != (item.technician_price or Decimal("0")):
                    item_changes.append({"field": "Technician Price", "old": str(item.technician_price or 0), "new": str(new_tech)})

                if item_changes:
                    changes.append({
                        "row": row_num,
                        "sku": sku,
                        "name": item.name,
                        "action": "update",
                        "changes": item_changes,
                    })
                else:
                    skipped += 1

        creates = sum(1 for c in changes if c["action"] == "create")
        deletes = sum(1 for c in changes if c["action"] == "delete")
        updates = sum(1 for c in changes if c["action"] == "update")

        return Response({
            "changes": changes,
            "skipped": skipped,
            "errors": errors,
            "summary": (
                f"{updates} to update, {creates} to create, {deletes} to delete, "
                f"{skipped} unchanged, {len(errors)} errors."
            ),
        })

    @action(
        detail=False,
        methods=["post"],
        url_path="bulk-update",
        permission_classes=[IsAdminUser],
    )
    def bulk_update(self, request):
        """
        Upload an XLSX file to bulk-update item names and prices.
        Also supports creating new items (blank SKU) and deleting existing
        items (Action = DELETE).
        Validates the file synchronously, then processes in a background
        thread. Results are delivered via WebSocket notification.
        """
        import threading

        import openpyxl

        xlsx_file = request.FILES.get("file")
        if not xlsx_file:
            return Response(
                {"detail": "No file uploaded. Send as 'file' in multipart form data."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not xlsx_file.name.endswith((".xlsx", ".xlsm")):
            return Response(
                {"detail": "Only .xlsx files are supported."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if xlsx_file.size > 5 * 1024 * 1024:
            return Response(
                {"detail": "File too large. Maximum 5 MB."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        except Exception:
            return Response(
                {"detail": "Could not parse the uploaded file. Ensure it is a valid .xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        if not rows:
            return Response(
                {"detail": "The file contains no data rows."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user_id = request.user.id

        excluded_rows_raw = request.data.get("excluded_rows", "[]")
        try:
            import json as _json
            excluded_rows = set(_json.loads(excluded_rows_raw))
        except Exception:
            excluded_rows = set()

        def _process_bulk_update():
            try:
                all_items = {
                    item.sku: item
                    for item in Item.objects.filter(is_deleted=False)
                }

                updated = []
                to_delete = []
                to_create = []
                skipped = []
                errors = []

                for row_num, row in enumerate(rows, 2):
                    if row_num in excluded_rows:
                        skipped.append(str(row[0] or ""))
                        continue

                    if len(row) < 7:
                        errors.append({"row": row_num, "error": "Row has fewer than 7 columns."})
                        continue

                    sku = str(row[0] or "").strip()
                    action_val = str(row[7] or "").strip().upper() if len(row) > 7 else ""

                    # ── CREATE ────────────────────────────────────────────────
                    if not sku:
                        new_name = str(row[1] or "").strip()
                        if not new_name:
                            continue

                        try:
                            new_retail = Decimal(str(row[4])) if row[4] is not None and str(row[4]).strip() != "" else Decimal("0")
                            new_cost = Decimal(str(row[3])) if row[3] is not None and str(row[3]).strip() != "" else Decimal("0")
                            new_wholesale = Decimal(str(row[5])) if row[5] is not None and str(row[5]).strip() != "" else Decimal("0")
                            new_tech = Decimal(str(row[6])) if row[6] is not None and str(row[6]).strip() != "" else Decimal("0")
                        except (InvalidOperation, ValueError) as e:
                            errors.append({"row": row_num, "error": f"Invalid price value: {e}"})
                            continue

                        category_name = str(row[2] or "").strip()
                        category = None
                        if category_name:
                            try:
                                category = ProductCategory.objects.get(name__iexact=category_name, is_deleted=False)
                            except ProductCategory.DoesNotExist:
                                pass

                        to_create.append({
                            "name": new_name,
                            "category": category,
                            "retail_price": new_retail,
                            "cost_price": new_cost,
                            "wholesale_price": new_wholesale,
                            "technician_price": new_tech,
                        })
                        continue

                    item = all_items.get(sku)
                    if not item:
                        errors.append({"row": row_num, "sku": sku, "error": "SKU not found."})
                        continue

                    # ── DELETE ────────────────────────────────────────────────
                    if action_val == "DELETE":
                        to_delete.append(item)
                        continue

                    # ── UPDATE ────────────────────────────────────────────────
                    new_name = str(row[1] or "").strip()

                    try:
                        new_cost = Decimal(str(row[3])) if row[3] is not None and str(row[3]).strip() != "" else None
                        new_retail = Decimal(str(row[4])) if row[4] is not None and str(row[4]).strip() != "" else None
                        new_wholesale = Decimal(str(row[5])) if row[5] is not None and str(row[5]).strip() != "" else None
                        new_tech = Decimal(str(row[6])) if row[6] is not None and str(row[6]).strip() != "" else None
                    except (InvalidOperation, ValueError) as e:
                        errors.append({"row": row_num, "sku": sku, "error": f"Invalid price value: {e}"})
                        continue

                    for label, val in [("cost", new_cost), ("retail", new_retail), ("wholesale", new_wholesale), ("technician", new_tech)]:
                        if val is not None and val < 0:
                            errors.append({"row": row_num, "sku": sku, "error": f"{label} price cannot be negative."})
                            break
                    else:
                        changed = False
                        if new_name and new_name != item.name:
                            item.name = new_name
                            changed = True
                        if new_cost is not None and new_cost != item.cost_price:
                            item.cost_price = new_cost
                            changed = True
                        if new_retail is not None and new_retail != item.retail_price:
                            item.retail_price = new_retail
                            changed = True
                        if new_wholesale is not None and new_wholesale != (item.wholesale_price or Decimal("0")):
                            item.wholesale_price = new_wholesale
                            changed = True
                        if new_tech is not None and new_tech != (item.technician_price or Decimal("0")):
                            item.technician_price = new_tech
                            changed = True

                        if changed:
                            updated.append(item)
                        else:
                            skipped.append(sku)

                with transaction.atomic():
                    for item in updated:
                        item.save()
                    for item in to_delete:
                        item.is_deleted = True
                        item.deleted_at = timezone.now()
                        item.save(track_history=False)
                    for data in to_create:
                        create_item_with_initial_stock(data)

                detail = (
                    f"Updated {len(updated)}, created {len(to_create)}, "
                    f"deleted {len(to_delete)}, skipped {len(skipped)} unchanged, "
                    f"{len(errors)} errors."
                )
                _notify_bulk_update(user_id, {
                    "updated": len(updated),
                    "created": len(to_create),
                    "deleted": len(to_delete),
                    "skipped": len(skipped),
                    "errors": errors,
                    "detail": detail,
                })
            except Exception:
                import logging
                logging.getLogger(__name__).exception("Bulk update failed")
                _notify_bulk_update_failed(user_id)

        threading.Thread(target=_process_bulk_update, daemon=True).start()

        return Response(
            {"detail": "Bulk update started. You will be notified when it's done."},
            status=status.HTTP_202_ACCEPTED,
        )

    @action(detail=False, methods=["get"], url_path="check-duplicates")
    def check_duplicates(self, request):
        """
        Check for potential duplicate items matching a given name.

        Query params:
            name  – the item name to check (required)
            category_id – optional category filter
            exclude_id  – optional item id to exclude (for edits)

        Returns a list of possible matches ranked by similarity:
            1. Exact (case-insensitive) match
            2. Normalized match (ignoring whitespace & case)
            3. Substring / contains match
            4. Token (word) overlap match — catches jumbled names
               e.g. "capacitor 10uf" vs "10uf capacitor"
        """
        import re

        raw_name = request.query_params.get("name", "").strip()
        category_id = request.query_params.get("category_id")
        exclude_id = request.query_params.get("exclude_id")

        if not raw_name or len(raw_name) < 2:
            return Response([])

        # Normalize: lowercase, collapse whitespace
        normalized = re.sub(r"\s+", " ", raw_name.strip().lower())
        tokens = set(normalized.split())

        qs = Item.objects.filter(is_deleted=False)
        if exclude_id:
            qs = qs.exclude(pk=exclude_id)

        # Annotate a normalized version of the stored name for comparison
        qs = qs.annotate(
            lower_name=Lower("name"),
        )

        # 1. Exact case-insensitive match
        exact = qs.filter(name__iexact=raw_name)
        if category_id:
            exact_cat = exact.filter(category_id=category_id)
        else:
            exact_cat = exact.none()

        # 2. Contains match (substring)
        contains = qs.filter(name__icontains=raw_name).exclude(
            pk__in=exact.values_list("pk", flat=True)
        )

        # 3. Token overlap — fetch candidates that contain ANY of the tokens
        token_q = Q()
        for token in tokens:
            if len(token) >= 2:  # skip very short tokens
                token_q |= Q(name__icontains=token)

        token_candidates = (
            qs.filter(token_q)
            .exclude(pk__in=exact.values_list("pk", flat=True))
            .exclude(pk__in=contains.values_list("pk", flat=True))
        )

        # Score token candidates by overlap ratio
        scored = []
        for item in token_candidates:
            item_tokens = set(item.name.lower().split())
            overlap = len(tokens & item_tokens)
            total = max(len(tokens | item_tokens), 1)
            ratio = overlap / total
            if ratio >= 0.4:  # at least 40% word overlap
                scored.append((item, ratio))
        scored.sort(key=lambda x: -x[1])

        # Build results
        results = []
        seen = set()

        def add(item, match_type):
            if item.pk not in seen:
                seen.add(item.pk)
                results.append({
                    "id": item.pk,
                    "name": item.name,
                    "sku": item.sku,
                    "category": item.category.name if item.category else None,
                    "category_id": item.category_id,
                    "match_type": match_type,
                })

        for item in exact_cat[:5]:
            add(item, "exact_same_category")
        for item in exact.exclude(pk__in=exact_cat.values_list("pk", flat=True))[:5]:
            add(item, "exact")
        for item in contains[:5]:
            add(item, "contains")
        for item, _ in scored[:5]:
            add(item, "similar")

        return Response(results)

    @action(detail=False, methods=["get"], url_path="custom-migration-summary", permission_classes=[IsAdminUser])
    def custom_migration_summary(self, request):
        """
        Returns all unique custom-item descriptions that still have item=null
        across SalesItem, ApplianceItemUsed, and ServiceItemUsed, grouped and
        annotated with usage counts and suggested inventory matches.
        """
        import re
        from sales.models import SalesItem
        from services.models import ApplianceItemUsed, ServiceItemUsed

        # ---- Aggregate unique descriptions from all three tables ----
        from django.db.models import Count, Avg

        si_rows = (
            SalesItem.objects.filter(item__isnull=True)
            .exclude(description="")
            .values("description")
            .annotate(count=Count("id"), avg_price=Avg("final_price_per_unit"))
        )

        appliance_rows = (
            ApplianceItemUsed.objects.filter(item__isnull=True)
            .exclude(custom_description="")
            .values("custom_description")
            .annotate(count=Count("id"), avg_price=Avg("custom_price"))
        )

        service_rows = (
            ServiceItemUsed.objects.filter(item__isnull=True)
            .exclude(custom_description="")
            .values("custom_description")
            .annotate(count=Count("id"), avg_price=Avg("custom_price"))
        )

        # Merge into a single dict keyed by normalised description
        aggregated: dict[str, dict] = {}
        for row in si_rows:
            key = row["description"].strip().lower()
            if key not in aggregated:
                aggregated[key] = {"description": row["description"].strip(), "count": 0, "avg_price": None, "sources": []}
            aggregated[key]["count"] += row["count"]
            aggregated[key]["avg_price"] = row["avg_price"]
            aggregated[key]["sources"].append("sales")

        for row in appliance_rows:
            key = row["custom_description"].strip().lower()
            if key not in aggregated:
                aggregated[key] = {"description": row["custom_description"].strip(), "count": 0, "avg_price": None, "sources": []}
            aggregated[key]["count"] += row["count"]
            if aggregated[key]["avg_price"] is None:
                aggregated[key]["avg_price"] = row["avg_price"]
            aggregated[key]["sources"].append("service_appliance")

        for row in service_rows:
            key = row["custom_description"].strip().lower()
            if key not in aggregated:
                aggregated[key] = {"description": row["custom_description"].strip(), "count": 0, "avg_price": None, "sources": []}
            aggregated[key]["count"] += row["count"]
            if aggregated[key]["avg_price"] is None:
                aggregated[key]["avg_price"] = row["avg_price"]
            aggregated[key]["sources"].append("service_level")

        # ---- Suggest matches for each description ----
        all_items = list(Item.objects.filter(is_deleted=False).values("id", "name", "sku", "retail_price", "is_tracked"))

        def find_suggestions(desc):
            normalised = re.sub(r"\s+", " ", desc.strip().lower())
            tokens = set(normalised.split())
            suggestions = []
            seen_ids: set[int] = set()

            for item in all_items:
                item_name_lower = item["name"].lower()
                if item_name_lower == normalised:
                    suggestions.insert(0, {**item, "match_type": "exact"})
                    seen_ids.add(item["id"])
                    continue
                if normalised in item_name_lower or item_name_lower in normalised:
                    if item["id"] not in seen_ids:
                        suggestions.append({**item, "match_type": "contains"})
                        seen_ids.add(item["id"])
                    continue
                item_tokens = set(item_name_lower.split())
                overlap = len(tokens & item_tokens)
                total = max(len(tokens | item_tokens), 1)
                if overlap / total >= 0.4:
                    if item["id"] not in seen_ids:
                        suggestions.append({**item, "match_type": "similar"})
                        seen_ids.add(item["id"])

            return suggestions[:5]

        results = []
        for entry in sorted(aggregated.values(), key=lambda x: -x["count"]):
            entry["suggestions"] = find_suggestions(entry["description"])
            entry["avg_price"] = float(entry["avg_price"]) if entry["avg_price"] else None
            entry["sources"] = list(set(entry["sources"]))
            results.append(entry)

        return Response(results)

    @action(detail=False, methods=["post"], url_path="link-custom-items", permission_classes=[IsAdminUser])
    def link_custom_items(self, request):
        """
        Links all item=null rows that match a given description to an existing or
        new untracked Item. Also supports creating the untracked item on the fly.

        Body:
          description    str   — the custom description to match (case-insensitive)
          item_id        int   — existing Item to link to (mutually exclusive with create_item)
          create_item    bool  — if True, create a new untracked Item from name+price
          item_name      str   — name for the new item (used when create_item=True)
          item_price     dec   — retail_price for the new item (used when create_item=True)
          item_cost      dec   — cost_price for the new item (optional)

        Returns: { linked: int, item: {id, name, is_tracked} }
        """
        from sales.models import SalesItem
        from services.models import ApplianceItemUsed, ServiceItemUsed
        from django.db import transaction

        description = (request.data.get("description") or "").strip()
        if not description:
            return Response({"detail": "description is required."}, status=400)

        item_id = request.data.get("item_id")
        create_item_flag = bool(request.data.get("create_item", False))

        if not item_id and not create_item_flag:
            return Response({"detail": "Provide either item_id or create_item=true."}, status=400)

        with transaction.atomic():
            if create_item_flag:
                item_name = (request.data.get("item_name") or description).strip()
                try:
                    item_price = float(request.data.get("item_price") or 0)
                    item_cost = float(request.data.get("item_cost") or 0)
                except (ValueError, TypeError):
                    return Response({"detail": "item_price and item_cost must be numbers."}, status=400)

                target_item = Item(
                    name=item_name,
                    retail_price=item_price,
                    cost_price=item_cost or None,
                    is_tracked=False,
                )
                target_item.save()
            else:
                try:
                    target_item = Item.objects.get(pk=item_id)
                except Item.DoesNotExist:
                    return Response({"detail": "Item not found."}, status=404)

            total_linked = 0

            # SalesItem: match by description
            si_qs = SalesItem.objects.filter(item__isnull=True, description__iexact=description)
            count = si_qs.count()
            si_qs.update(item=target_item)
            total_linked += count

            # ApplianceItemUsed: match by custom_description
            aiu_qs = ApplianceItemUsed.objects.filter(item__isnull=True, custom_description__iexact=description)
            count = aiu_qs.count()
            aiu_qs.update(item=target_item)
            total_linked += count

            # ServiceItemUsed: match by custom_description
            siu_qs = ServiceItemUsed.objects.filter(item__isnull=True, custom_description__iexact=description)
            count = siu_qs.count()
            siu_qs.update(item=target_item)
            total_linked += count

        return Response({
            "linked": total_linked,
            "item": {
                "id": target_item.id,
                "name": target_item.name,
                "sku": target_item.sku,
                "retail_price": float(target_item.retail_price),
                "is_tracked": target_item.is_tracked,
            },
        })


class StallViewSet(viewsets.ModelViewSet):
    queryset = Stall.objects.all()

    serializer_class = StallSerializer

    permission_classes = [IsAdminUser]

    # Disable create/update/delete through this viewset — stalls are system-managed.

    http_method_names = ["get", "head", "options"]

    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]

    filterset_fields = ["name"]

    search_fields = ["name"]

    ordering_fields = "__all__"

    def get_queryset(self):
        return filter_by_date_range(self.request, super().get_queryset())

    # Explicit handlers to provide a clear message for disallowed methods on stalls
    def create(self, request, *args, **kwargs):
        return Response(
            {
                "detail": "Stalls are system-managed (read-only). Creation is not allowed."
            },
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def update(self, request, *args, **kwargs):
        return Response(
            {
                "detail": "Stalls are system-managed (read-only). Updates are not allowed."
            },
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def partial_update(self, request, *args, **kwargs):
        return Response(
            {
                "detail": "Stalls are system-managed (read-only). Updates are not allowed."
            },
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def destroy(self, request, *args, **kwargs):
        return Response(
            {
                "detail": "Stalls are system-managed (read-only). Deletion is not allowed."
            },
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )


def _notify_bulk_update(user_id, result):
    """Push bulk_update_complete event via WebSocket."""
    try:
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer

        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f"notifications_{user_id}",
                {
                    "type": "send_notification",
                    "data": {
                        "event": "export_ready",
                        "export_type": "bulk_update",
                        "title": "Bulk Update Complete",
                        "message": result["detail"],
                        "result": result,
                    },
                },
            )
    except Exception:
        import logging
        logging.getLogger(__name__).exception("Failed to send bulk_update via WebSocket")


def _notify_bulk_update_failed(user_id):
    """Push bulk_update_failed event via WebSocket."""
    try:
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer

        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f"notifications_{user_id}",
                {
                    "type": "send_notification",
                    "data": {
                        "event": "export_failed",
                        "export_type": "bulk_update",
                        "title": "Bulk Update Failed",
                        "message": "Failed to process the bulk update. Please try again.",
                    },
                },
            )
    except Exception:
        import logging
        logging.getLogger(__name__).exception("Failed to send bulk_update_failed via WebSocket")


def _notify_stall_stock_bulk_update(user_id, result):
    """Push stall_stock_bulk_update_complete event via WebSocket."""
    try:
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer

        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f"notifications_{user_id}",
                {
                    "type": "send_notification",
                    "data": {
                        "event": "export_ready",
                        "export_type": "stall_stock_bulk_update",
                        "title": "Stall Stock Bulk Update Complete",
                        "message": result["detail"],
                        "result": result,
                    },
                },
            )
    except Exception:
        import logging
        logging.getLogger(__name__).exception("Failed to send stall_stock_bulk_update via WebSocket")


def _notify_stall_stock_bulk_update_failed(user_id):
    """Push stall_stock_bulk_update_failed event via WebSocket."""
    try:
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer

        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f"notifications_{user_id}",
                {
                    "type": "send_notification",
                    "data": {
                        "event": "export_failed",
                        "export_type": "stall_stock_bulk_update",
                        "title": "Stall Stock Bulk Update Failed",
                        "message": "Failed to process the stall stock bulk update. Please try again.",
                    },
                },
            )
    except Exception:
        import logging
        logging.getLogger(__name__).exception("Failed to send stall_stock_bulk_update_failed via WebSocket")


def _notify_stockroom_bulk_update(user_id, result):
    """Push stockroom_bulk_update_complete event via WebSocket."""
    try:
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer

        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f"notifications_{user_id}",
                {
                    "type": "send_notification",
                    "data": {
                        "event": "export_ready",
                        "export_type": "stockroom_bulk_update",
                        "title": "Stockroom Bulk Update Complete",
                        "message": result["detail"],
                        "result": result,
                    },
                },
            )
    except Exception:
        import logging
        logging.getLogger(__name__).exception("Failed to send stockroom_bulk_update via WebSocket")


def _notify_stockroom_bulk_update_failed(user_id):
    """Push stockroom_bulk_update_failed event via WebSocket."""
    try:
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer

        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f"notifications_{user_id}",
                {
                    "type": "send_notification",
                    "data": {
                        "event": "export_failed",
                        "export_type": "stockroom_bulk_update",
                        "title": "Stockroom Bulk Update Failed",
                        "message": "Failed to process the stockroom bulk update. Please try again.",
                    },
                },
            )
    except Exception:
        import logging
        logging.getLogger(__name__).exception("Failed to send stockroom_bulk_update_failed via WebSocket")


class StockViewSet(SoftDeleteViewSetMixin, viewsets.ModelViewSet):
    queryset = Stock.objects.select_related(
        'item__category', 'item__stockroom_stock', 'stall'
    ).all()
    permission_classes = [IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = StockFilter
    search_fields = ["stall__name", "item__name"]
    ordering_fields = "__all__"

    def get_serializer_class(self):
        if self.action in ["list", "retrieve"]:
            return StockReadSerializer
        elif self.action in ["partial_update", "update"]:
            return StockPatchSerializer
        elif self.action == "create":
            return StockWriteSerializer
        return StockReadSerializer

    def get_queryset(self):
        queryset = super().get_queryset().filter(
            is_deleted=False,
            item__is_deleted=False,
        )
        # Annotate with available quantity for filtering
        queryset = queryset.annotate(
            available_expr=models.F('quantity') - models.F('reserved_quantity')
        )
        return filter_by_date_range(self.request, queryset)

    @action(detail=False, methods=["get"], url_path="filters")
    def get_filters(self, request):
        filters_config = {
            "status": {
                "options": get_status_options,
            },
        }

        ordering_config = [
            {"label": "Item Name", "value": "item__name"},
            {"label": "Quantity", "value": "quantity"},
        ]

        return get_role_based_filter_response(request, filters_config, ordering_config)

    @action(detail=False, methods=["get"], url_path="status-counts")
    def status_counts(self, request):
        qs = self.get_queryset()
        no_stock = qs.filter(available_expr__lte=0).count()
        low_stock = qs.filter(
            available_expr__gt=0,
            available_expr__lte=models.F("low_stock_threshold"),
        ).count()
        high_stock = qs.filter(
            available_expr__gt=models.F("low_stock_threshold"),
        ).count()
        return Response({
            "no_stock": no_stock,
            "low_stock": low_stock,
            "high_stock": high_stock,
        })

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated])
    @transaction.atomic
    def restock(self, request, pk=None):
        stock = self.get_object()
        if not user_can_manage_stall(request.user, stock.stall):
            return Response(
                {"detail": "You do not have permission to restock this stall."},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = StockRestockSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        quantity = serializer.validated_data["quantity"]

        try:
            stock_room_stock = StockRoomStock.objects.get(item=stock.item)
        except StockRoomStock.DoesNotExist:
            return Response(
                {"detail": "No stock found in stock room for this item."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if stock_room_stock.quantity < quantity:
            return Response(
                {
                    "detail": f"Not enough stock in stock room. Available: {stock_room_stock.quantity}."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Deduct from stock room
        stock_room_stock.quantity -= quantity
        stock_room_stock.save()

        # Add to stall
        stock.quantity += quantity
        stock.save()

        manager_user = (
            get_user_model()
            .objects.filter(assigned_stall=stock.stall, role__in=["manager", "clerk"], is_active=True)
            .first()
        )

        if manager_user:
            Notification.objects.create(
                user=manager_user,
                type="stock_restocked",
                title="Stock Restocked",
                data={
                    "stall": stock.stall.name,
                    "item": stock.item.name,
                    "item_id": stock.item.id,
                    "stock_id": stock.id,
                    "quantity": float(quantity),
                    "new_total": float(stock.quantity),
                },
                message=f"{quantity} {stock.item.unit_of_measure} of '{stock.item.name}' restocked to {stock.stall.name}.",
            )

        return Response(
            {"detail": f"Restocked successfully. New quantity: {stock.quantity}"}
        )

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated])
    @transaction.atomic
    def add_stock(self, request, pk=None):
        """
        Temporary endpoint to directly add stock to stall without stock room.
        This bypasses the stock room process for quick inventory sync.
        """
        stock = self.get_object()

        # Check permissions
        if not (request.user.role == "admin" or user_can_manage_stall(request.user, stock.stall)):
            return Response(
                {"detail": "You do not have permission to add stock to this stall."},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = StockRestockSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        quantity = serializer.validated_data["quantity"]

        # Directly add to stall stock
        stock.quantity += quantity
        stock.save()

        # Notify manager
        manager_user = (
            get_user_model()
            .objects.filter(assigned_stall=stock.stall, role__in=["manager", "clerk"], is_active=True)
            .first()
        )

        if manager_user:
            Notification.objects.create(
                user=manager_user,
                type="stock_restocked",
                title="Stock Added",
                data={
                    "stall": stock.stall.name,
                    "item": stock.item.name,
                    "item_id": stock.item.id,
                    "stock_id": stock.id,
                    "quantity": float(quantity),
                    "new_total": float(stock.quantity),
                },
                message=f"{quantity} {stock.item.unit_of_measure} of '{stock.item.name}' added to {stock.stall.name} (direct add).",
            )

        return Response(
            {
                "detail": f"Stock added successfully. New quantity: {stock.quantity}",
                "quantity": float(stock.quantity),
            }
        )

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated], url_path="pull-out")
    @transaction.atomic
    def pull_out(self, request, pk=None):
        """
        Pull out / remove stock from stall (defective, damaged, etc.).
        Deducts quantity without creating a sales transaction.
        """
        stock = self.get_object()

        if not (request.user.role == "admin" or user_can_manage_stall(request.user, stock.stall)):
            return Response(
                {"detail": "You do not have permission to pull out stock from this stall."},
                status=status.HTTP_403_FORBIDDEN,
            )

        quantity = request.data.get("quantity")

        if quantity is None:
            return Response(
                {"detail": "Quantity is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            from decimal import Decimal, InvalidOperation
            quantity = Decimal(str(quantity))
        except (TypeError, ValueError, InvalidOperation):
            return Response(
                {"detail": "Quantity must be a number."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if quantity <= 0:
            return Response(
                {"detail": "Quantity must be greater than zero."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if stock.available_quantity < quantity:
            return Response(
                {"detail": f"Not enough available stock. Available: {stock.available_quantity}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        old_quantity = stock.quantity
        stock.quantity -= quantity
        stock.save()

        return Response({
            "detail": f"Pulled out {quantity} {stock.item.unit_of_measure} of '{stock.item.name}' from {stock.stall.name}.",
            "item_name": stock.item.name,
            "quantity_removed": float(quantity),
            "old_quantity": float(old_quantity),
            "new_quantity": float(stock.quantity),

        })

    @action(detail=True, methods=["get", "post"], permission_classes=[IsAdminUser], url_path="audit")
    @transaction.atomic
    def audit(self, request, pk=None):
        """
        Stock audit/reconciliation tool (admin only).

        GET: Returns the current stock breakdown and active reservations.
        POST: Accepts physical_count and adjusts system quantity to match,
              preserving reserved_quantity.
        """
        from services.models import ApplianceItemUsed, Service

        stock = self.get_object()

        # Gather active services that have reserved items from this stock
        active_statuses = ["pending", "in_progress", "on_hold"]
        reserved_items = (
            ApplianceItemUsed.objects.filter(
                stall_stock=stock,
                is_cancelled=False,
                appliance__service__status__in=active_statuses,
            )
            .select_related(
                "appliance__service__client",
                "item",
            )
            .order_by("-appliance__service__created_at")
        )

        reservations = []
        for aiu in reserved_items:
            service = aiu.appliance.service
            reservations.append({
                "service_id": service.id,
                "client_name": str(service.client) if service.client else "N/A",
                "service_type": service.service_type,
                "service_status": service.status,
                "item_name": aiu.item.name if aiu.item else "N/A",
                "quantity_used": float(aiu.quantity),
                "created_at": service.created_at.isoformat(),
            })

        breakdown = {
            "stock_id": stock.id,
            "item_name": stock.item.name,
            "item_unit": stock.item.unit_of_measure,
            "stall_name": stock.stall.name if stock.stall else "N/A",
            "system_quantity": float(stock.quantity),
            "reserved_quantity": float(stock.reserved_quantity),
            "available_quantity": float(stock.available_quantity),
            "reservations": reservations,
        }

        if request.method == "GET":
            return Response(breakdown)

        # POST - reconcile
        serializer = StockAuditSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        physical_count = serializer.validated_data["physical_count"]

        old_quantity = stock.quantity
        # The physical count represents the TOTAL items physically present,
        # which includes items reserved for active services.
        # So system quantity should be set to the physical count.
        stock.quantity = physical_count

        # Ensure reserved_quantity doesn't exceed new quantity
        if stock.reserved_quantity > stock.quantity:
            stock.reserved_quantity = stock.quantity

        stock.save(update_fields=["quantity", "reserved_quantity", "updated_at"])

        discrepancy = float(physical_count) - float(old_quantity)

        return Response({
            **breakdown,
            "system_quantity": float(stock.quantity),
            "reserved_quantity": float(stock.reserved_quantity),
            "available_quantity": float(stock.available_quantity),
            "physical_count": float(physical_count),
            "old_quantity": float(old_quantity),
            "discrepancy": discrepancy,
            "adjusted": True,
        })

    @action(
        detail=False,
        methods=["get"],
        url_path="bulk-template",
        permission_classes=[IsAdminUser],
    )
    def bulk_template(self, request):
        """Download an XLSX template pre-filled with all active stall stock entries."""
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        stocks = (
            Stock.objects.filter(is_deleted=False, item__is_deleted=False)
            .select_related("item__category", "stall")
            .order_by("item__name")
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stall Stock"

        headers = ["Stock ID", "SKU", "Item Name", "Category", "Stall", "Quantity", "Low Stock Threshold", "Track Stock"]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, stock in enumerate(stocks, 2):
            ws.cell(row=row_idx, column=1, value=stock.id).border = thin_border
            ws.cell(row=row_idx, column=2, value=stock.item.sku).border = thin_border
            ws.cell(row=row_idx, column=3, value=stock.item.name).border = thin_border
            ws.cell(row=row_idx, column=4, value=stock.item.category.name if stock.item.category else "").border = thin_border
            ws.cell(row=row_idx, column=5, value=stock.stall.name if stock.stall else "").border = thin_border
            ws.cell(row=row_idx, column=6, value=float(stock.quantity)).border = thin_border
            ws.cell(row=row_idx, column=7, value=float(stock.low_stock_threshold)).border = thin_border
            ws.cell(row=row_idx, column=8, value=stock.track_stock).border = thin_border

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="stall_stock_template.xlsx"'
        return response

    @action(
        detail=False,
        methods=["post"],
        url_path="bulk-preview",
        permission_classes=[IsAdminUser],
    )
    def bulk_preview(self, request):
        """
        Upload an XLSX file to preview stall stock changes before applying them.
        Returns a list of changes (old vs new) without saving anything.
        """
        import openpyxl

        xlsx_file = request.FILES.get("file")
        if not xlsx_file:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        if not xlsx_file.name.endswith((".xlsx", ".xlsm")):
            return Response({"detail": "Only .xlsx files are supported."}, status=status.HTTP_400_BAD_REQUEST)
        if xlsx_file.size > 5 * 1024 * 1024:
            return Response({"detail": "File too large. Maximum 5 MB."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        except Exception:
            return Response({"detail": "Could not parse the uploaded file."}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        if not rows:
            return Response({"detail": "The file contains no data rows."}, status=status.HTTP_400_BAD_REQUEST)

        all_stocks = {stock.id: stock for stock in Stock.objects.filter(is_deleted=False, item__is_deleted=False).select_related("item", "stall")}

        changes = []
        skipped = 0
        errors = []

        for row_num, row in enumerate(rows, 2):
            if len(row) < 8:
                errors.append({"row": row_num, "error": "Row has fewer than 8 columns."})
                continue

            try:
                stock_id = int(row[0]) if row[0] is not None else None
            except (ValueError, TypeError):
                errors.append({"row": row_num, "error": "Invalid Stock ID."})
                continue

            if stock_id is None:
                continue

            stock = all_stocks.get(stock_id)
            if not stock:
                errors.append({"row": row_num, "stock_id": stock_id, "error": "Stock ID not found."})
                continue

            try:
                new_qty = Decimal(str(row[5])) if row[5] is not None and str(row[5]).strip() != "" else None
                new_threshold = Decimal(str(row[6])) if row[6] is not None and str(row[6]).strip() != "" else None
            except (InvalidOperation, ValueError) as e:
                errors.append({"row": row_num, "stock_id": stock_id, "error": f"Invalid numeric value: {e}"})
                continue

            new_track = None
            if row[7] is not None and str(row[7]).strip() != "":
                val = str(row[7]).strip().lower()
                new_track = val in ("true", "1", "yes")

            for label, val in [("quantity", new_qty), ("threshold", new_threshold)]:
                if val is not None and val < 0:
                    errors.append({"row": row_num, "stock_id": stock_id, "error": f"{label} cannot be negative."})
                    break
            else:
                item_changes = []
                if new_qty is not None and new_qty != stock.quantity:
                    item_changes.append({"field": "Quantity", "old": str(stock.quantity), "new": str(new_qty)})
                if new_threshold is not None and new_threshold != stock.low_stock_threshold:
                    item_changes.append({"field": "Low Stock Threshold", "old": str(stock.low_stock_threshold), "new": str(new_threshold)})
                if new_track is not None and new_track != stock.track_stock:
                    item_changes.append({"field": "Track Stock", "old": str(stock.track_stock), "new": str(new_track)})

                if item_changes:
                    changes.append({
                        "row": row_num,
                        "stock_id": stock_id,
                        "sku": stock.item.sku,
                        "name": stock.item.name,
                        "stall": stock.stall.name if stock.stall else "",
                        "changes": item_changes,
                    })
                else:
                    skipped += 1

        return Response({
            "changes": changes,
            "skipped": skipped,
            "errors": errors,
            "summary": f"{len(changes)} items to update, {skipped} unchanged, {len(errors)} errors.",
        })

    @action(
        detail=False,
        methods=["post"],
        url_path="bulk-update",
        permission_classes=[IsAdminUser],
    )
    def bulk_update(self, request):
        """
        Upload an XLSX file to bulk-update stall stock quantities and thresholds.
        Validates synchronously, processes in background thread, notifies via WebSocket.
        """
        import threading

        import openpyxl

        xlsx_file = request.FILES.get("file")
        if not xlsx_file:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        if not xlsx_file.name.endswith((".xlsx", ".xlsm")):
            return Response({"detail": "Only .xlsx files are supported."}, status=status.HTTP_400_BAD_REQUEST)
        if xlsx_file.size > 5 * 1024 * 1024:
            return Response({"detail": "File too large. Maximum 5 MB."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        except Exception:
            return Response({"detail": "Could not parse the uploaded file."}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        if not rows:
            return Response({"detail": "The file contains no data rows."}, status=status.HTTP_400_BAD_REQUEST)

        user_id = request.user.id

        excluded_rows_raw = request.data.get("excluded_rows", "[]")
        try:
            import json as _json
            excluded_rows = set(_json.loads(excluded_rows_raw))
        except Exception:
            excluded_rows = set()

        def _process():
            try:
                all_stocks = {
                    stock.id: stock
                    for stock in Stock.objects.filter(is_deleted=False, item__is_deleted=False).select_related("item", "stall")
                }

                updated = []
                skipped = []
                errors = []

                for row_num, row in enumerate(rows, 2):
                    if row_num in excluded_rows:
                        skipped.append(row[0])
                        continue

                    if len(row) < 8:
                        errors.append({"row": row_num, "error": "Row has fewer than 8 columns."})
                        continue

                    try:
                        stock_id = int(row[0]) if row[0] is not None else None
                    except (ValueError, TypeError):
                        errors.append({"row": row_num, "error": "Invalid Stock ID."})
                        continue

                    if stock_id is None:
                        continue

                    stock = all_stocks.get(stock_id)
                    if not stock:
                        errors.append({"row": row_num, "stock_id": stock_id, "error": "Stock ID not found."})
                        continue

                    try:
                        new_qty = Decimal(str(row[5])) if row[5] is not None and str(row[5]).strip() != "" else None
                        new_threshold = Decimal(str(row[6])) if row[6] is not None and str(row[6]).strip() != "" else None
                    except (InvalidOperation, ValueError) as e:
                        errors.append({"row": row_num, "stock_id": stock_id, "error": f"Invalid numeric value: {e}"})
                        continue

                    new_track = None
                    if row[7] is not None and str(row[7]).strip() != "":
                        val = str(row[7]).strip().lower()
                        new_track = val in ("true", "1", "yes")

                    for label, val in [("quantity", new_qty), ("threshold", new_threshold)]:
                        if val is not None and val < 0:
                            errors.append({"row": row_num, "stock_id": stock_id, "error": f"{label} cannot be negative."})
                            break
                    else:
                        changed = False
                        if new_qty is not None and new_qty != stock.quantity:
                            stock.quantity = new_qty
                            changed = True
                        if new_threshold is not None and new_threshold != stock.low_stock_threshold:
                            stock.low_stock_threshold = new_threshold
                            changed = True
                        if new_track is not None and new_track != stock.track_stock:
                            stock.track_stock = new_track
                            changed = True

                        if changed:
                            updated.append(stock)
                        else:
                            skipped.append(stock_id)

                with transaction.atomic():
                    for stock in updated:
                        stock.save(update_fields=["quantity", "low_stock_threshold", "track_stock", "updated_at"])

                detail = f"Updated {len(updated)} stall stock entries, skipped {len(skipped)} unchanged, {len(errors)} errors."
                _notify_stall_stock_bulk_update(user_id, {
                    "updated": len(updated),
                    "skipped": len(skipped),
                    "errors": errors,
                    "detail": detail,
                })
            except Exception:
                import logging
                logging.getLogger(__name__).exception("Stall stock bulk update failed")
                _notify_stall_stock_bulk_update_failed(user_id)

        threading.Thread(target=_process, daemon=True).start()

        return Response(
            {"detail": "Bulk update started. You will be notified when it's done."},
            status=status.HTTP_202_ACCEPTED,
        )


class StockRoomStockViewSet(SoftDeleteViewSetMixin, viewsets.ModelViewSet):
    queryset = StockRoomStock.objects.select_related('item__category').all()
    serializer_class = StockRoomStockSerializer
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = StockRoomFilter
    search_fields = ["item__name"]
    ordering_fields = "__all__"

    def get_queryset(self):
        qs = super().get_queryset().filter(
            is_deleted=False,
            item__is_deleted=False,
        )
        return filter_by_date_range(self.request, qs)

    @action(detail=False, methods=["get"], url_path="filters")
    def get_filters(self, request):
        filters_config = {
            "category": {
                "options": get_product_category_options,
            },
            "status": {
                "options": get_status_options,
            },
        }

        ordering_config = [
            {"label": "Item Name", "value": "item__name"},
            {"label": "Quantity", "value": "quantity"},
            {"label": "Last Updated", "value": "updated_at"},
        ]

        return get_role_based_filter_response(request, filters_config, ordering_config)

    @action(detail=False, methods=["get"], url_path="status-counts")
    def status_counts(self, request):
        qs = self.get_queryset()
        no_stock = qs.filter(quantity=0).count()
        low_stock = qs.filter(
            quantity__gt=0,
            quantity__lte=models.F("low_stock_threshold"),
        ).count()
        high_stock = qs.filter(
            quantity__gt=models.F("low_stock_threshold"),
        ).count()
        return Response({
            "no_stock": no_stock,
            "low_stock": low_stock,
            "high_stock": high_stock,
        })

    @action(detail=True, methods=["post"], permission_classes=[IsAdminUser])
    @transaction.atomic
    def restock(self, request, pk=None):
        stock_room_stock = self.get_object()

        serializer = StockRestockSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        quantity = serializer.validated_data["quantity"]

        stock_room_stock.quantity += quantity
        stock_room_stock.save()

        return Response(
            {
                "detail": f"Stock room restocked successfully. New quantity: {stock_room_stock.quantity}."
            }
        )

    @action(detail=True, methods=["get", "post"], permission_classes=[IsAdminUser], url_path="audit")
    @transaction.atomic
    def audit(self, request, pk=None):
        """
        Stock room audit/reconciliation tool (admin only).

        GET: Returns the current stock breakdown.
        POST: Accepts physical_count and adjusts system quantity to match.
        """
        stock = self.get_object()

        breakdown = {
            "stock_id": stock.id,
            "item_name": stock.item.name,
            "item_unit": stock.item.unit_of_measure,
            "system_quantity": float(stock.quantity),
        }

        if request.method == "GET":
            return Response(breakdown)

        # POST - reconcile
        serializer = StockAuditSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        physical_count = serializer.validated_data["physical_count"]

        old_quantity = stock.quantity
        stock.quantity = physical_count
        stock.save(update_fields=["quantity", "updated_at"])

        discrepancy = float(physical_count) - float(old_quantity)

        return Response({
            **breakdown,
            "system_quantity": float(stock.quantity),
            "physical_count": float(physical_count),
            "old_quantity": float(old_quantity),
            "discrepancy": discrepancy,
            "adjusted": True,
        })

    @action(
        detail=False,
        methods=["get"],
        url_path="bulk-template",
        permission_classes=[IsAdminUser],
    )
    def bulk_template(self, request):
        """Download an XLSX template pre-filled with all active stockroom stock entries."""
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        stocks = (
            StockRoomStock.objects.filter(is_deleted=False, item__is_deleted=False)
            .select_related("item__category")
            .order_by("item__name")
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stockroom Stock"

        headers = ["Stock ID", "SKU", "Item Name", "Category", "Quantity", "Low Stock Threshold"]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, stock in enumerate(stocks, 2):
            ws.cell(row=row_idx, column=1, value=stock.id).border = thin_border
            ws.cell(row=row_idx, column=2, value=stock.item.sku).border = thin_border
            ws.cell(row=row_idx, column=3, value=stock.item.name).border = thin_border
            ws.cell(row=row_idx, column=4, value=stock.item.category.name if stock.item.category else "").border = thin_border
            ws.cell(row=row_idx, column=5, value=float(stock.quantity)).border = thin_border
            ws.cell(row=row_idx, column=6, value=float(stock.low_stock_threshold)).border = thin_border

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="stockroom_stock_template.xlsx"'
        return response

    @action(
        detail=False,
        methods=["post"],
        url_path="bulk-preview",
        permission_classes=[IsAdminUser],
    )
    def bulk_preview(self, request):
        """
        Upload an XLSX file to preview stockroom stock changes before applying them.
        Returns a list of changes (old vs new) without saving anything.
        """
        import openpyxl

        xlsx_file = request.FILES.get("file")
        if not xlsx_file:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        if not xlsx_file.name.endswith((".xlsx", ".xlsm")):
            return Response({"detail": "Only .xlsx files are supported."}, status=status.HTTP_400_BAD_REQUEST)
        if xlsx_file.size > 5 * 1024 * 1024:
            return Response({"detail": "File too large. Maximum 5 MB."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        except Exception:
            return Response({"detail": "Could not parse the uploaded file."}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        if not rows:
            return Response({"detail": "The file contains no data rows."}, status=status.HTTP_400_BAD_REQUEST)

        all_stocks = {
            stock.id: stock
            for stock in StockRoomStock.objects.filter(is_deleted=False, item__is_deleted=False).select_related("item")
        }

        changes = []
        skipped = 0
        errors = []

        for row_num, row in enumerate(rows, 2):
            if len(row) < 6:
                errors.append({"row": row_num, "error": "Row has fewer than 6 columns."})
                continue

            try:
                stock_id = int(row[0]) if row[0] is not None else None
            except (ValueError, TypeError):
                errors.append({"row": row_num, "error": "Invalid Stock ID."})
                continue

            if stock_id is None:
                continue

            stock = all_stocks.get(stock_id)
            if not stock:
                errors.append({"row": row_num, "stock_id": stock_id, "error": "Stock ID not found."})
                continue

            try:
                new_qty = Decimal(str(row[4])) if row[4] is not None and str(row[4]).strip() != "" else None
                new_threshold = Decimal(str(row[5])) if row[5] is not None and str(row[5]).strip() != "" else None
            except (InvalidOperation, ValueError) as e:
                errors.append({"row": row_num, "stock_id": stock_id, "error": f"Invalid numeric value: {e}"})
                continue

            for label, val in [("quantity", new_qty), ("threshold", new_threshold)]:
                if val is not None and val < 0:
                    errors.append({"row": row_num, "stock_id": stock_id, "error": f"{label} cannot be negative."})
                    break
            else:
                item_changes = []
                if new_qty is not None and new_qty != stock.quantity:
                    item_changes.append({"field": "Quantity", "old": str(stock.quantity), "new": str(new_qty)})
                if new_threshold is not None and new_threshold != stock.low_stock_threshold:
                    item_changes.append({"field": "Low Stock Threshold", "old": str(stock.low_stock_threshold), "new": str(new_threshold)})

                if item_changes:
                    changes.append({
                        "row": row_num,
                        "stock_id": stock_id,
                        "sku": stock.item.sku,
                        "name": stock.item.name,
                        "changes": item_changes,
                    })
                else:
                    skipped += 1

        return Response({
            "changes": changes,
            "skipped": skipped,
            "errors": errors,
            "summary": f"{len(changes)} items to update, {skipped} unchanged, {len(errors)} errors.",
        })

    @action(
        detail=False,
        methods=["post"],
        url_path="bulk-update",
        permission_classes=[IsAdminUser],
    )
    def bulk_update(self, request):
        """
        Upload an XLSX file to bulk-update stockroom stock quantities and thresholds.
        Validates synchronously, processes in background thread, notifies via WebSocket.
        """
        import threading

        import openpyxl

        xlsx_file = request.FILES.get("file")
        if not xlsx_file:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        if not xlsx_file.name.endswith((".xlsx", ".xlsm")):
            return Response({"detail": "Only .xlsx files are supported."}, status=status.HTTP_400_BAD_REQUEST)
        if xlsx_file.size > 5 * 1024 * 1024:
            return Response({"detail": "File too large. Maximum 5 MB."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        except Exception:
            return Response({"detail": "Could not parse the uploaded file."}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        if not rows:
            return Response({"detail": "The file contains no data rows."}, status=status.HTTP_400_BAD_REQUEST)

        user_id = request.user.id

        excluded_rows_raw = request.data.get("excluded_rows", "[]")
        try:
            import json as _json
            excluded_rows = set(_json.loads(excluded_rows_raw))
        except Exception:
            excluded_rows = set()

        def _process():
            try:
                all_stocks = {
                    stock.id: stock
                    for stock in StockRoomStock.objects.filter(is_deleted=False, item__is_deleted=False).select_related("item")
                }

                updated = []
                skipped = []
                errors = []

                for row_num, row in enumerate(rows, 2):
                    if row_num in excluded_rows:
                        skipped.append(row[0])
                        continue

                    if len(row) < 6:
                        errors.append({"row": row_num, "error": "Row has fewer than 6 columns."})
                        continue

                    try:
                        stock_id = int(row[0]) if row[0] is not None else None
                    except (ValueError, TypeError):
                        errors.append({"row": row_num, "error": "Invalid Stock ID."})
                        continue

                    if stock_id is None:
                        continue

                    stock = all_stocks.get(stock_id)
                    if not stock:
                        errors.append({"row": row_num, "stock_id": stock_id, "error": "Stock ID not found."})
                        continue

                    try:
                        new_qty = Decimal(str(row[4])) if row[4] is not None and str(row[4]).strip() != "" else None
                        new_threshold = Decimal(str(row[5])) if row[5] is not None and str(row[5]).strip() != "" else None
                    except (InvalidOperation, ValueError) as e:
                        errors.append({"row": row_num, "stock_id": stock_id, "error": f"Invalid numeric value: {e}"})
                        continue

                    for label, val in [("quantity", new_qty), ("threshold", new_threshold)]:
                        if val is not None and val < 0:
                            errors.append({"row": row_num, "stock_id": stock_id, "error": f"{label} cannot be negative."})
                            break
                    else:
                        changed = False
                        if new_qty is not None and new_qty != stock.quantity:
                            stock.quantity = new_qty
                            changed = True
                        if new_threshold is not None and new_threshold != stock.low_stock_threshold:
                            stock.low_stock_threshold = new_threshold
                            changed = True

                        if changed:
                            updated.append(stock)
                        else:
                            skipped.append(stock_id)

                with transaction.atomic():
                    for stock in updated:
                        stock.save(update_fields=["quantity", "low_stock_threshold", "updated_at"])

                detail = f"Updated {len(updated)} stockroom entries, skipped {len(skipped)} unchanged, {len(errors)} errors."
                _notify_stockroom_bulk_update(user_id, {
                    "updated": len(updated),
                    "skipped": len(skipped),
                    "errors": errors,
                    "detail": detail,
                })
            except Exception:
                import logging
                logging.getLogger(__name__).exception("Stockroom bulk update failed")
                _notify_stockroom_bulk_update_failed(user_id)

        threading.Thread(target=_process, daemon=True).start()

        return Response(
            {"detail": "Bulk update started. You will be notified when it's done."},
            status=status.HTTP_202_ACCEPTED,
        )


class ProductCategoryViewSet(SoftDeleteViewSetMixin, viewsets.ModelViewSet):
    queryset = ProductCategory.objects.all()
    serializer_class = ProductCategorySerializer
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["name"]
    search_fields = ["name"]
    ordering_fields = "__all__"

    def get_queryset(self):
        qs = super().get_queryset().filter(is_deleted=False)
        return filter_by_date_range(self.request, qs)


class StockRequestViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for managing stock requests.

    Clerks create stock requests implicitly when adding items with insufficient
    stock. Admins can list, approve, decline, or batch-approve requests.
    """

    queryset = StockRequest.objects.select_related(
        "item", "stall", "service", "requested_by", "approved_by",
        "appliance_item", "service_item",
    ).all()
    serializer_class = StockRequestSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["status", "source"]
    search_fields = ["item__name", "item__sku", "notes"]
    ordering_fields = ["created_at", "requested_quantity", "status"]

    def get_queryset(self):
        qs = super().get_queryset()
        if self.request.user.role != "admin":
            qs = qs.filter(requested_by=self.request.user)
        return qs

    @action(detail=False, methods=["get"], url_path="pending-count")
    def pending_count(self, request):
        count = self.get_queryset().filter(status="pending").count()
        return Response({"count": count})

    @action(detail=True, methods=["post"], permission_classes=[IsAdminUser])
    @transaction.atomic
    def approve(self, request, pk=None):
        """Approve a stock request: add stock and reserve for the service item."""
        from services.business_logic import StockReservationManager
        from django.db import transaction as db_transaction

        with db_transaction.atomic():
            # Lock the stock request row to prevent double-approval
            stock_request = StockRequest.objects.select_for_update().get(pk=pk)

            if stock_request.status != "pending":
                return Response(
                    {"detail": f"Cannot approve a request with status '{stock_request.status}'."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            stock_record, _ = Stock.objects.get_or_create(
                item=stock_request.item,
                stall=stock_request.stall,
                is_deleted=False,
                defaults={"quantity": 0, "reserved_quantity": 0},
            )
            stock_record = Stock.objects.select_for_update().get(pk=stock_record.pk)
            stock_record.quantity += stock_request.requested_quantity
            stock_record.save(update_fields=["quantity", "updated_at"])

            item_usage = stock_request.appliance_item or stock_request.service_item
            if item_usage and item_usage.item:
                try:
                    StockReservationManager.reserve_stock(
                        item=item_usage.item,
                        quantity=item_usage.quantity,
                        stall_stock=stock_record,
                    )
                    if not item_usage.stall_stock_id:
                        item_usage.stall_stock = stock_record
                    item_usage.stock_request_status = "approved"
                    item_usage.save(update_fields=["stall_stock", "stock_request_status"])
                except Exception:
                    if item_usage.stock_request_status == "pending":
                        item_usage.stock_request_status = "approved"
                        item_usage.save(update_fields=["stock_request_status"])

            stock_request.status = "approved"
            stock_request.approved_by = request.user
            stock_request.approved_at = timezone.now()
            stock_request.save(update_fields=[
                "status", "approved_by", "approved_at", "updated_at",
            ])

            if stock_request.requested_by and stock_request.requested_by.is_active:
                Notification.objects.create(
                    user=stock_request.requested_by,
                    type="stock_request_approved",
                    title="Stock Request Approved",
                    message=(
                        f"Your stock request for {stock_request.requested_quantity} "
                        f"{stock_request.item.unit_of_measure} of '{stock_request.item.name}' "
                        f"has been approved."
                    ),
                    data={
                        "stock_request_id": stock_request.id,
                        "item_name": stock_request.item.name,
                        "quantity": float(stock_request.requested_quantity),
                        "service_id": stock_request.service_id,
                    },
                )

        return Response(StockRequestSerializer(stock_request).data)

    @action(detail=True, methods=["post"], permission_classes=[IsAdminUser])
    @transaction.atomic
    def decline(self, request, pk=None):
        """Decline a stock request."""
        stock_request = self.get_object()

        if stock_request.status != "pending":
            return Response(
                {"detail": f"Cannot decline a request with status '{stock_request.status}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        decline_reason = request.data.get("reason", "")

        item_usage = stock_request.appliance_item or stock_request.service_item
        if item_usage:
            item_usage.stock_request_status = "declined"
            item_usage.save(update_fields=["stock_request_status"])

        stock_request.status = "declined"
        stock_request.decline_reason = decline_reason
        stock_request.declined_at = timezone.now()
        stock_request.save(update_fields=[
            "status", "decline_reason", "declined_at", "updated_at",
        ])

        if stock_request.requested_by and stock_request.requested_by.is_active:
            msg = (
                f"Your stock request for {stock_request.requested_quantity} "
                f"{stock_request.item.unit_of_measure} of '{stock_request.item.name}' "
                f"has been declined."
            )
            if decline_reason:
                msg += f" Reason: {decline_reason}"

            Notification.objects.create(
                user=stock_request.requested_by,
                type="stock_request_declined",
                title="Stock Request Declined",
                message=msg,
                data={
                    "stock_request_id": stock_request.id,
                    "item_name": stock_request.item.name,
                    "service_id": stock_request.service_id,
                },
            )

        return Response(StockRequestSerializer(stock_request).data)

    @action(detail=False, methods=["post"], permission_classes=[IsAdminUser], url_path="batch-approve")
    @transaction.atomic
    def batch_approve(self, request):
        """Approve multiple stock requests at once."""
        from services.business_logic import StockReservationManager

        ids = request.data.get("ids", [])
        if not ids:
            return Response(
                {"detail": "No request IDs provided."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        requests_qs = StockRequest.objects.select_for_update().filter(
            id__in=ids, status="pending"
        ).select_related("item", "stall", "appliance_item", "service_item", "requested_by")

        approved_count = 0
        for sr in requests_qs:
            stock_record, _ = Stock.objects.get_or_create(
                item=sr.item, stall=sr.stall, is_deleted=False,
                defaults={"quantity": 0, "reserved_quantity": 0},
            )
            stock_record = Stock.objects.select_for_update().get(pk=stock_record.pk)
            stock_record.quantity += sr.requested_quantity
            stock_record.save(update_fields=["quantity", "updated_at"])

            item_usage = sr.appliance_item or sr.service_item
            if item_usage and item_usage.item:
                try:
                    StockReservationManager.reserve_stock(
                        item=item_usage.item,
                        quantity=item_usage.quantity,
                        stall_stock=stock_record,
                    )
                    if not item_usage.stall_stock_id:
                        item_usage.stall_stock = stock_record
                    item_usage.stock_request_status = "approved"
                    item_usage.save(update_fields=["stall_stock", "stock_request_status"])
                except Exception:
                    if item_usage.stock_request_status == "pending":
                        item_usage.stock_request_status = "approved"
                        item_usage.save(update_fields=["stock_request_status"])

            sr.status = "approved"
            sr.approved_by = request.user
            sr.approved_at = timezone.now()
            sr.save(update_fields=["status", "approved_by", "approved_at", "updated_at"])
            approved_count += 1

            if sr.requested_by and sr.requested_by.is_active:
                Notification.objects.create(
                    user=sr.requested_by,
                    type="stock_request_approved",
                    title="Stock Request Approved",
                    message=(
                        f"Your stock request for {sr.requested_quantity} "
                        f"{sr.item.unit_of_measure} of '{sr.item.name}' "
                        f"has been approved."
                    ),
                    data={
                        "stock_request_id": sr.id,
                        "item_name": sr.item.name,
                        "quantity": float(sr.requested_quantity),
                        "service_id": sr.service_id,
                    },
                )

        return Response({"approved_count": approved_count})


class CustomItemTemplateViewSet(viewsets.ModelViewSet):
    """
    CRUD for custom item templates. Admin-only create/update/delete.
    All authenticated users can list/retrieve active templates.
    """

    serializer_class = CustomItemTemplateSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name", "description"]
    ordering_fields = ["name", "default_price", "created_at"]

    def get_queryset(self):
        qs = CustomItemTemplate.objects.select_related("created_by")
        # Non-admins only see active templates
        if getattr(self.request.user, "role", None) != "admin":
            qs = qs.filter(is_active=True)
        is_active = self.request.query_params.get("is_active")
        if is_active is not None:
            qs = qs.filter(is_active=str(is_active).lower() in ("1", "true", "yes"))
        return qs

    def get_permissions(self):
        if self.action in ("create", "update", "partial_update", "destroy"):
            return [IsAdminUser()]
        return [IsAuthenticated()]
