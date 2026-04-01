from clients.api.filters import ClientFilter
from clients.api.serializers import ClientSerializer
from clients.models import Client
from django.db import models
from django.db.models import functions
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, permissions, serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from utils.filters.role_filters import get_role_based_filter_response
from utils.query import filter_by_date_range
from utils.soft_delete import SoftDeleteViewSetMixin


class ClientViewSet(SoftDeleteViewSetMixin, viewsets.ModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = ClientFilter
    search_fields = [
        "full_name",
        "contact_number",
        "province",
        "city",
        "barangay",
        "address",
    ]
    ordering_fields = "__all__"

    def get_queryset(self):
        # Apply role/date-based filtering
        qs = super().get_queryset()
        qs = qs.filter(is_deleted=False)
        return filter_by_date_range(self.request, qs)

    @action(detail=False, methods=["get"], url_path="filters")
    def get_filters(self, request):
        filters_config = {
            "is_blocklisted": {
                "options": lambda: [
                    {"label": "Blocklisted", "value": "true"},
                    {"label": "Not Blocklisted", "value": "false"},
                ],
            },
            "is_deleted": {
                "options": lambda: [
                    {"label": "Deleted", "value": "true"},
                    {"label": "Not Deleted", "value": "false"},
                ],
            },
            "province": {
                "options": lambda: Client.objects.values_list("province", flat=True)
                .distinct()
                .order_by("province"),
            },
            "city": {
                "options": lambda: Client.objects.values_list("city", flat=True)
                .distinct()
                .order_by("city"),
            },
        }
        ordering_config = [
            {"label": "Full Name", "value": "full_name"},
            {"label": "Created Date", "value": "created_at"},
            {"label": "City", "value": "city"},
            {"label": "Province", "value": "province"},
        ]
        return get_role_based_filter_response(request, filters_config, ordering_config)

    def perform_create(self, serializer):
        full_name = serializer.validated_data.get("full_name")
        contact_number = serializer.validated_data.get("contact_number") or None

        if Client.objects.filter(
            full_name=full_name, contact_number=contact_number
        ).exists():
            raise serializers.ValidationError(
                {
                    "non_field_errors": [
                        "A client with this full name and contact number already exists."
                    ]
                }
            )

        if (
            contact_number
            and Client.objects.filter(contact_number=contact_number).exists()
        ):
            raise serializers.ValidationError(
                {
                    "non_field_errors": [
                        "A client with this contact number already exists."
                    ]
                }
            )

        serializer.save(contact_number=contact_number)

    @action(detail=False, methods=["get"], url_path="recent")
    def recent_clients(self, request):
        """
        Return clients from recent sales transactions and services,
        ordered by most recent interaction.
        """
        from django.db.models import Max

        limit = int(request.query_params.get("limit", 10))

        # Get clients with recent sales or services, ordered by latest activity
        recent = (
            Client.objects.filter(is_deleted=False)
            .annotate(
                latest_sale=Max("salestransaction__created_at"),
                latest_service=Max("services__created_at"),
            )
            .filter(
                models.Q(latest_sale__isnull=False)
                | models.Q(latest_service__isnull=False)
            )
            .order_by(
                models.functions.Coalesce("latest_sale", "latest_service").desc()
            )[:limit]
        )

        serializer = self.get_serializer(recent, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="toggle-favorite")
    def toggle_favorite(self, request, pk=None):
        """Toggle the is_favorite flag on a client."""
        client = self.get_object()
        client.is_favorite = not client.is_favorite
        client.save(update_fields=["is_favorite", "updated_at"])
        return Response(
            {"is_favorite": client.is_favorite},
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"], url_path="favorites")
    def favorite_clients(self, request):
        """Return all favorite clients."""
        favorites = Client.objects.filter(is_deleted=False, is_favorite=True).order_by("full_name")
        serializer = self.get_serializer(favorites, many=True)
        return Response(serializer.data)

    @action(
        detail=False,
        methods=["get"],
        url_path="bulk-template",
        permission_classes=[permissions.IsAdminUser],
    )
    def bulk_template(self, request):
        """Download an XLSX file pre-filled with all active clients."""
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        clients = Client.objects.filter(is_deleted=False).order_by("full_name")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clients"

        headers = [
            "ID",
            "Full Name",
            "Contact Number",
            "Province",
            "City",
            "Barangay",
            "Address",
        ]
        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, client in enumerate(clients, 2):
            ws.cell(row=row_idx, column=1, value=client.id).border = thin_border
            ws.cell(
                row=row_idx, column=2, value=client.full_name
            ).border = thin_border
            ws.cell(
                row=row_idx, column=3, value=client.contact_number or ""
            ).border = thin_border
            ws.cell(
                row=row_idx, column=4, value=client.province
            ).border = thin_border
            ws.cell(row=row_idx, column=5, value=client.city).border = thin_border
            ws.cell(
                row=row_idx, column=6, value=client.barangay or ""
            ).border = thin_border
            ws.cell(
                row=row_idx, column=7, value=client.address or ""
            ).border = thin_border

        ws.sheet_properties.tabColor = "1F4E79"

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        from io import BytesIO

        from django.http import HttpResponse

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = (
            'attachment; filename="client_update_template.xlsx"'
        )
        return resp

    @action(
        detail=False,
        methods=["post"],
        url_path="bulk-preview",
        permission_classes=[permissions.IsAdminUser],
    )
    def bulk_preview(self, request):
        """
        Upload an XLSX to preview client changes before applying.
        Returns a diff of proposed changes.
        """
        import openpyxl

        xlsx_file = request.FILES.get("file")
        if not xlsx_file:
            return Response(
                {"detail": "No file uploaded."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not xlsx_file.name.endswith((".xlsx", ".xlsm")):
            return Response(
                {"detail": "Only .xlsx files are supported."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        except Exception:
            return Response(
                {"detail": "Could not parse the uploaded file."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        if not rows:
            return Response(
                {"detail": "The file contains no data rows."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        all_clients = {c.id: c for c in Client.objects.filter(is_deleted=False)}

        changes = []
        skipped = 0
        errors = []

        for row_num, row in enumerate(rows, 2):
            if len(row) < 7:
                errors.append(
                    {"row": row_num, "error": "Row has fewer than 7 columns."}
                )
                continue

            try:
                client_id = int(row[0]) if row[0] is not None else None
            except (ValueError, TypeError):
                errors.append(
                    {
                        "row": row_num,
                        "error": f"Invalid ID: {row[0]}",
                    }
                )
                continue

            if not client_id:
                continue

            client = all_clients.get(client_id)
            if not client:
                errors.append(
                    {"row": row_num, "sku": str(client_id), "error": "Client ID not found."}
                )
                continue

            new_name = str(row[1] or "").strip()
            new_contact = str(row[2] or "").strip() or None
            new_province = str(row[3] or "").strip()
            new_city = str(row[4] or "").strip()
            new_barangay = str(row[5] or "").strip() or None
            new_address = str(row[6] or "").strip() or None

            client_changes = []
            if new_name and new_name != client.full_name:
                client_changes.append(
                    {"field": "Full Name", "old": client.full_name, "new": new_name}
                )
            if new_contact != (client.contact_number or None):
                client_changes.append(
                    {
                        "field": "Contact Number",
                        "old": client.contact_number or "",
                        "new": new_contact or "",
                    }
                )
            if new_province and new_province != client.province:
                client_changes.append(
                    {"field": "Province", "old": client.province, "new": new_province}
                )
            if new_city and new_city != client.city:
                client_changes.append(
                    {"field": "City", "old": client.city, "new": new_city}
                )
            if new_barangay != (client.barangay or None):
                client_changes.append(
                    {
                        "field": "Barangay",
                        "old": client.barangay or "",
                        "new": new_barangay or "",
                    }
                )
            if new_address != (client.address or None):
                client_changes.append(
                    {
                        "field": "Address",
                        "old": client.address or "",
                        "new": new_address or "",
                    }
                )

            if client_changes:
                changes.append(
                    {
                        "row": row_num,
                        "sku": str(client_id),
                        "name": client.full_name,
                        "changes": client_changes,
                    }
                )
            else:
                skipped += 1

        return Response(
            {
                "changes": changes,
                "skipped": skipped,
                "errors": errors,
                "summary": f"{len(changes)} clients to update, {skipped} unchanged, {len(errors)} errors.",
            }
        )

    @action(
        detail=False,
        methods=["post"],
        url_path="bulk-update",
        permission_classes=[permissions.IsAdminUser],
    )
    def bulk_update(self, request):
        """
        Upload an XLSX file to bulk-update client records.
        Validates synchronously, then processes in a background thread.
        """
        import threading

        import openpyxl

        xlsx_file = request.FILES.get("file")
        if not xlsx_file:
            return Response(
                {"detail": "No file uploaded."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not xlsx_file.name.endswith((".xlsx", ".xlsm")):
            return Response(
                {"detail": "Only .xlsx files are supported."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if xlsx_file.size > 5 * 1024 * 1024:
            return Response(
                {"detail": "File too large. Maximum 5 MB."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        except Exception:
            return Response(
                {"detail": "Could not parse the uploaded file."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        if not rows:
            return Response(
                {"detail": "The file contains no data rows."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user_id = request.user.id

        def _process_bulk_update():
            from django.db import transaction as db_transaction

            try:
                all_clients = {
                    c.id: c for c in Client.objects.filter(is_deleted=False)
                }

                updated = []
                skipped = []
                errors = []

                for row_num, row in enumerate(rows, 2):
                    if len(row) < 7:
                        errors.append(
                            {"row": row_num, "error": "Row has fewer than 7 columns."}
                        )
                        continue

                    try:
                        client_id = int(row[0]) if row[0] is not None else None
                    except (ValueError, TypeError):
                        errors.append(
                            {"row": row_num, "error": f"Invalid ID: {row[0]}"}
                        )
                        continue

                    if not client_id:
                        continue

                    client = all_clients.get(client_id)
                    if not client:
                        errors.append(
                            {
                                "row": row_num,
                                "sku": str(client_id),
                                "error": "Client ID not found.",
                            }
                        )
                        continue

                    new_name = str(row[1] or "").strip()
                    new_contact = str(row[2] or "").strip() or None
                    new_province = str(row[3] or "").strip()
                    new_city = str(row[4] or "").strip()
                    new_barangay = str(row[5] or "").strip() or None
                    new_address = str(row[6] or "").strip() or None

                    changed = False
                    if new_name and new_name != client.full_name:
                        client.full_name = new_name
                        changed = True
                    if new_contact != (client.contact_number or None):
                        client.contact_number = new_contact
                        changed = True
                    if new_province and new_province != client.province:
                        client.province = new_province
                        changed = True
                    if new_city and new_city != client.city:
                        client.city = new_city
                        changed = True
                    if new_barangay != (client.barangay or None):
                        client.barangay = new_barangay
                        changed = True
                    if new_address != (client.address or None):
                        client.address = new_address
                        changed = True

                    if changed:
                        updated.append(client)
                    else:
                        skipped.append(str(client_id))

                with db_transaction.atomic():
                    for client in updated:
                        client.save()

                detail = f"Updated {len(updated)} clients, skipped {len(skipped)} unchanged, {len(errors)} errors."
                _notify_client_bulk_update(
                    user_id,
                    {
                        "updated": len(updated),
                        "skipped": len(skipped),
                        "errors": errors,
                        "detail": detail,
                    },
                )
            except Exception:
                import logging

                logging.getLogger(__name__).exception("Client bulk update failed")
                _notify_client_bulk_update_failed(user_id)

        threading.Thread(target=_process_bulk_update, daemon=True).start()

        return Response(
            {"detail": "Bulk update started. You will be notified when it's done."},
            status=status.HTTP_202_ACCEPTED,
        )


def _notify_client_bulk_update(user_id, result):
    """Push client_bulk_update_complete event via WebSocket."""
    try:
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer

        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f"notifications_{user_id}",
                {
                    "type": "send_notification",
                    "data": {
                        "event": "export_ready",
                        "export_type": "client_bulk_update",
                        "title": "Client Bulk Update Complete",
                        "message": result["detail"],
                        "result": result,
                    },
                },
            )
    except Exception:
        import logging

        logging.getLogger(__name__).exception(
            "Failed to send client_bulk_update via WebSocket"
        )


def _notify_client_bulk_update_failed(user_id):
    """Push client_bulk_update_failed event via WebSocket."""
    try:
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer

        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f"notifications_{user_id}",
                {
                    "type": "send_notification",
                    "data": {
                        "event": "export_failed",
                        "export_type": "client_bulk_update",
                        "title": "Client Bulk Update Failed",
                        "message": "Failed to process the bulk update. Please try again.",
                    },
                },
            )
    except Exception:
        import logging

        logging.getLogger(__name__).exception(
            "Failed to send client_bulk_update_failed via WebSocket"
        )
