from rest_framework import generics, parsers, permissions, filters, viewsets, status
from rest_framework.decorators import action
from rest_framework.exceptions import NotFound
from rest_framework.response import Response
from rest_framework.views import APIView
from utils.query import filter_by_date_range
from django.utils import timezone
from users.models import CustomUser, SystemSettings, CashAdvanceMovement
from users.api.serializers import EmployeesSerializer, UserSerializer, SystemSettingsSerializer, CashAdvanceMovementSerializer
from django_filters.rest_framework import DjangoFilterBackend


# List all users (admin only)
class UserListView(generics.ListAPIView):
    queryset = CustomUser.all_objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAdminUser]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]

    filterset_fields = ["username", "contact_number", "first_name", "last_name", "role"]
    search_fields = ["username", "contact_number", "first_name", "last_name"]
    ordering_fields = "__all__"

    def get_queryset(self):
        return filter_by_date_range(self.request, super().get_queryset())


# Admin: view, update, or soft delete any specific user
class AdminUserDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = CustomUser.all_objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAdminUser]

    def get_object(self):
        try:
            user = self.queryset.get(pk=self.kwargs["pk"])
            if user.is_deleted:
                raise NotFound(
                    detail="User not found."
                )  # Prevent accessing soft-deleted users
            return user
        except CustomUser.DoesNotExist:
            raise NotFound(detail="User not found.")

    def perform_destroy(self, instance):
        instance.is_deleted = True
        instance.is_active = False
        instance.deleted_at = timezone.now()
        instance.save(update_fields=["is_deleted", "is_active", "deleted_at"])

    def get_serializer_context(self):
        return {"request": self.request}


class EmployeesListView(generics.ListCreateAPIView):
    serializer_class = EmployeesSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = CustomUser.objects.exclude(role="admin").filter(is_deleted=False)
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = [
        "username",
        "contact_number",
        "first_name",
        "last_name",
        "email",
        "address",
        "province",
        "city",
        "barangay",
        "role",
    ]
    search_fields = [
        "username",
        "contact_number",
        "first_name",
        "last_name",
        "email",
        "address",
        "province",
        "city",
        "barangay",
    ]
    ordering_fields = "__all__"

    def get_queryset(self):
        qs = CustomUser.objects.filter(is_deleted=False)
        if self.request.query_params.get("include_admins") != "true":
            qs = qs.exclude(role="admin")
        return filter_by_date_range(self.request, qs)


class UseraDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = EmployeesSerializer
    queryset = CustomUser.objects.filter(is_deleted=False)
    permission_classes = [permissions.IsAuthenticated]

    def filter_queryset(self, queryset):
        return queryset

    def perform_destroy(self, instance):
        instance.is_deleted = True
        instance.is_active = False
        instance.deleted_at = timezone.now()
        instance.save(update_fields=["is_deleted", "is_active", "deleted_at"])

    def get_serializer_context(self):
        return {"request": self.request}


class EmployeeArchivedListView(generics.ListAPIView):
    """List all archived (soft-deleted) employees."""
    serializer_class = EmployeesSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = [
        "username", "contact_number", "first_name", "last_name",
        "email", "address", "province", "city", "barangay",
    ]
    search_fields = [
        "username", "contact_number", "first_name", "last_name",
        "email", "address", "province", "city", "barangay",
    ]
    ordering_fields = "__all__"

    def get_queryset(self):
        return CustomUser.all_objects.exclude(role="admin").filter(is_deleted=True)


class EmployeeRestoreView(APIView):
    """Restore an archived employee."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        from django.shortcuts import get_object_or_404
        instance = get_object_or_404(CustomUser.all_objects.all(), pk=pk)
        if not instance.is_deleted:
            return Response(
                {"detail": "This record is not archived."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        instance.is_deleted = False
        instance.is_active = True
        instance.deleted_at = None
        instance.save(update_fields=["is_deleted", "is_active", "deleted_at"])
        serializer = EmployeesSerializer(instance, context={"request": request})
        return Response(serializer.data)


class MyProfileView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        if self.request.user.is_deleted:
            raise NotFound(detail="User not found.")
        return self.request.user

    def perform_destroy(self, instance):
        instance.is_deleted = True
        instance.save()

    def get_serializer_context(self):
        return {"request": self.request}


class SystemSettingsView(generics.RetrieveUpdateAPIView):
    """
    GET: Retrieve system settings (any authenticated user can view)
    PUT/PATCH: Update system settings (admin only)
    """
    serializer_class = SystemSettingsSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        """Always return the singleton settings instance"""
        return SystemSettings.get_settings()

    def get_permissions(self):
        """Allow any authenticated user to view, but restrict updates by field sensitivity"""
        if self.request.method in ['PUT', 'PATCH']:
            sensitive_fields = {'maintenance_mode', 'check_stock_on_sale'}
            if sensitive_fields & set(self.request.data.keys()):
                return [IsSuperAdminUser()]
            return [IsAdminOrManager()]
        return [permissions.IsAuthenticated()]


class CashAdvanceMovementViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing cash ban balance movements.
    - List: View all movements (admin/manager) or own movements (employee)
    - Create: Record a new movement — credit (+) or debit (-) (admin/manager only)
    - Retrieve/Delete: Manage specific movement (admin/manager only)
    """
    serializer_class = CashAdvanceMovementSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ['employee', 'date', 'movement_type']
    search_fields = ['employee__first_name', 'employee__last_name', 'description', 'reference']
    ordering_fields = ['date', 'amount', 'created_at']
    ordering = ['-date', '-created_at']

    def get_queryset(self):
        """
        Admin/Manager: See all movements
        Other users: See only their own movements
        """
        queryset = CashAdvanceMovement.objects.filter(is_deleted=False).select_related(
            'employee', 'created_by'
        )

        user = self.request.user
        if user.role in ['admin', 'manager']:
            return filter_by_date_range(self.request, queryset)
        else:
            return queryset.filter(employee=user)

    def get_permissions(self):
        """Only admin/manager can create, update, or delete"""
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated(), IsAdminOrManager()]
        return [permissions.IsAuthenticated()]

    def perform_destroy(self, instance):
        """Soft delete and reverse the balance change (only if movement was already applied)"""
        instance.is_deleted = True
        instance.save(update_fields=['is_deleted'])
        # Only reverse the balance change if the movement was already applied (not pending)
        if not instance.is_pending:
            if instance.movement_type == CashAdvanceMovement.MovementType.CREDIT:
                instance.employee.cash_ban_balance -= instance.amount
            else:
                instance.employee.cash_ban_balance += instance.amount
            instance.employee.save(update_fields=['cash_ban_balance'])


class IsAdminOrManager(permissions.BasePermission):
    """Permission class to check if user is admin or manager."""
    def has_permission(self, request, view):
        return request.user.is_authenticated and (
            request.user.role in ['admin', 'manager']
        )


class IsSuperAdminUser(permissions.BasePermission):
    """Permission class for Django superusers only."""
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.is_superuser


class ServerMaintenanceView(APIView):
    """
    Admin-only endpoint for server maintenance tasks.
    GET  — Disk, memory, Docker stats, container info, cron jobs, management commands
    POST — Cleanup actions, restart containers, view logs, run management commands
    """
    permission_classes = [IsSuperAdminUser]

    # Defined cron schedule (runs on host, not in container)
    CRON_JOBS = [
        {"id": "auto_close_attendance", "schedule": "Daily 9:00 PM", "description": "Auto-close open attendance sessions for employees who forgot to clock out", "log_file": "cron-auto-close-attendance.log", "category": "attendance"},
        {"id": "mark_daily_absences", "schedule": "Daily 11:30 PM", "description": "Mark employees absent if no clock-in/out and no approved leave", "log_file": "cron-mark-absences.log", "category": "attendance"},
        {"id": "delete_old_notifications", "schedule": "Daily 2:00 AM", "description": "Delete read notifications older than 7 days", "log_file": "cron-delete-old-notifications.log", "category": "maintenance"},
        {"id": "cleanup_archived_quotations", "schedule": "Daily 3:00 AM", "description": "Delete archived quotations older than 14 days", "log_file": "cron-cleanup-archived-quotations.log", "category": "maintenance"},
        {"id": "truncate_large_logs", "schedule": "Daily 5:00 AM", "description": "Truncate cron log files over 10MB, keep last 500 lines", "log_file": "cron-log-maintenance.log", "category": "maintenance"},
        {"id": "disk_usage_check", "schedule": "Daily 6:00 AM", "description": "Monitor disk usage, auto-cleanup at 90%+ (apt cache, journal, Docker)", "log_file": "cron-disk-usage.log", "category": "maintenance"},
        {"id": "weekly_attendance_payroll_fix", "schedule": "Friday 11:00 PM", "description": "Verify and fix attendance for the past week, then refresh payroll for Saturday payday", "log_file": "cron-weekly-attendance-payroll-fix.log", "category": "payroll"},
        {"id": "docker_system_prune", "schedule": "Sunday 3:00 AM", "description": "Remove unused containers, networks, and dangling images", "log_file": "cron-docker-prune.log", "category": "maintenance"},
        {"id": "cleanup_unused_images", "schedule": "Sunday 3:30 AM", "description": "Delete unused user profile images not linked to any employee", "log_file": "cron-cleanup-images.log", "category": "maintenance"},
        {"id": "update_holiday_years", "schedule": "Jan 1, 2:00 AM", "description": "Shift all holidays to the next calendar year", "log_file": "cron-yearly-update-holidays.log", "category": "payroll"},
        {"id": "replenish_leave_balances", "schedule": "Jan 1, 3:00 AM", "description": "Reset annual leave balances (7 sick, 3 emergency days)", "log_file": "cron-yearly-replenish-leaves.log", "category": "attendance"},
        {"id": "archive_old_payrolls", "schedule": "Jan 2, 2:00 AM", "description": "Export and delete payroll data from the previous year", "log_file": "cron-yearly-archive-payrolls.log", "category": "payroll"},
    ]

    # Management commands safe for manual triggering
    TRIGGERABLE_COMMANDS = [
        {"id": "auto_close_attendance", "label": "Auto-Close Attendance", "description": "Close any open attendance sessions", "app": "attendance", "category": "attendance", "destructive": False},
        {"id": "mark_daily_absences", "label": "Mark Daily Absences", "description": "Mark employees absent for today if no attendance record", "app": "attendance", "category": "attendance", "destructive": False},
        {"id": "fix_attendance_time_entries", "label": "Fix Attendance Entries", "description": "Recalculate paid hours, lateness, and penalties for all attendance records", "app": "attendance", "args": ["--verify-only"], "category": "attendance", "destructive": False},
        {"id": "recalculate_pending_attendance", "label": "Recalculate Attendance", "description": "Recompute attendance metrics for pending records", "app": "attendance", "category": "attendance", "destructive": False},
        {"id": "refresh_payroll_from_attendance", "label": "Refresh Payroll", "description": "Recompute payroll calculations from attendance data", "app": "payroll", "args": ["--force"], "category": "payroll", "destructive": False},
        {"id": "delete_old_notifications", "label": "Delete Old Notifications", "description": "Remove read notifications older than 7 days", "app": "notifications", "args": ["--all", "--days", "7"], "category": "maintenance", "destructive": False},
        {"id": "cleanup_unused_images", "label": "Cleanup Unused Images", "description": "Delete orphaned profile images not linked to any user", "app": "users", "category": "maintenance", "destructive": False},
        {"id": "cleanup_archived_quotations", "label": "Cleanup Archived Quotations", "description": "Delete archived quotations older than 14 days", "app": "quotations", "args": ["--days", "14"], "category": "maintenance", "destructive": False},
        {"id": "recalculate_remittances", "label": "Recalculate Remittances", "description": "Recalculate sales totals on all remittance records", "app": "remittances", "category": "sales", "destructive": False},
        {"id": "recalculate_service_revenue", "label": "Recalculate Service Revenue", "description": "Recalculate revenue figures for all services", "app": "services", "category": "sales", "destructive": False},
        {"id": "find_duplicate_items", "label": "Find Duplicate Items", "description": "Scan inventory for duplicate items by name", "app": "inventory", "category": "inventory", "destructive": False},
        {"id": "fix_service_transaction_types", "label": "Fix Service Transaction Types", "description": "Update service-linked sales transactions from 'sale' to 'service' type", "app": "sales", "category": "sales", "destructive": False},
        {"id": "remove_duplicate_clients", "label": "Remove Duplicate Clients", "description": "Remove duplicate client records by contact number, keeping the oldest", "app": "clients", "category": "maintenance", "destructive": True},
        {"id": "add_philippine_holidays", "label": "Add PH Holidays", "description": "Add Philippine holidays for the current year (skip existing)", "app": "payroll", "args": ["--skip-existing"], "category": "payroll", "destructive": False},
        {"id": "fix_zero_balance_payment_status", "label": "Fix Free Service Payment Status", "description": "Fix completed services with free labor/no parts stuck as unpaid", "app": "services", "category": "sales", "destructive": False},
        {"id": "migrate_custom_items", "label": "Migrate Custom Items", "description": "Create untracked items and link historical custom/free-text rows to inventory items", "app": "inventory", "args": ["--apply"], "category": "inventory", "destructive": True},
    ]

    def get(self, request):
        """Return server stats: disk, memory, docker, containers."""
        import shutil
        import subprocess
        import os

        result = {}

        # Disk usage
        total, used, free = shutil.disk_usage("/")
        result["disk"] = {
            "total_gb": round(total / (1024 ** 3), 2),
            "used_gb": round(used / (1024 ** 3), 2),
            "free_gb": round(free / (1024 ** 3), 2),
            "percent_used": round((used / total) * 100, 1),
        }

        # Memory usage
        try:
            with open("/proc/meminfo", "r") as f:
                meminfo = {}
                for line in f:
                    parts = line.split(":")
                    if len(parts) == 2:
                        key = parts[0].strip()
                        val = parts[1].strip().split()[0]  # value in kB
                        meminfo[key] = int(val)
                total_mem = meminfo.get("MemTotal", 0)
                available = meminfo.get("MemAvailable", 0)
                used_mem = total_mem - available
                result["memory"] = {
                    "total_gb": round(total_mem / (1024 ** 2), 2),
                    "used_gb": round(used_mem / (1024 ** 2), 2),
                    "available_gb": round(available / (1024 ** 2), 2),
                    "percent_used": round((used_mem / total_mem) * 100, 1) if total_mem else 0,
                }
        except (FileNotFoundError, ValueError, ZeroDivisionError):
            result["memory"] = None

        # Docker disk usage (if socket available)
        try:
            docker_output = subprocess.run(
                ["docker", "system", "df", "--format", "{{.Type}}\t{{.Size}}\t{{.Reclaimable}}"],
                capture_output=True, text=True, timeout=10,
            )
            if docker_output.returncode == 0:
                lines = docker_output.stdout.strip().split("\n")
                result["docker"] = []
                for line in lines:
                    parts = line.split("\t")
                    if len(parts) >= 3:
                        reclaimable = parts[2].strip()
                        # Sanitize negative/malformed reclaimable values
                        if reclaimable.startswith("-") or "e+" in reclaimable or "e-" in reclaimable:
                            reclaimable = "0B (0%)"
                        result["docker"].append({
                            "type": parts[0],
                            "size": parts[1],
                            "reclaimable": reclaimable,
                        })
        except (FileNotFoundError, subprocess.TimeoutExpired):
            result["docker"] = None

        # Docker containers status
        try:
            containers_output = subprocess.run(
                [
                    "docker", "ps", "-a",
                    "--format", "{{.Names}}\t{{.Status}}\t{{.State}}\t{{.Size}}"
                ],
                capture_output=True, text=True, timeout=10,
            )
            if containers_output.returncode == 0 and containers_output.stdout.strip():
                result["containers"] = []
                for line in containers_output.stdout.strip().split("\n"):
                    parts = line.split("\t")
                    if len(parts) >= 3:
                        result["containers"].append({
                            "name": parts[0],
                            "status": parts[1],
                            "state": parts[2],
                            "size": parts[3] if len(parts) > 3 else "",
                        })
        except (FileNotFoundError, subprocess.TimeoutExpired):
            result["containers"] = None

        # Top large files in media directory
        media_root = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "media")
        try:
            large_files = []
            for dirpath, _, filenames in os.walk(media_root):
                for fname in filenames:
                    fpath = os.path.join(dirpath, fname)
                    try:
                        fsize = os.path.getsize(fpath)
                        if fsize > 1 * 1024 * 1024:  # > 1MB
                            large_files.append({
                                "path": os.path.relpath(fpath, media_root),
                                "size_mb": round(fsize / (1024 * 1024), 2),
                            })
                    except OSError:
                        pass
            large_files.sort(key=lambda x: x["size_mb"], reverse=True)
            result["large_media_files"] = large_files[:20]
            # Total media size
            total_media = sum(
                os.path.getsize(os.path.join(dp, f))
                for dp, _, fns in os.walk(media_root) for f in fns
                if os.path.isfile(os.path.join(dp, f))
            )
            result["media_total_mb"] = round(total_media / (1024 * 1024), 2)
        except (FileNotFoundError, OSError):
            result["large_media_files"] = []
            result["media_total_mb"] = 0

        # Cron jobs with log info
        host_logs_dir = "/host-logs"
        cron_jobs = []
        for job in self.CRON_JOBS:
            job_info = {**job}
            log_path = os.path.join(host_logs_dir, job["log_file"])
            if os.path.isfile(log_path):
                try:
                    stat = os.stat(log_path)
                    job_info["log_size_kb"] = round(stat.st_size / 1024, 1)
                    job_info["last_modified"] = timezone.datetime.fromtimestamp(
                        stat.st_mtime, tz=timezone.get_current_timezone()
                    ).isoformat()
                    # Read last few lines to detect success/failure
                    with open(log_path, "r", errors="replace") as f:
                        lines = f.readlines()
                        tail = "".join(lines[-5:]) if lines else ""
                        if "✅" in tail or "Success" in tail or "complete" in tail.lower():
                            job_info["last_status"] = "success"
                        elif "❌" in tail or "Failed" in tail or "Error" in tail:
                            job_info["last_status"] = "error"
                        else:
                            job_info["last_status"] = "unknown"
                except OSError:
                    job_info["last_status"] = "no_log"
            else:
                job_info["last_status"] = "no_log"
                job_info["log_size_kb"] = 0
                job_info["last_modified"] = None
            cron_jobs.append(job_info)
        result["cron_jobs"] = cron_jobs

        # Available management commands for manual triggering
        result["management_commands"] = self.TRIGGERABLE_COMMANDS

        return Response(result)

    def post(self, request):
        """Run maintenance actions."""
        import subprocess
        import logging
        import threading

        logger = logging.getLogger(__name__)
        action_type = request.data.get("action", "full_cleanup")

        allowed_actions = [
            "docker_prune", "log_cleanup", "full_cleanup",
            "restart_containers", "container_logs",
            "run_command", "view_cron_log", "install_cron_jobs",
            "delete_chats",
            "db_backup", "list_backups", "delete_backup",
        ]
        if action_type not in allowed_actions:
            return Response(
                {"error": f"Invalid action. Allowed: {allowed_actions}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # --- Synchronous actions (fast, return data immediately) ---

        # Container logs
        if action_type == "container_logs":
            container = request.data.get("container", "rvdc_backend-api-1")
            tail_lines = min(int(request.data.get("lines", 100)), 500)
            allowed_containers = ["rvdc_backend-api-1", "rvdc_backend-redis-1", "rvdc_backend-db-1", "rvdc_backend-go2rtc-1"]
            if container not in allowed_containers:
                return Response(
                    {"error": f"Container not allowed. Allowed: {allowed_containers}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            try:
                logs_result = subprocess.run(
                    ["docker", "logs", "--tail", str(tail_lines), container],
                    capture_output=True, text=True, timeout=15,
                )
                log_output = logs_result.stdout or logs_result.stderr or ""
                return Response({
                    "success": True,
                    "action": "container_logs",
                    "container": container,
                    "logs": log_output[-10000:],
                })
            except (FileNotFoundError, subprocess.TimeoutExpired) as e:
                return Response({
                    "success": False,
                    "action": "container_logs",
                    "error": str(e)[:200],
                })

        # View cron log file
        if action_type == "view_cron_log":
            import os
            log_file = request.data.get("log_file", "")
            # Validate against known cron job log files only
            allowed_log_files = {job["log_file"] for job in self.CRON_JOBS}
            if log_file not in allowed_log_files:
                return Response(
                    {"error": "Log file not allowed."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            log_path = os.path.join("/host-logs", log_file)
            try:
                with open(log_path, "r", errors="replace") as f:
                    content = f.read()
                # Return last 15000 chars
                return Response({
                    "success": True,
                    "action": "view_cron_log",
                    "log_file": log_file,
                    "logs": content[-15000:],
                })
            except FileNotFoundError:
                return Response({
                    "success": True,
                    "action": "view_cron_log",
                    "log_file": log_file,
                    "logs": "No log file found. This job may not have run yet.",
                })
            except OSError as e:
                return Response({
                    "success": False,
                    "action": "view_cron_log",
                    "error": str(e)[:200],
                })

        # List database backups
        if action_type == "list_backups":
            import os
            backups_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                "backups",
            )
            os.makedirs(backups_dir, exist_ok=True)
            backups = []
            for fname in sorted(os.listdir(backups_dir), reverse=True):
                if fname.startswith("rvdc_backup_") and fname.endswith((".sql.gz", ".sql")):
                    fpath = os.path.join(backups_dir, fname)
                    try:
                        stat = os.stat(fpath)
                        backups.append({
                            "filename": fname,
                            "size_mb": round(stat.st_size / (1024 * 1024), 2),
                            "created_at": timezone.datetime.fromtimestamp(
                                stat.st_mtime, tz=timezone.get_current_timezone()
                            ).isoformat(),
                        })
                    except OSError:
                        pass
            return Response({
                "success": True,
                "action": "list_backups",
                "backups": backups,
            })

        # Delete a specific backup
        if action_type == "delete_backup":
            import os
            import re
            filename = request.data.get("filename", "")
            if not re.match(r'^rvdc_backup_\d{8}_\d{6}\.sql(\.gz)?$', filename):
                return Response(
                    {"error": "Invalid backup filename."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            backups_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                "backups",
            )
            fpath = os.path.join(backups_dir, filename)
            # Ensure the resolved path is within backups_dir
            if not os.path.realpath(fpath).startswith(os.path.realpath(backups_dir)):
                return Response(
                    {"error": "Invalid path."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if not os.path.isfile(fpath):
                return Response(
                    {"error": "Backup file not found."},
                    status=status.HTTP_404_NOT_FOUND,
                )
            try:
                os.remove(fpath)
                return Response({
                    "success": True,
                    "action": "delete_backup",
                    "message": f"Deleted {filename}",
                })
            except OSError as e:
                return Response({
                    "success": False,
                    "action": "delete_backup",
                    "error": str(e)[:200],
                })

        # --- Background actions (threaded, push result via WebSocket) ---
        user_id = request.user.id
        user_name = request.user.get_full_name()
        container_arg = request.data.get("container", "")
        command_id = request.data.get("command", "")

        # Validate run_command before starting thread
        command_config = None
        if action_type == "run_command":
            command_config = next(
                (cmd for cmd in self.TRIGGERABLE_COMMANDS if cmd["id"] == command_id),
                None,
            )
            if not command_config:
                return Response(
                    {"error": f"Unknown or disallowed command: {command_id}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Require admin credentials for destructive/aggressive actions
        destructive_actions = {"full_cleanup", "install_cron_jobs", "delete_chats"}
        requires_auth = (
            action_type in destructive_actions
            or (action_type == "run_command" and command_config and command_config.get("destructive"))
        )
        if requires_auth:
            from django.contrib.auth import authenticate as auth_check
            admin_username = request.data.get("admin_username", "")
            admin_password = request.data.get("admin_password", "")
            if not admin_username or not admin_password:
                return Response(
                    {"error": "Admin credentials required for this action."},
                    status=status.HTTP_403_FORBIDDEN,
                )
            admin_user = auth_check(username=admin_username, password=admin_password)
            if not admin_user or admin_user.role != "admin":
                return Response(
                    {"error": "Invalid admin credentials."},
                    status=status.HTTP_403_FORBIDDEN,
                )

        def _run_maintenance():
            _logger = logging.getLogger(__name__)
            results = []

            # Docker system prune
            if action_type in ("docker_prune", "full_cleanup"):
                try:
                    prune_result = subprocess.run(
                        ["docker", "system", "prune", "-f"],
                        capture_output=True, text=True, timeout=300,
                    )
                    output = prune_result.stdout.strip()[-500:] if prune_result.stdout else ""
                    results.append({
                        "task": "docker_system_prune",
                        "success": prune_result.returncode == 0,
                        "output": output,
                        "error": prune_result.stderr.strip()[-200:] if prune_result.returncode != 0 else "",
                    })
                except FileNotFoundError:
                    results.append({
                        "task": "docker_system_prune",
                        "success": False,
                        "error": "Docker not available in this environment",
                    })
                except subprocess.TimeoutExpired:
                    results.append({
                        "task": "docker_system_prune",
                        "success": False,
                        "error": "Timed out after 300s",
                    })

                # Image prune (unused images older than 7 days)
                try:
                    img_result = subprocess.run(
                        ["docker", "image", "prune", "-a", "--filter", "until=168h", "-f"],
                        capture_output=True, text=True, timeout=300,
                    )
                    output = img_result.stdout.strip()[-500:] if img_result.stdout else ""
                    results.append({
                        "task": "docker_image_prune",
                        "success": img_result.returncode == 0,
                        "output": output,
                        "error": img_result.stderr.strip()[-200:] if img_result.returncode != 0 else "",
                    })
                except (FileNotFoundError, subprocess.TimeoutExpired) as e:
                    results.append({
                        "task": "docker_image_prune",
                        "success": False,
                        "error": str(e)[:200],
                    })

            # Log cleanup
            if action_type in ("log_cleanup", "full_cleanup"):
                log_files_truncated = 0
                try:
                    import glob
                    import os
                    for log_path in glob.glob("/host-logs/cron-*.log"):
                        try:
                            size = os.path.getsize(log_path)
                            if size > 5 * 1024 * 1024:
                                with open(log_path, "w") as f:
                                    f.write(f"--- Log truncated by admin on {timezone.now().isoformat()} ---\n")
                                log_files_truncated += 1
                        except OSError:
                            pass
                    results.append({
                        "task": "log_cleanup",
                        "success": True,
                        "output": f"Truncated {log_files_truncated} log files over 5MB",
                    })
                except Exception as e:
                    results.append({
                        "task": "log_cleanup",
                        "success": False,
                        "error": str(e)[:200],
                    })

            # Restart containers
            if action_type == "restart_containers":
                allowed_containers = ["rvdc_backend-api-1", "rvdc_backend-redis-1", "rvdc_backend-db-1", "rvdc_backend-go2rtc-1"]
                targets = [container_arg] if container_arg and container_arg in allowed_containers else (
                    allowed_containers if not container_arg else []
                )
                if not targets and container_arg:
                    results.append({
                        "task": "restart",
                        "success": False,
                        "error": f"Container not allowed. Allowed: {allowed_containers}",
                    })
                for c_name in targets:
                    try:
                        restart_result = subprocess.run(
                            ["docker", "restart", c_name],
                            capture_output=True, text=True, timeout=60,
                        )
                        results.append({
                            "task": f"restart_{c_name}",
                            "success": restart_result.returncode == 0,
                            "output": f"{c_name} restarted" if restart_result.returncode == 0 else "",
                            "error": restart_result.stderr.strip()[:200] if restart_result.returncode != 0 else "",
                        })
                    except (FileNotFoundError, subprocess.TimeoutExpired) as e:
                        results.append({
                            "task": f"restart_{c_name}",
                            "success": False,
                            "error": str(e)[:200],
                        })

            # Run management command
            if action_type == "run_command" and command_config:
                from io import StringIO
                from django.core.management import call_command

                cmd_name = command_config["id"]
                cmd_args = command_config.get("args", [])
                try:
                    out = StringIO()
                    err = StringIO()
                    call_command(cmd_name, *cmd_args, stdout=out, stderr=err)
                    output = out.getvalue().strip()
                    error_output = err.getvalue().strip()
                    results.append({
                        "task": cmd_name,
                        "success": True,
                        "output": (output or "Command completed successfully.")[-1000:],
                    })
                    if error_output:
                        results[-1]["output"] += f"\n{error_output[-500:]}"
                except Exception as e:
                    results.append({
                        "task": cmd_name,
                        "success": False,
                        "error": str(e)[:500],
                    })

            # Install cron jobs on host
            if action_type == "install_cron_jobs":
                import os
                import shutil

                container_name = "rvdc_backend-api-1"
                staging_dir = "/host-logs/cron-scripts-staging"
                try:
                    # Generate all cron scripts and crontab entries to staging dir
                    os.makedirs(staging_dir, exist_ok=True)
                    cron_scripts = ServerMaintenanceView._generate_cron_scripts(container_name)
                    for filename, content in cron_scripts.items():
                        fpath = os.path.join(staging_dir, filename)
                        with open(fpath, "w", newline="\n") as f:
                            f.write(content)
                    crontab_entries = ServerMaintenanceView._generate_crontab_entries()
                    with open(os.path.join(staging_dir, "rvdc-crontab-entries.txt"), "w", newline="\n") as f:
                        f.write(crontab_entries)

                    # Deploy to host via temp Docker container with host path mounts
                    deploy_script = (
                        'set -e\n'
                        'mkdir -p /opt/cron-scripts\n'
                        'cp /staging/*.sh /opt/cron-scripts/\n'
                        'chmod +x /opt/cron-scripts/*.sh\n'
                        'CRONTAB_FILE="/crontab-dir/root"\n'
                        'if [ -f "$CRONTAB_FILE" ]; then\n'
                        '  sed "/^# ==.*RVDC/,/^# ==.*END RVDC/d" "$CRONTAB_FILE" > /tmp/clean.txt 2>/dev/null || true\n'
                        'else\n'
                        '  touch /tmp/clean.txt\n'
                        'fi\n'
                        'cat /tmp/clean.txt > "$CRONTAB_FILE"\n'
                        'cat /staging/rvdc-crontab-entries.txt >> "$CRONTAB_FILE"\n'
                        'chmod 600 "$CRONTAB_FILE"\n'
                        'SCRIPT_COUNT=$(ls -1 /opt/cron-scripts/*.sh 2>/dev/null | wc -l)\n'
                        'echo "Deployed $SCRIPT_COUNT cron scripts and updated crontab"\n'
                    )
                    deploy_result = subprocess.run(
                        [
                            "docker", "run", "--rm",
                            "-v", "/var/log/cron-scripts-staging:/staging:ro",
                            "-v", "/opt/cron-scripts:/opt/cron-scripts",
                            "-v", "/var/spool/cron/crontabs:/crontab-dir",
                            "alpine:latest", "sh", "-c", deploy_script,
                        ],
                        capture_output=True, text=True, timeout=120,
                    )
                    shutil.rmtree(staging_dir, ignore_errors=True)

                    if deploy_result.returncode == 0:
                        results.append({
                            "task": "install_cron_jobs",
                            "success": True,
                            "output": deploy_result.stdout.strip()[-500:] or "Cron jobs installed successfully",
                        })
                    else:
                        results.append({
                            "task": "install_cron_jobs",
                            "success": False,
                            "error": (deploy_result.stderr.strip() or deploy_result.stdout.strip())[-500:],
                        })
                except Exception as e:
                    shutil.rmtree(staging_dir, ignore_errors=True)
                    results.append({
                        "task": "install_cron_jobs",
                        "success": False,
                        "error": str(e)[:500],
                    })

            # Delete all chat messages from Redis
            if action_type == "delete_chats":
                try:
                    import redis
                    from django.conf import settings

                    redis_url = getattr(settings, "CHANNEL_LAYERS", {}).get(
                        "default", {}
                    ).get("CONFIG", {}).get("hosts", [("redis", 6379)])[0]
                    if isinstance(redis_url, str):
                        r = redis.from_url(redis_url)
                    else:
                        host, port = redis_url if isinstance(redis_url, (list, tuple)) else ("redis", 6379)
                        r = redis.Redis(host=host, port=port, db=0)

                    deleted_count = 0
                    patterns = ["chat:*", "chat:unread:*", "chat:online",
                                "chat:heartbeat:*", "chat:last_seen:*"]
                    for pattern in patterns:
                        cursor = 0
                        while True:
                            cursor, keys = r.scan(cursor=cursor, match=pattern, count=200)
                            if keys:
                                deleted_count += r.delete(*keys)
                            if cursor == 0:
                                break

                    results.append({
                        "task": "delete_chats",
                        "success": True,
                        "output": f"Deleted {deleted_count} chat keys from Redis",
                    })
                except Exception as e:
                    results.append({
                        "task": "delete_chats",
                        "success": False,
                        "error": str(e)[:500],
                    })

            # Create database backup
            if action_type == "db_backup":
                import os
                import re
                try:
                    from django.conf import settings as django_settings
                    db_conf = django_settings.DATABASES["default"]
                    db_name = db_conf["NAME"]
                    db_user = db_conf["USER"]

                    backups_dir = os.path.join(
                        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                        "backups",
                    )
                    os.makedirs(backups_dir, exist_ok=True)

                    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"rvdc_backup_{timestamp}.sql.gz"
                    backup_path = os.path.join(backups_dir, filename)

                    # Find the db container name
                    container_result = subprocess.run(
                        ["docker", "ps", "--filter", "ancestor=postgres:16",
                         "--format", "{{.Names}}"],
                        capture_output=True, text=True, timeout=10,
                    )
                    db_container = container_result.stdout.strip().split("\n")[0]
                    if not db_container:
                        db_container = "rvdc_backend-db-1"

                    # Run pg_dump inside the db container, pipe through gzip
                    dump_cmd = (
                        f'docker exec {db_container} pg_dump -U {db_user} -d {db_name} '
                        f'--no-owner --no-privileges'
                    )
                    # Validate db_user and db_name to prevent injection
                    if not re.match(r'^[a-zA-Z0-9_]+$', db_user):
                        raise ValueError("Invalid database user name")
                    if not re.match(r'^[a-zA-Z0-9_]+$', db_name):
                        raise ValueError("Invalid database name")

                    dump_result = subprocess.run(
                        [
                            "docker", "exec", db_container,
                            "pg_dump", "-U", db_user, "-d", db_name,
                            "--no-owner", "--no-privileges",
                        ],
                        capture_output=True, timeout=600,
                    )
                    if dump_result.returncode != 0:
                        error_msg = dump_result.stderr.decode("utf-8", errors="replace")[:500]
                        raise RuntimeError(f"pg_dump failed: {error_msg}")

                    # Compress the dump
                    import gzip
                    with gzip.open(backup_path, "wb") as f:
                        f.write(dump_result.stdout)

                    size_mb = round(os.path.getsize(backup_path) / (1024 * 1024), 2)
                    results.append({
                        "task": "db_backup",
                        "success": True,
                        "output": f"Backup created: {filename} ({size_mb} MB)",
                    })
                except Exception as e:
                    results.append({
                        "task": "db_backup",
                        "success": False,
                        "error": str(e)[:500],
                    })

            all_success = all(r.get("success") for r in results)
            _logger.info(
                "Server maintenance by %s: action=%s results=%s",
                user_name, action_type, results,
            )

            # Push result to admin via WebSocket notification channel
            try:
                from asgiref.sync import async_to_sync
                from channels.layers import get_channel_layer

                channel_layer = get_channel_layer()
                if channel_layer:
                    action_labels = {
                        "docker_prune": "Docker Cleanup",
                        "log_cleanup": "Log Cleanup",
                        "full_cleanup": "Full System Cleanup",
                        "restart_containers": "Container Restart",
                        "run_command": command_config["label"] if command_config else "Command",
                        "install_cron_jobs": "Install Cron Jobs",
                        "delete_chats": "Delete Chats",
                        "db_backup": "Database Backup",
                    }
                    label = action_labels.get(action_type, action_type)
                    failed = [r for r in results if not r.get("success")]
                    if all_success:
                        title = f"{label} completed"
                        message = "; ".join(r.get("output", r["task"]) for r in results if r.get("output"))
                        if not message:
                            message = "All tasks completed successfully."
                    else:
                        title = f"{label} finished with issues"
                        message = "; ".join(r.get("error", r["task"]) for r in failed)

                    async_to_sync(channel_layer.group_send)(
                        f"notifications_{user_id}",
                        {
                            "type": "send_notification",
                            "data": {
                                "event": "maintenance_result",
                                "success": all_success,
                                "action": action_type,
                                "title": title,
                                "message": message,
                                "results": results,
                            },
                        },
                    )
            except Exception:
                _logger.exception("Failed to send maintenance result via WebSocket")

        thread = threading.Thread(target=_run_maintenance, daemon=True)
        thread.start()

        action_labels = {
            "docker_prune": "Docker Cleanup",
            "log_cleanup": "Log Cleanup",
            "full_cleanup": "Full System Cleanup",
            "restart_containers": "Container Restart",
            "run_command": command_config["label"] if command_config else "Command",
            "install_cron_jobs": "Install Cron Jobs",
            "delete_chats": "Delete Chats",
            "db_backup": "Database Backup",
        }
        return Response({
            "accepted": True,
            "action": action_type,
            "message": f"{action_labels.get(action_type, action_type)} started. You'll be notified when it completes.",
        }, status=status.HTTP_202_ACCEPTED)

    @staticmethod
    def _generate_cron_scripts(container_name):
        """Generate all cron script file contents keyed by filename."""

        def _mgmt(log_file, title, command, extra_args=""):
            return (
                '#!/bin/bash\n'
                f'LOG_FILE="/var/log/{log_file}"\n'
                f'CONTAINER_NAME="{container_name}"\n'
                'export TZ=Asia/Manila\n\n'
                'echo "=== ' + title + " - $(date '+%Y-%m-%d %H:%M:%S %Z') ===\" >> \"$LOG_FILE\"\n"
                f'docker exec "$CONTAINER_NAME" python manage.py {command}'
                f'{" " + extra_args if extra_args else ""} >> "$LOG_FILE" 2>&1\n\n'
                'if [ $? -eq 0 ]; then\n'
                "    echo \"✅ Success - $(date '+%Y-%m-%d %H:%M:%S %Z')\" >> \"$LOG_FILE\"\n"
                'else\n'
                "    echo \"❌ Failed - $(date '+%Y-%m-%d %H:%M:%S %Z')\" >> \"$LOG_FILE\"\n"
                'fi\n'
                'echo "" >> "$LOG_FILE"\n'
            )

        scripts = {
            "auto-close-attendance.sh": _mgmt(
                "cron-auto-close-attendance.log", "Auto-Close Attendance", "auto_close_attendance",
            ),
            "mark-absences.sh": _mgmt(
                "cron-mark-absences.log", "Mark Daily Absences", "mark_daily_absences",
            ),
            "delete-old-notifications.sh": _mgmt(
                "cron-delete-old-notifications.log", "Delete Old Notifications",
                "delete_old_notifications", "--all --days 7",
            ),
            "cleanup-archived-quotations.sh": _mgmt(
                "cron-cleanup-archived-quotations.log", "Cleanup Archived Quotations",
                "cleanup_archived_quotations", "--days 14",
            ),
            "cleanup-unused-images.sh": _mgmt(
                "cron-cleanup-images.log", "Cleanup Unused Images", "cleanup_unused_images",
            ),
        }

        # Weekly attendance + payroll fix (multi-step)
        scripts["weekly-attendance-payroll-fix.sh"] = (
            '#!/bin/bash\n'
            f'LOG_FILE="/var/log/cron-weekly-attendance-payroll-fix.log"\n'
            f'CONTAINER_NAME="{container_name}"\n'
            'export TZ=Asia/Manila\n\n'
            'LAST_SATURDAY=$(date -d "last saturday -7 days" +%Y-%m-%d)\n'
            'LAST_FRIDAY=$(date -d "yesterday" +%Y-%m-%d)\n\n'
            "echo \"=== Weekly Attendance & Payroll Fix - $(date '+%Y-%m-%d %H:%M:%S %Z') ===\" >> \"$LOG_FILE\"\n"
            'echo "Processing period: $LAST_SATURDAY to $LAST_FRIDAY" >> "$LOG_FILE"\n\n'
            'echo "--- Step 1: Verifying attendance ---" >> "$LOG_FILE"\n'
            'docker exec "$CONTAINER_NAME" python manage.py fix_attendance_time_entries \\\n'
            '    --verify-only \\\n'
            '    --start-date "$LAST_SATURDAY" \\\n'
            '    --end-date "$LAST_FRIDAY" >> "$LOG_FILE" 2>&1\n\n'
            'echo "--- Step 2: Fixing attendance ---" >> "$LOG_FILE"\n'
            'docker exec "$CONTAINER_NAME" python manage.py fix_attendance_time_entries \\\n'
            '    --force \\\n'
            '    --start-date "$LAST_SATURDAY" \\\n'
            '    --end-date "$LAST_FRIDAY" >> "$LOG_FILE" 2>&1\n\n'
            'echo "--- Step 3: Refreshing payroll ---" >> "$LOG_FILE"\n'
            'docker exec "$CONTAINER_NAME" python manage.py refresh_payroll_from_attendance \\\n'
            '    --force \\\n'
            '    --start-date "$LAST_SATURDAY" \\\n'
            '    --end-date "$LAST_FRIDAY" >> "$LOG_FILE" 2>&1\n\n'
            'if [ $? -eq 0 ]; then\n'
            "    echo \"✅ Weekly fix complete - $(date '+%Y-%m-%d %H:%M:%S %Z')\" >> \"$LOG_FILE\"\n"
            'else\n'
            "    echo \"❌ Weekly fix failed - $(date '+%Y-%m-%d %H:%M:%S %Z')\" >> \"$LOG_FILE\"\n"
            'fi\n'
            'echo "" >> "$LOG_FILE"\n'
        )

        scripts["yearly-update-holidays.sh"] = (
            '#!/bin/bash\n'
            f'LOG_FILE="/var/log/cron-yearly-update-holidays.log"\n'
            f'CONTAINER_NAME="{container_name}"\n'
            'export TZ=Asia/Manila\n\n'
            'CURRENT_YEAR=$(date +%Y)\n'
            'NEXT_YEAR=$((CURRENT_YEAR + 1))\n\n'
            "echo \"=== Update Holiday Years - $(date '+%Y-%m-%d %H:%M:%S %Z') ===\" >> \"$LOG_FILE\"\n"
            'echo "Updating holidays to year: $NEXT_YEAR" >> "$LOG_FILE"\n\n'
            'docker exec "$CONTAINER_NAME" python manage.py update_holiday_years \\\n'
            '    --year "$NEXT_YEAR" --force >> "$LOG_FILE" 2>&1\n\n'
            'if [ $? -eq 0 ]; then\n'
            "    echo \"✅ Holiday years updated - $(date '+%Y-%m-%d %H:%M:%S %Z')\" >> \"$LOG_FILE\"\n"
            'else\n'
            "    echo \"❌ Holiday update failed - $(date '+%Y-%m-%d %H:%M:%S %Z')\" >> \"$LOG_FILE\"\n"
            'fi\n'
            'echo "" >> "$LOG_FILE"\n'
        )

        scripts["yearly-replenish-leaves.sh"] = (
            '#!/bin/bash\n'
            f'LOG_FILE="/var/log/cron-yearly-replenish-leaves.log"\n'
            f'CONTAINER_NAME="{container_name}"\n'
            'export TZ=Asia/Manila\n\n'
            "echo \"=== Replenish Leave Balances - $(date '+%Y-%m-%d %H:%M:%S %Z') ===\" >> \"$LOG_FILE\"\n\n"
            'docker exec "$CONTAINER_NAME" python manage.py replenish_leave_balances \\\n'
            '    --reset --sick-leave 7 --emergency-leave 3 --force >> "$LOG_FILE" 2>&1\n\n'
            'if [ $? -eq 0 ]; then\n'
            "    echo \"✅ Leave balances replenished - $(date '+%Y-%m-%d %H:%M:%S %Z')\" >> \"$LOG_FILE\"\n"
            'else\n'
            "    echo \"❌ Leave replenishment failed - $(date '+%Y-%m-%d %H:%M:%S %Z')\" >> \"$LOG_FILE\"\n"
            'fi\n'
            'echo "" >> "$LOG_FILE"\n'
        )

        scripts["yearly-archive-payrolls.sh"] = (
            '#!/bin/bash\n'
            f'LOG_FILE="/var/log/cron-yearly-archive-payrolls.log"\n'
            f'CONTAINER_NAME="{container_name}"\n'
            'export TZ=Asia/Manila\n\n'
            'CURRENT_YEAR=$(date +%Y)\n'
            'LAST_YEAR=$((CURRENT_YEAR - 1))\n\n'
            "echo \"=== Archive Old Payrolls - $(date '+%Y-%m-%d %H:%M:%S %Z') ===\" >> \"$LOG_FILE\"\n"
            'echo "Archiving payrolls from year: $LAST_YEAR and earlier" >> "$LOG_FILE"\n\n'
            'docker exec "$CONTAINER_NAME" python manage.py archive_old_payrolls \\\n'
            '    --year "$LAST_YEAR" --export --delete --force >> "$LOG_FILE" 2>&1\n\n'
            'if [ $? -eq 0 ]; then\n'
            "    echo \"✅ Payrolls archived - $(date '+%Y-%m-%d %H:%M:%S %Z')\" >> \"$LOG_FILE\"\n"
            'else\n'
            "    echo \"❌ Payroll archival failed - $(date '+%Y-%m-%d %H:%M:%S %Z')\" >> \"$LOG_FILE\"\n"
            'fi\n'
            'echo "" >> "$LOG_FILE"\n'
        )

        # Host-level scripts (no docker exec needed)
        scripts["docker-system-prune.sh"] = (
            '#!/bin/bash\n'
            'LOG_FILE="/var/log/cron-docker-prune.log"\n'
            'export TZ=Asia/Manila\n\n'
            "echo \"=== Docker System Prune - $(date '+%Y-%m-%d %H:%M:%S %Z') ===\" >> \"$LOG_FILE\"\n\n"
            'echo "--- Pruning containers, networks, and dangling images ---" >> "$LOG_FILE"\n'
            'docker system prune -f >> "$LOG_FILE" 2>&1\n\n'
            'echo "--- Removing unused images older than 7 days ---" >> "$LOG_FILE"\n'
            'docker image prune -a --filter "until=168h" -f >> "$LOG_FILE" 2>&1\n\n'
            'echo "--- Docker disk usage after prune ---" >> "$LOG_FILE"\n'
            'docker system df >> "$LOG_FILE" 2>&1\n\n'
            "echo \"✅ Docker prune complete - $(date '+%Y-%m-%d %H:%M:%S %Z')\" >> \"$LOG_FILE\"\n"
            'echo "" >> "$LOG_FILE"\n'
        )

        scripts["disk-usage-check.sh"] = (
            '#!/bin/bash\n'
            'LOG_FILE="/var/log/cron-disk-usage.log"\n'
            'export TZ=Asia/Manila\n\n'
            'DISK_THRESHOLD=80\n'
            'DISK_CRITICAL=90\n'
            "DISK_USAGE=$(df / | awk 'NR==2 {print $5}' | sed 's/%//')\n"
            "DISK_AVAIL=$(df -h / | awk 'NR==2 {print $4}')\n\n"
            "echo \"=== Disk Usage Check - $(date '+%Y-%m-%d %H:%M:%S %Z') ===\" >> \"$LOG_FILE\"\n"
            'echo "Disk usage: ${DISK_USAGE}% (Available: ${DISK_AVAIL})" >> "$LOG_FILE"\n\n'
            'if [ "$DISK_USAGE" -ge "$DISK_CRITICAL" ]; then\n'
            '    echo "CRITICAL: Disk usage ${DISK_USAGE}% exceeds ${DISK_CRITICAL}%!" >> "$LOG_FILE"\n'
            '    echo "Running emergency cleanup..." >> "$LOG_FILE"\n'
            '    apt-get clean 2>/dev/null\n'
            '    journalctl --vacuum-size=50M 2>/dev/null\n'
            '    for logfile in /var/log/cron-*.log; do\n'
            '        if [ -f "$logfile" ] && [ $(wc -l < "$logfile") -gt 1000 ]; then\n'
            '            tail -1000 "$logfile" > "${logfile}.tmp" && mv "${logfile}.tmp" "$logfile"\n'
            '            echo "  Truncated: $logfile" >> "$LOG_FILE"\n'
            '        fi\n'
            '    done\n'
            '    docker system prune -f >> "$LOG_FILE" 2>&1\n'
            "    NEW_USAGE=$(df / | awk 'NR==2 {print $5}' | sed 's/%//')\n"
            '    echo "After cleanup: ${NEW_USAGE}%" >> "$LOG_FILE"\n'
            'elif [ "$DISK_USAGE" -ge "$DISK_THRESHOLD" ]; then\n'
            '    echo "WARNING: Disk usage ${DISK_USAGE}% exceeds ${DISK_THRESHOLD}% threshold" >> "$LOG_FILE"\n'
            'else\n'
            '    echo "Disk usage OK" >> "$LOG_FILE"\n'
            'fi\n'
            'echo "" >> "$LOG_FILE"\n'
        )

        scripts["truncate-large-logs.sh"] = (
            '#!/bin/bash\n'
            'LOG_FILE="/var/log/cron-log-maintenance.log"\n'
            'export TZ=Asia/Manila\n'
            'MAX_LOG_SIZE_KB=10240\n\n'
            "echo \"=== Log Truncation - $(date '+%Y-%m-%d %H:%M:%S %Z') ===\" >> \"$LOG_FILE\"\n\n"
            'for logfile in /var/log/cron-*.log; do\n'
            '    if [ -f "$logfile" ] && [ "$logfile" != "$LOG_FILE" ]; then\n'
            '        SIZE_KB=$(du -k "$logfile" | cut -f1)\n'
            '        if [ "$SIZE_KB" -gt "$MAX_LOG_SIZE_KB" ]; then\n'
            '            LINES=$(wc -l < "$logfile")\n'
            '            tail -500 "$logfile" > "${logfile}.tmp" && mv "${logfile}.tmp" "$logfile"\n'
            '            echo "  Truncated $logfile (was ${SIZE_KB}KB, ${LINES} lines)" >> "$LOG_FILE"\n'
            '        fi\n'
            '    fi\n'
            'done\n\n'
            'for syslog in /var/log/syslog /var/log/kern.log /var/log/auth.log; do\n'
            '    if [ -f "$syslog" ]; then\n'
            '        SIZE_KB=$(du -k "$syslog" | cut -f1)\n'
            '        if [ "$SIZE_KB" -gt 51200 ]; then\n'
            '            tail -2000 "$syslog" > "${syslog}.tmp" && mv "${syslog}.tmp" "$syslog"\n'
            '            echo "  Truncated $syslog (was ${SIZE_KB}KB)" >> "$LOG_FILE"\n'
            '        fi\n'
            '    fi\n'
            'done\n\n'
            'journalctl --vacuum-size=100M >> "$LOG_FILE" 2>&1\n\n'
            "echo \"✅ Log maintenance complete\" >> \"$LOG_FILE\"\n"
            'echo "" >> "$LOG_FILE"\n'
        )

        return scripts

    @staticmethod
    def _generate_crontab_entries():
        """Generate RVDC crontab entries for root crontab."""
        return (
            '# ============================================================================\n'
            '# RVDC Attendance & Payroll Management Cron Jobs\n'
            '# Auto-installed via Server Maintenance Dashboard\n'
            '# All times are in Philippines Time (UTC+8)\n'
            '# ============================================================================\n\n'
            '# DAILY TASKS\n'
            '# 9:00 PM PHT (1:00 PM UTC) - Auto-close open attendance records\n'
            '0 13 * * * /opt/cron-scripts/auto-close-attendance.sh\n\n'
            '# 11:30 PM PHT (3:30 PM UTC) - Mark absent employees\n'
            '30 15 * * * /opt/cron-scripts/mark-absences.sh\n\n'
            '# 2:00 AM PHT (6:00 PM prev day UTC) - Delete old notifications\n'
            '0 18 * * * /opt/cron-scripts/delete-old-notifications.sh\n\n'
            '# 3:00 AM PHT (7:00 PM prev day UTC) - Cleanup archived quotations\n'
            '0 19 * * * /opt/cron-scripts/cleanup-archived-quotations.sh\n\n'
            '# 5:00 AM PHT (9:00 PM prev day UTC) - Truncate large log files\n'
            '0 21 * * * /opt/cron-scripts/truncate-large-logs.sh\n\n'
            '# 6:00 AM PHT (10:00 PM prev day UTC) - Disk usage check\n'
            '0 22 * * * /opt/cron-scripts/disk-usage-check.sh\n\n'
            '# WEEKLY TASKS\n'
            '# Friday 11:00 PM PHT (3:00 PM UTC) - Fix attendance + refresh payroll\n'
            '0 15 * * 5 /opt/cron-scripts/weekly-attendance-payroll-fix.sh\n\n'
            '# Sunday 3:00 AM PHT (7:00 PM Sat UTC) - Docker system prune\n'
            '0 19 * * 6 /opt/cron-scripts/docker-system-prune.sh\n\n'
            '# Sunday 3:30 AM PHT (7:30 PM Sat UTC) - Cleanup unused profile images\n'
            '30 19 * * 6 /opt/cron-scripts/cleanup-unused-images.sh\n\n'
            '# YEARLY TASKS\n'
            '# Jan 1 2:00 AM PHT (6:00 PM Dec 31 UTC) - Update holiday years\n'
            '0 18 31 12 * /opt/cron-scripts/yearly-update-holidays.sh\n\n'
            '# Jan 1 3:00 AM PHT (7:00 PM Dec 31 UTC) - Replenish leave balances\n'
            '0 19 31 12 * /opt/cron-scripts/yearly-replenish-leaves.sh\n\n'
            '# Jan 2 2:00 AM PHT (6:00 PM Jan 1 UTC) - Archive old payrolls\n'
            '0 18 1 1 * /opt/cron-scripts/yearly-archive-payrolls.sh\n\n'
            '# MONTHLY - Delete old log files (90+ days)\n'
            '0 19 * * * [ "$(date +\\%d)" = "01" ] && find /var/log/cron-*.log -type f -mtime +90 -delete\n\n'
            '# ============================================================================\n'
            '# END RVDC Cron Jobs\n'
            '# ============================================================================\n'
        )


# -----------------------------------------------------------------------
# Database Backup Download
# -----------------------------------------------------------------------

class BackupDownloadView(APIView):
    """Download a database backup file. Admin-only."""
    permission_classes = [permissions.IsAdminUser]

    def get(self, request, filename):
        import os
        import re
        from django.http import FileResponse

        if not re.match(r'^rvdc_backup_\d{8}_\d{6}\.sql(\.gz)?$', filename):
            return Response(
                {"error": "Invalid backup filename."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        backups_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
            "backups",
        )
        fpath = os.path.join(backups_dir, filename)

        if not os.path.realpath(fpath).startswith(os.path.realpath(backups_dir)):
            return Response(
                {"error": "Invalid path."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not os.path.isfile(fpath):
            return Response(
                {"error": "Backup file not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        content_type = "application/gzip" if filename.endswith(".gz") else "application/sql"
        response = FileResponse(open(fpath, "rb"), content_type=content_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


# -----------------------------------------------------------------------
# Database Backup Upload & Restore
# -----------------------------------------------------------------------

class BackupUploadRestoreView(APIView):
    """Upload a .sql.gz or .sql backup and restore the PostgreSQL database. Admin-only."""
    permission_classes = [permissions.IsAdminUser]
    parser_classes = [parsers.MultiPartParser]

    def post(self, request):
        import gzip
        import logging
        import os
        import re
        import subprocess
        import threading

        logger = logging.getLogger(__name__)

        # Require admin credentials
        from django.contrib.auth import authenticate as auth_check
        admin_username = request.data.get("admin_username", "")
        admin_password = request.data.get("admin_password", "")
        if not admin_username or not admin_password:
            return Response(
                {"error": "Admin credentials required for database restore."},
                status=status.HTTP_403_FORBIDDEN,
            )
        admin_user = auth_check(username=admin_username, password=admin_password)
        if not admin_user or admin_user.role != "admin":
            return Response(
                {"error": "Invalid admin credentials."},
                status=status.HTTP_403_FORBIDDEN,
            )

        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response(
                {"error": "No file uploaded."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        filename = uploaded_file.name
        if not filename.endswith((".sql.gz", ".sql")):
            return Response(
                {"error": "Only .sql or .sql.gz files are allowed."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Limit file size to 500MB
        if uploaded_file.size > 500 * 1024 * 1024:
            return Response(
                {"error": "File too large. Maximum 500MB."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Save uploaded file to backups dir
        backups_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
            "backups",
        )
        os.makedirs(backups_dir, exist_ok=True)

        # Sanitize: use a safe name
        safe_name = re.sub(r'[^a-zA-Z0-9._-]', '_', filename)
        save_path = os.path.join(backups_dir, safe_name)
        if not os.path.realpath(save_path).startswith(os.path.realpath(backups_dir)):
            return Response(
                {"error": "Invalid filename."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with open(save_path, "wb") as f:
            for chunk in uploaded_file.chunks():
                f.write(chunk)

        user_id = request.user.id
        user_name = request.user.get_full_name()

        def _run_restore():
            _logger = logging.getLogger(__name__)
            results = []
            try:
                from datetime import datetime as _dt
                from django.conf import settings as django_settings
                db_conf = django_settings.DATABASES["default"]
                db_name = db_conf["NAME"]
                db_user = db_conf["USER"]

                if not re.match(r'^[a-zA-Z0-9_]+$', db_user):
                    raise ValueError("Invalid database user name")
                if not re.match(r'^[a-zA-Z0-9_]+$', db_name):
                    raise ValueError("Invalid database name")

                # Find the db container
                container_result = subprocess.run(
                    ["docker", "ps", "--filter", "ancestor=postgres:16",
                     "--format", "{{.Names}}"],
                    capture_output=True, text=True, timeout=10,
                )
                db_container = container_result.stdout.strip().split("\n")[0]
                if not db_container:
                    db_container = "rvdc_backend-db-1"

                # --- Auto-backup before restore (safety net) ---
                try:
                    pre_restore_name = f"pre_restore_{_dt.now().strftime('%Y%m%d_%H%M%S')}.sql.gz"
                    pre_restore_path = os.path.join(backups_dir, pre_restore_name)
                    dump_result = subprocess.run(
                        [
                            "docker", "exec", db_container,
                            "pg_dump", "-U", db_user, "-d", db_name,
                            "--no-owner", "--no-privileges",
                        ],
                        capture_output=True, timeout=600,
                    )
                    if dump_result.returncode == 0:
                        with gzip.open(pre_restore_path, "wb") as gz_f:
                            gz_f.write(dump_result.stdout)
                        _logger.info("Pre-restore backup created: %s", pre_restore_name)
                    else:
                        _logger.warning("Pre-restore backup failed, continuing anyway: %s",
                                        dump_result.stderr[:300] if dump_result.stderr else "unknown")
                except Exception as pre_err:
                    _logger.warning("Pre-restore backup failed, continuing anyway: %s", pre_err)

                # Decompress if .gz
                if save_path.endswith(".gz"):
                    sql_path = save_path[:-3]  # remove .gz
                    with gzip.open(save_path, "rb") as gz_in:
                        with open(sql_path, "wb") as sql_out:
                            while True:
                                chunk = gz_in.read(1024 * 1024)
                                if not chunk:
                                    break
                                sql_out.write(chunk)
                else:
                    sql_path = save_path

                # Copy SQL file into the container
                copy_result = subprocess.run(
                    ["docker", "cp", sql_path, f"{db_container}:/tmp/restore.sql"],
                    capture_output=True, text=True, timeout=60,
                )
                if copy_result.returncode != 0:
                    raise RuntimeError(f"Failed to copy file to container: {copy_result.stderr[:300]}")

                # Drop and recreate the database, then restore
                # First terminate existing connections
                subprocess.run(
                    [
                        "docker", "exec", db_container,
                        "psql", "-U", db_user, "-d", "postgres", "-c",
                        f"SELECT pg_terminate_backend(pid) FROM pg_stat_activity WHERE datname = '{db_name}' AND pid <> pg_backend_pid();",
                    ],
                    capture_output=True, text=True, timeout=30,
                )

                # Drop database
                drop_result = subprocess.run(
                    [
                        "docker", "exec", db_container,
                        "psql", "-U", db_user, "-d", "postgres", "-c",
                        f"DROP DATABASE IF EXISTS {db_name};",
                    ],
                    capture_output=True, text=True, timeout=30,
                )
                if drop_result.returncode != 0:
                    raise RuntimeError(f"Failed to drop database: {drop_result.stderr[:300]}")

                # Create database
                create_result = subprocess.run(
                    [
                        "docker", "exec", db_container,
                        "psql", "-U", db_user, "-d", "postgres", "-c",
                        f"CREATE DATABASE {db_name} OWNER {db_user};",
                    ],
                    capture_output=True, text=True, timeout=30,
                )
                if create_result.returncode != 0:
                    raise RuntimeError(f"Failed to create database: {create_result.stderr[:300]}")

                # Restore from SQL
                restore_result = subprocess.run(
                    [
                        "docker", "exec", db_container,
                        "psql", "-U", db_user, "-d", db_name, "-f", "/tmp/restore.sql",
                    ],
                    capture_output=True, text=True, timeout=600,
                )

                # Cleanup temp file in container
                subprocess.run(
                    ["docker", "exec", db_container, "rm", "-f", "/tmp/restore.sql"],
                    capture_output=True, text=True, timeout=10,
                )

                # Cleanup decompressed sql if we created one
                if save_path.endswith(".gz") and os.path.isfile(sql_path):
                    os.remove(sql_path)

                if restore_result.returncode != 0:
                    error_lines = restore_result.stderr.strip().split("\n")[-5:]
                    raise RuntimeError(f"Restore errors: {''.join(error_lines)[:500]}")

                results.append({
                    "task": "restore_backup",
                    "success": True,
                    "output": f"Database restored from {filename}",
                })
            except Exception as e:
                results.append({
                    "task": "restore_backup",
                    "success": False,
                    "error": str(e)[:500],
                })

            all_success = all(r.get("success") for r in results)
            _logger.info(
                "Database restore by %s: file=%s results=%s",
                user_name, filename, results,
            )

            try:
                from asgiref.sync import async_to_sync
                from channels.layers import get_channel_layer

                channel_layer = get_channel_layer()
                if channel_layer:
                    if all_success:
                        title = "Database Restore completed"
                        message = f"Database restored from {filename}"
                    else:
                        title = "Database Restore failed"
                        message = "; ".join(r.get("error", r["task"]) for r in results if not r.get("success"))

                    async_to_sync(channel_layer.group_send)(
                        f"notifications_{user_id}",
                        {
                            "type": "send_notification",
                            "data": {
                                "event": "maintenance_result",
                                "success": all_success,
                                "action": "restore_backup",
                                "title": title,
                                "message": message,
                                "results": results,
                            },
                        },
                    )
            except Exception:
                _logger.exception("Failed to send restore result via WebSocket")

        thread = threading.Thread(target=_run_restore, daemon=True)
        thread.start()

        logger.info("Database restore started by %s from file %s", user_name, filename)
        return Response({
            "accepted": True,
            "action": "restore_backup",
            "message": f"Database restore from {filename} started. You'll be notified when it completes.",
        }, status=status.HTTP_202_ACCEPTED)


# -----------------------------------------------------------------------
# Employee Bulk Update
# -----------------------------------------------------------------------

class EmployeeBulkTemplateView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        """Download an XLSX file pre-filled with all active employees."""
        import io
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        employees = CustomUser.objects.filter(is_deleted=False).order_by("last_name", "first_name")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employees"

        headers = [
            "ID", "First Name", "Last Name", "Role", "Contact Number",
            "Province", "City", "Barangay", "Address",
            "Basic Salary", "SSS #", "PhilHealth #", "TIN #",
        ]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, emp in enumerate(employees, 2):
            ws.cell(row=row_idx, column=1, value=emp.id).border = thin_border
            ws.cell(row=row_idx, column=2, value=emp.first_name or "").border = thin_border
            ws.cell(row=row_idx, column=3, value=emp.last_name or "").border = thin_border
            ws.cell(row=row_idx, column=4, value=emp.role or "").border = thin_border
            ws.cell(row=row_idx, column=5, value=emp.contact_number or "").border = thin_border
            ws.cell(row=row_idx, column=6, value=emp.province or "").border = thin_border
            ws.cell(row=row_idx, column=7, value=emp.city or "").border = thin_border
            ws.cell(row=row_idx, column=8, value=emp.barangay or "").border = thin_border
            ws.cell(row=row_idx, column=9, value=emp.address or "").border = thin_border
            ws.cell(row=row_idx, column=10, value=float(emp.basic_salary) if emp.basic_salary else "").border = thin_border
            ws.cell(row=row_idx, column=11, value=emp.sss_number or "").border = thin_border
            ws.cell(row=row_idx, column=12, value=emp.philhealth_number or "").border = thin_border
            ws.cell(row=row_idx, column=13, value=emp.tin_number or "").border = thin_border

        ws.sheet_properties.tabColor = "1F4E79"
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="employee_template.xlsx"'
        return resp


class EmployeeBulkPreviewView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def post(self, request):
        """Upload an XLSX to preview employee changes."""
        import openpyxl
        from decimal import Decimal, InvalidOperation

        xlsx_file = request.FILES.get("file")
        if not xlsx_file:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        if not xlsx_file.name.endswith((".xlsx", ".xlsm")):
            return Response({"detail": "Only .xlsx files are supported."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        except Exception:
            return Response({"detail": "Could not parse the uploaded file."}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        if not rows:
            return Response({"detail": "The file contains no data rows."}, status=status.HTTP_400_BAD_REQUEST)

        all_employees = {e.id: e for e in CustomUser.objects.filter(is_deleted=False)}

        changes = []
        skipped = 0
        errors = []

        for row_num, row in enumerate(rows, 2):
            if len(row) < 13:
                errors.append({"row": row_num, "error": "Row has fewer than 13 columns."})
                continue

            try:
                emp_id = int(row[0]) if row[0] is not None else None
            except (ValueError, TypeError):
                errors.append({"row": row_num, "error": f"Invalid ID: {row[0]}"})
                continue

            if not emp_id:
                continue

            emp = all_employees.get(emp_id)
            if not emp:
                errors.append({"row": row_num, "sku": str(emp_id), "error": "Employee ID not found."})
                continue

            new_first = str(row[1] or "").strip()
            new_last = str(row[2] or "").strip()
            new_role = str(row[3] or "").strip()
            new_contact = str(row[4] or "").strip() or None
            new_province = str(row[5] or "").strip() or None
            new_city = str(row[6] or "").strip() or None
            new_barangay = str(row[7] or "").strip() or None
            new_address = str(row[8] or "").strip() or None

            try:
                new_salary = Decimal(str(row[9])) if row[9] is not None and str(row[9]).strip() != "" else None
            except (InvalidOperation, ValueError) as e:
                errors.append({"row": row_num, "sku": str(emp_id), "error": f"Invalid salary: {e}"})
                continue

            new_sss = str(row[10] or "").strip() or None
            new_philhealth = str(row[11] or "").strip() or None
            new_tin = str(row[12] or "").strip() or None

            valid_roles = {"admin", "manager", "clerk", "technician"}
            if new_role and new_role not in valid_roles:
                errors.append({"row": row_num, "sku": str(emp_id), "error": f"Invalid role: {new_role}"})
                continue

            item_changes = []
            if new_first and new_first != (emp.first_name or ""):
                item_changes.append({"field": "First Name", "old": emp.first_name or "", "new": new_first})
            if new_last and new_last != (emp.last_name or ""):
                item_changes.append({"field": "Last Name", "old": emp.last_name or "", "new": new_last})
            if new_role and new_role != emp.role:
                item_changes.append({"field": "Role", "old": emp.role, "new": new_role})
            if new_contact != (emp.contact_number or None):
                item_changes.append({"field": "Contact Number", "old": emp.contact_number or "", "new": new_contact or ""})
            if new_province != (emp.province or None):
                item_changes.append({"field": "Province", "old": emp.province or "", "new": new_province or ""})
            if new_city != (emp.city or None):
                item_changes.append({"field": "City", "old": emp.city or "", "new": new_city or ""})
            if new_barangay != (emp.barangay or None):
                item_changes.append({"field": "Barangay", "old": emp.barangay or "", "new": new_barangay or ""})
            if new_address != (emp.address or None):
                item_changes.append({"field": "Address", "old": emp.address or "", "new": new_address or ""})
            if new_salary is not None and new_salary != (emp.basic_salary or Decimal("0")):
                item_changes.append({"field": "Basic Salary", "old": str(emp.basic_salary or 0), "new": str(new_salary)})
            if new_sss != (emp.sss_number or None):
                item_changes.append({"field": "SSS #", "old": emp.sss_number or "", "new": new_sss or ""})
            if new_philhealth != (emp.philhealth_number or None):
                item_changes.append({"field": "PhilHealth #", "old": emp.philhealth_number or "", "new": new_philhealth or ""})
            if new_tin != (emp.tin_number or None):
                item_changes.append({"field": "TIN #", "old": emp.tin_number or "", "new": new_tin or ""})

            if item_changes:
                changes.append({
                    "row": row_num, "sku": str(emp_id),
                    "name": f"{emp.first_name or ''} {emp.last_name or ''}".strip(),
                    "changes": item_changes,
                })
            else:
                skipped += 1

        return Response({
            "changes": changes, "skipped": skipped, "errors": errors,
            "summary": f"{len(changes)} employees to update, {skipped} unchanged, {len(errors)} errors.",
        })


class EmployeeBulkUpdateView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def post(self, request):
        """Upload an XLSX file to bulk-update employees."""
        import threading
        import openpyxl

        xlsx_file = request.FILES.get("file")
        if not xlsx_file:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        if not xlsx_file.name.endswith((".xlsx", ".xlsm")):
            return Response({"detail": "Only .xlsx files are supported."}, status=status.HTTP_400_BAD_REQUEST)
        if xlsx_file.size > 5 * 1024 * 1024:
            return Response({"detail": "File too large. Maximum 5 MB."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        except Exception:
            return Response({"detail": "Could not parse the uploaded file."}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        if not rows:
            return Response({"detail": "The file contains no data rows."}, status=status.HTTP_400_BAD_REQUEST)

        user_id = request.user.id

        def _process():
            from decimal import Decimal, InvalidOperation
            from django.db import transaction
            try:
                all_employees = {e.id: e for e in CustomUser.objects.filter(is_deleted=False)}
                valid_roles = {"admin", "manager", "clerk", "technician"}
                updated = []
                skipped = []
                errors = []

                for row_num, row in enumerate(rows, 2):
                    if len(row) < 13:
                        errors.append({"row": row_num, "error": "Row has fewer than 13 columns."})
                        continue

                    try:
                        emp_id = int(row[0]) if row[0] is not None else None
                    except (ValueError, TypeError):
                        errors.append({"row": row_num, "error": f"Invalid ID: {row[0]}"})
                        continue

                    if not emp_id:
                        continue

                    emp = all_employees.get(emp_id)
                    if not emp:
                        errors.append({"row": row_num, "sku": str(emp_id), "error": "Employee ID not found."})
                        continue

                    new_first = str(row[1] or "").strip()
                    new_last = str(row[2] or "").strip()
                    new_role = str(row[3] or "").strip()
                    new_contact = str(row[4] or "").strip() or None
                    new_province = str(row[5] or "").strip() or None
                    new_city = str(row[6] or "").strip() or None
                    new_barangay = str(row[7] or "").strip() or None
                    new_address = str(row[8] or "").strip() or None

                    try:
                        new_salary = Decimal(str(row[9])) if row[9] is not None and str(row[9]).strip() != "" else None
                    except (InvalidOperation, ValueError) as e:
                        errors.append({"row": row_num, "sku": str(emp_id), "error": f"Invalid salary: {e}"})
                        continue

                    new_sss = str(row[10] or "").strip() or None
                    new_philhealth = str(row[11] or "").strip() or None
                    new_tin = str(row[12] or "").strip() or None

                    if new_role and new_role not in valid_roles:
                        errors.append({"row": row_num, "sku": str(emp_id), "error": f"Invalid role: {new_role}"})
                        continue

                    changed = False
                    if new_first and new_first != (emp.first_name or ""):
                        emp.first_name = new_first
                        changed = True
                    if new_last and new_last != (emp.last_name or ""):
                        emp.last_name = new_last
                        changed = True
                    if new_role and new_role != emp.role:
                        emp.role = new_role
                        changed = True
                    if new_contact != (emp.contact_number or None):
                        emp.contact_number = new_contact
                        changed = True
                    if new_province != (emp.province or None):
                        emp.province = new_province
                        changed = True
                    if new_city != (emp.city or None):
                        emp.city = new_city
                        changed = True
                    if new_barangay != (emp.barangay or None):
                        emp.barangay = new_barangay
                        changed = True
                    if new_address != (emp.address or None):
                        emp.address = new_address
                        changed = True
                    if new_salary is not None and new_salary != (emp.basic_salary or Decimal("0")):
                        emp.basic_salary = new_salary
                        changed = True
                    if new_sss != (emp.sss_number or None):
                        emp.sss_number = new_sss
                        changed = True
                    if new_philhealth != (emp.philhealth_number or None):
                        emp.philhealth_number = new_philhealth
                        changed = True
                    if new_tin != (emp.tin_number or None):
                        emp.tin_number = new_tin
                        changed = True

                    if changed:
                        updated.append(emp)
                    else:
                        skipped.append(str(emp_id))

                with transaction.atomic():
                    for e in updated:
                        e.save()

                detail = f"Updated {len(updated)} employees, skipped {len(skipped)} unchanged, {len(errors)} errors."
                _notify_employee_bulk_update(user_id, {
                    "updated": len(updated), "skipped": len(skipped),
                    "errors": errors, "detail": detail,
                })
            except Exception:
                import logging
                logging.getLogger(__name__).exception("Employee bulk update failed")
                _notify_employee_bulk_update_failed(user_id)

        threading.Thread(target=_process, daemon=True).start()
        return Response(
            {"detail": "Bulk update started. You will be notified when it's done."},
            status=status.HTTP_202_ACCEPTED,
        )


def _notify_employee_bulk_update(user_id, result):
    try:
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer
        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f"notifications_{user_id}",
                {"type": "send_notification", "data": {
                    "event": "export_ready", "export_type": "employee_bulk_update",
                    "title": "Employee Bulk Update Complete", "message": result["detail"],
                    "result": result,
                }},
            )
    except Exception:
        import logging
        logging.getLogger(__name__).exception("Failed to send employee_bulk_update via WebSocket")


def _notify_employee_bulk_update_failed(user_id):
    try:
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer
        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f"notifications_{user_id}",
                {"type": "send_notification", "data": {
                    "event": "export_failed", "export_type": "employee_bulk_update",
                    "title": "Employee Bulk Update Failed",
                    "message": "Failed to process the bulk update. Please try again.",
                }},
            )
    except Exception:
        import logging
        logging.getLogger(__name__).exception("Failed to send employee_bulk_update_failed via WebSocket")
