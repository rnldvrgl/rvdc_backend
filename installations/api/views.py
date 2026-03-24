from django.core.cache import cache
from django_filters.rest_framework import DjangoFilterBackend
from installations.api.filters import (
    AirconModelFilter,
    AirconUnitFilter,
)
from installations.api.serializers import (
    AirconBrandSerializer,
    AirconInstallationCreateSerializer,
    AirconModelSerializer,
    AirconReservationSerializer,
    AirconSaleSerializer,
    AirconUnitSerializer,
    FreeCleaningEligibilitySerializer,
    FreeCleaningRedemptionSerializer,
    WarrantyClaimApproveSerializer,
    WarrantyClaimCancelSerializer,
    WarrantyClaimCreateSerializer,
    WarrantyClaimRejectSerializer,
    WarrantyClaimSerializer,
    WarrantyEligibilitySerializer,
)
from installations.models import (
    AirconBrand,
    AirconModel,
    AirconUnit,
    WarrantyClaim,
)
from rest_framework import filters as drf_filters
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from utils.filters.options import (
    get_aircon_brand_options,
    get_aircon_model_options,
    get_aircon_type_options,
)
from utils.filters.role_filters import get_role_based_filter_response
from utils.query import (
    get_role_filtered_queryset,
)
from utils.soft_delete import SoftDeleteViewSetMixin


class AirconBrandViewSet(viewsets.ModelViewSet):
    queryset = AirconBrand.objects.all()
    serializer_class = AirconBrandSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        drf_filters.SearchFilter,
        drf_filters.OrderingFilter,
    ]
    search_fields = ["name"]
    ordering_fields = ["name"]

    # No role filtering - brands are global resources


class AirconModelViewSet(viewsets.ModelViewSet):
    queryset = AirconModel.objects.select_related('brand').prefetch_related('price_history').all()
    serializer_class = AirconModelSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        drf_filters.SearchFilter,
        drf_filters.OrderingFilter,
    ]
    filterset_class = AirconModelFilter
    search_fields = ["name", "brand__name"]
    ordering_fields = ["name", "retail_price"]

    # No role filtering - models are global resources

    @action(detail=False, methods=["get"], url_path="filters")
    def get_filters(self, request):
        filters_config = {
            "brand": {"options": get_aircon_brand_options},
            "aircon_type": {"options": get_aircon_type_options},
            "is_inverter": {
                "options": lambda: [
                    {"label": "Yes", "value": "true"},
                    {"label": "No", "value": "false"},
                ]
            },
            "has_discount": {
                "options": lambda: [
                    {"label": "Yes", "value": "true"},
                    {"label": "No", "value": "false"},
                ]
            },
        }
        ordering_config = [
            {"label": "Name", "value": "name"},
            {"label": "Retail Price", "value": "retail_price"},
        ]
        return get_role_based_filter_response(request, filters_config, ordering_config)

    @action(
        detail=False,
        methods=["get"],
        url_path="bulk-template",
        permission_classes=[IsAuthenticated],
    )
    def bulk_template(self, request):
        """Download an XLSX file pre-filled with all aircon models."""
        import io
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        models = AirconModel.objects.select_related("brand").order_by("brand__name", "name")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Aircon Models"

        headers = [
            "ID", "Brand", "Name", "Aircon Type", "Horsepower", "Is Inverter",
            "Retail Price", "Cost Price", "Promo Price",
            "Parts Warranty (months)", "Compressor Warranty (months)", "Labor Warranty (months)",
        ]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, m in enumerate(models, 2):
            ws.cell(row=row_idx, column=1, value=m.id).border = thin_border
            ws.cell(row=row_idx, column=2, value=m.brand.name).border = thin_border
            ws.cell(row=row_idx, column=3, value=m.name).border = thin_border
            ws.cell(row=row_idx, column=4, value=m.aircon_type).border = thin_border
            ws.cell(row=row_idx, column=5, value=m.horsepower).border = thin_border
            ws.cell(row=row_idx, column=6, value="Yes" if m.is_inverter else "No").border = thin_border
            ws.cell(row=row_idx, column=7, value=float(m.retail_price)).border = thin_border
            ws.cell(row=row_idx, column=8, value=float(m.cost_price)).border = thin_border
            ws.cell(row=row_idx, column=9, value=float(m.promo_price) if m.promo_price else "").border = thin_border
            ws.cell(row=row_idx, column=10, value=m.parts_warranty_months).border = thin_border
            ws.cell(row=row_idx, column=11, value=m.compressor_warranty_months).border = thin_border
            ws.cell(row=row_idx, column=12, value=m.labor_warranty_months).border = thin_border

        ws.sheet_properties.tabColor = "1F4E79"
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="aircon_model_template.xlsx"'
        return resp

    @action(
        detail=False,
        methods=["post"],
        url_path="bulk-preview",
        permission_classes=[IsAuthenticated],
    )
    def bulk_preview(self, request):
        """Upload an XLSX to preview aircon model changes."""
        import openpyxl
        from decimal import Decimal, InvalidOperation

        xlsx_file = request.FILES.get("file")
        if not xlsx_file:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        if not xlsx_file.name.endswith((".xlsx", ".xlsm")):
            return Response({"detail": "Only .xlsx files are supported."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        except Exception:
            return Response({"detail": "Could not parse the uploaded file."}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        if not rows:
            return Response({"detail": "The file contains no data rows."}, status=status.HTTP_400_BAD_REQUEST)

        all_models = {m.id: m for m in AirconModel.objects.select_related("brand").all()}

        changes = []
        skipped = 0
        errors = []

        for row_num, row in enumerate(rows, 2):
            if len(row) < 12:
                errors.append({"row": row_num, "error": "Row has fewer than 12 columns."})
                continue

            try:
                model_id = int(row[0]) if row[0] is not None else None
            except (ValueError, TypeError):
                errors.append({"row": row_num, "error": f"Invalid ID: {row[0]}"})
                continue

            if not model_id:
                continue

            model = all_models.get(model_id)
            if not model:
                errors.append({"row": row_num, "sku": str(model_id), "error": "Model ID not found."})
                continue

            new_name = str(row[2] or "").strip()

            try:
                new_retail = Decimal(str(row[6])) if row[6] is not None and str(row[6]).strip() != "" else None
                new_cost = Decimal(str(row[7])) if row[7] is not None and str(row[7]).strip() != "" else None
                new_promo = Decimal(str(row[8])) if row[8] is not None and str(row[8]).strip() != "" else None
            except (InvalidOperation, ValueError) as e:
                errors.append({"row": row_num, "sku": str(model_id), "error": f"Invalid price: {e}"})
                continue

            try:
                new_parts = int(row[9]) if row[9] is not None and str(row[9]).strip() != "" else None
                new_compressor = int(row[10]) if row[10] is not None and str(row[10]).strip() != "" else None
                new_labor = int(row[11]) if row[11] is not None and str(row[11]).strip() != "" else None
            except (ValueError, TypeError) as e:
                errors.append({"row": row_num, "sku": str(model_id), "error": f"Invalid warranty value: {e}"})
                continue

            item_changes = []
            if new_name and new_name != model.name:
                item_changes.append({"field": "Name", "old": model.name, "new": new_name})
            if new_retail is not None and new_retail != model.retail_price:
                item_changes.append({"field": "Retail Price", "old": str(model.retail_price), "new": str(new_retail)})
            if new_cost is not None and new_cost != model.cost_price:
                item_changes.append({"field": "Cost Price", "old": str(model.cost_price), "new": str(new_cost)})
            if new_promo is not None and new_promo != (model.promo_price or Decimal("0")):
                item_changes.append({"field": "Promo Price", "old": str(model.promo_price or 0), "new": str(new_promo)})
            if new_parts is not None and new_parts != model.parts_warranty_months:
                item_changes.append({"field": "Parts Warranty", "old": str(model.parts_warranty_months), "new": str(new_parts)})
            if new_compressor is not None and new_compressor != model.compressor_warranty_months:
                item_changes.append({"field": "Compressor Warranty", "old": str(model.compressor_warranty_months), "new": str(new_compressor)})
            if new_labor is not None and new_labor != model.labor_warranty_months:
                item_changes.append({"field": "Labor Warranty", "old": str(model.labor_warranty_months), "new": str(new_labor)})

            if item_changes:
                changes.append({"row": row_num, "sku": str(model_id), "name": f"{model.brand.name} {model.name}", "changes": item_changes})
            else:
                skipped += 1

        return Response({
            "changes": changes, "skipped": skipped, "errors": errors,
            "summary": f"{len(changes)} models to update, {skipped} unchanged, {len(errors)} errors.",
        })

    @action(
        detail=False,
        methods=["post"],
        url_path="bulk-update",
        permission_classes=[IsAuthenticated],
    )
    def bulk_update(self, request):
        """Upload an XLSX file to bulk-update aircon models."""
        import threading
        import openpyxl

        xlsx_file = request.FILES.get("file")
        if not xlsx_file:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        if not xlsx_file.name.endswith((".xlsx", ".xlsm")):
            return Response({"detail": "Only .xlsx files are supported."}, status=status.HTTP_400_BAD_REQUEST)
        if xlsx_file.size > 5 * 1024 * 1024:
            return Response({"detail": "File too large. Maximum 5 MB."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        except Exception:
            return Response({"detail": "Could not parse the uploaded file."}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        if not rows:
            return Response({"detail": "The file contains no data rows."}, status=status.HTTP_400_BAD_REQUEST)

        user_id = request.user.id

        def _process():
            from decimal import Decimal, InvalidOperation
            from django.db import transaction
            try:
                all_models = {m.id: m for m in AirconModel.objects.select_related("brand").all()}
                updated = []
                skipped = []
                errors = []

                for row_num, row in enumerate(rows, 2):
                    if len(row) < 12:
                        errors.append({"row": row_num, "error": "Row has fewer than 12 columns."})
                        continue

                    try:
                        model_id = int(row[0]) if row[0] is not None else None
                    except (ValueError, TypeError):
                        errors.append({"row": row_num, "error": f"Invalid ID: {row[0]}"})
                        continue

                    if not model_id:
                        continue

                    model = all_models.get(model_id)
                    if not model:
                        errors.append({"row": row_num, "sku": str(model_id), "error": "Model ID not found."})
                        continue

                    new_name = str(row[2] or "").strip()

                    try:
                        new_retail = Decimal(str(row[6])) if row[6] is not None and str(row[6]).strip() != "" else None
                        new_cost = Decimal(str(row[7])) if row[7] is not None and str(row[7]).strip() != "" else None
                        new_promo = Decimal(str(row[8])) if row[8] is not None and str(row[8]).strip() != "" else None
                    except (InvalidOperation, ValueError) as e:
                        errors.append({"row": row_num, "sku": str(model_id), "error": f"Invalid price: {e}"})
                        continue

                    try:
                        new_parts = int(row[9]) if row[9] is not None and str(row[9]).strip() != "" else None
                        new_compressor = int(row[10]) if row[10] is not None and str(row[10]).strip() != "" else None
                        new_labor = int(row[11]) if row[11] is not None and str(row[11]).strip() != "" else None
                    except (ValueError, TypeError) as e:
                        errors.append({"row": row_num, "sku": str(model_id), "error": f"Invalid warranty value: {e}"})
                        continue

                    changed = False
                    if new_name and new_name != model.name:
                        model.name = new_name
                        changed = True
                    if new_retail is not None and new_retail != model.retail_price:
                        model.retail_price = new_retail
                        changed = True
                    if new_cost is not None and new_cost != model.cost_price:
                        model.cost_price = new_cost
                        changed = True
                    if new_promo is not None and new_promo != (model.promo_price or Decimal("0")):
                        model.promo_price = new_promo
                        changed = True
                    if new_parts is not None and new_parts != model.parts_warranty_months:
                        model.parts_warranty_months = new_parts
                        changed = True
                    if new_compressor is not None and new_compressor != model.compressor_warranty_months:
                        model.compressor_warranty_months = new_compressor
                        changed = True
                    if new_labor is not None and new_labor != model.labor_warranty_months:
                        model.labor_warranty_months = new_labor
                        changed = True

                    if changed:
                        updated.append(model)
                    else:
                        skipped.append(str(model_id))

                with transaction.atomic():
                    for m in updated:
                        m.save()

                detail = f"Updated {len(updated)} aircon models, skipped {len(skipped)} unchanged, {len(errors)} errors."
                _notify_aircon_model_bulk_update(user_id, {
                    "updated": len(updated), "skipped": len(skipped),
                    "errors": errors, "detail": detail,
                })
            except Exception:
                import logging
                logging.getLogger(__name__).exception("Aircon model bulk update failed")
                _notify_aircon_model_bulk_update_failed(user_id)

        threading.Thread(target=_process, daemon=True).start()
        return Response(
            {"detail": "Bulk update started. You will be notified when it's done."},
            status=status.HTTP_202_ACCEPTED,
        )


def _notify_aircon_model_bulk_update(user_id, result):
    try:
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer
        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f"notifications_{user_id}",
                {"type": "send_notification", "data": {
                    "event": "export_ready", "export_type": "aircon_model_bulk_update",
                    "title": "Aircon Model Bulk Update Complete", "message": result["detail"],
                    "result": result,
                }},
            )
    except Exception:
        import logging
        logging.getLogger(__name__).exception("Failed to send aircon_model_bulk_update via WebSocket")


def _notify_aircon_model_bulk_update_failed(user_id):
    try:
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer
        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f"notifications_{user_id}",
                {"type": "send_notification", "data": {
                    "event": "export_failed", "export_type": "aircon_model_bulk_update",
                    "title": "Aircon Model Bulk Update Failed",
                    "message": "Failed to process the bulk update. Please try again.",
                }},
            )
    except Exception:
        import logging
        logging.getLogger(__name__).exception("Failed to send aircon_model_bulk_update_failed via WebSocket")


class AirconUnitViewSet(SoftDeleteViewSetMixin, viewsets.ModelViewSet):
    """
    Aircon unit inventory management.
    """
    allow_hard_delete = True
    """
    
    This viewset manages the inventory of aircon units available for installation.
    Units are added to inventory and later linked to sales/installations through
    the installation workflow.

    Endpoints:
    - GET /aircon-units/ - List all units in inventory
          Query params:
          - is_available_for_sale=true: Units available to purchase
          - is_available_for_installation=true: Units available for installation scheduling
          - is_reserved=true/false: Filter by reservation status
    - POST /aircon-units/ - Add new unit to inventory
    - GET /aircon-units/{id}/ - Get unit details
    - PUT/PATCH /aircon-units/{id}/ - Update unit information
    - DELETE /aircon-units/{id}/ - Remove unit from inventory
    - GET /aircon-units/available/ - List available units for sale
    - POST /aircon-units/sell/ - Sell one or more units (creates sale transaction)
    - POST /aircon-units/{id}/reserve/ - Reserve a unit for a client
    - POST /aircon-units/{id}/release-reservation/ - Release reservation
    - POST /aircon-units/{id}/create-installation/ - Create installation service (reserves unit if not sold)
    - GET /aircon-units/stock-report/ - Get inventory stock report
    """

    queryset = AirconUnit.objects.select_related(
        'model__brand', 'stall', 'installation_service', 'installation_service__client',
        'reserved_by', 'sale__client'
    ).prefetch_related(
        'model__price_history',
    ).all()
    serializer_class = AirconUnitSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        drf_filters.SearchFilter,
        drf_filters.OrderingFilter,
    ]
    filterset_class = AirconUnitFilter
    search_fields = ["serial_number", "model__name", "model__brand__name"]
    ordering_fields = ["serial_number", "created_at"]
    ordering = ["-created_at"]

    def get_queryset(self):
        return get_role_filtered_queryset(
            self.request, super().get_queryset().filter(is_deleted=False)
        )

    @action(detail=False, methods=["get"], url_path="available")
    def available_units(self, request):
        """
        Get available units for sale.

        Query params:
        - model: Filter by model ID
        - brand: Filter by brand ID
        """
        from installations.business_logic import AirconInventoryManager

        model_id = request.query_params.get('model')
        brand_id = request.query_params.get('brand')

        model = None
        brand = None

        if model_id:
            from installations.models import AirconModel
            try:
                model = AirconModel.objects.get(id=model_id)
            except AirconModel.DoesNotExist:
                pass

        if brand_id:
            from installations.models import AirconBrand
            try:
                brand = AirconBrand.objects.get(id=brand_id)
            except AirconBrand.DoesNotExist:
                pass

        units = AirconInventoryManager.get_available_units(model=model, brand=brand)
        serializer = self.get_serializer(units, many=True)

        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path="sell")
    def sell_units(self, request):
        """
        Sell one or more aircon units.

        Request body:
        {
            "unit_ids": [1, 2, 3],
            "client_id": 123,
            "payment_type": "cash"
        }

        Response:
        {
            "units": [...],
            "sale_transaction": {...},
            "total_amount": "50000.00"
        }
        """
        serializer = AirconSaleSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        result = serializer.save()

        # Return sale details
        return Response({
            'units': AirconUnitSerializer(result.get('units') or [result.get('unit')], many=True).data,
            'sale_transaction_id': result['sale_transaction'].id if result.get('sale_transaction') else None,
            'total_amount': str(result.get('total_amount') or result.get('sale_price')),
            'client_id': result['client'].id,
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["post"], url_path="reserve")
    def reserve_unit(self, request, pk=None):
        """
        Reserve a unit for a client.

        Request body:
        {
            "client_id": 123
        }
        """
        unit = self.get_object()

        data = {
            'unit_id': unit.id,
            'client_id': request.data.get('client_id')
        }

        serializer = AirconReservationSerializer(data=data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        result = serializer.save()

        return Response(
            AirconUnitSerializer(result).data,
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=["post"], url_path="release-reservation")
    def release_reservation(self, request, pk=None):
        """Release reservation on a unit."""
        from installations.business_logic import AirconInventoryManager

        unit = self.get_object()

        if not unit.is_reserved:
            return Response(
                {'error': 'Unit is not reserved.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        result = AirconInventoryManager.release_reservation(unit)

        return Response(
            AirconUnitSerializer(result).data,
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=["post"], url_path="create-installation")
    def create_installation(self, request, pk=None):
        """
        Create installation service for a sold unit.

        Request body:
        {
            "scheduled_date": "2024-01-15",
            "scheduled_time": "14:00:00",
            "labor_fee": "500.00",
            "apply_free_installation": false,
            "copper_tube_length": 25
        }
        """
        unit = self.get_object()

        data = request.data.copy()
        data['unit_id'] = unit.id

        serializer = AirconInstallationCreateSerializer(data=data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        result = serializer.save()

        return Response({
            'service_id': result['service'].id,
            'installation_id': result['installation'].id,
            'unit_id': result['unit'].id,
            'appliance_id': result['appliance'].id,
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["get"], url_path="stock-report")
    def stock_report(self, request):
        """
        Get aircon unit inventory stock report.

        Query params:
        - model: Filter by model ID

        Response:
        {
            "total": 100,
            "available": 75,
            "reserved": 10,
            "sold": 15
        }
        """
        from installations.business_logic import AirconInventoryManager

        model_id = request.query_params.get('model')
        model = None

        if model_id:
            from installations.models import AirconModel
            try:
                model = AirconModel.objects.get(id=model_id)
            except AirconModel.DoesNotExist:
                pass

        report = AirconInventoryManager.check_stock_level(model=model)

        return Response(report, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="filters")
    def get_filters(self, request):
        filters_config = {
            "model": {"options": get_aircon_model_options},
            "is_sold": {
                "options": lambda: [
                    {"label": "Sold", "value": "true"},
                    {"label": "Not Sold", "value": "false"},
                ]
            },
            "is_installed": {
                "options": lambda: [
                    {"label": "Installed", "value": "true"},
                    {"label": "Not Installed", "value": "false"},
                ]
            },
            "is_available": {
                "options": lambda: [
                    {"label": "Available", "value": "true"},
                    {"label": "Not Available", "value": "false"},
                ]
            },
        }
        ordering_config = [
            {"label": "Serial Number", "value": "serial_number"},
            {"label": "Created At", "value": "created_at"},
        ]
        return get_role_based_filter_response(request, filters_config, ordering_config)


class WarrantyClaimViewSet(viewsets.ModelViewSet):
    """
    Warranty claim management and free cleaning redemption.

    Endpoints:
    - GET /warranty-claims/ - List all warranty claims
    - POST /warranty-claims/ - Create new warranty claim
    - GET /warranty-claims/{id}/ - Get claim details
    - PUT/PATCH /warranty-claims/{id}/ - Update claim
    - DELETE /warranty-claims/{id}/ - Delete claim
    - POST /warranty-claims/{id}/approve/ - Approve claim and create service
    - POST /warranty-claims/{id}/reject/ - Reject claim with reason
    - POST /warranty-claims/{id}/cancel/ - Cancel claim
    - POST /warranty-claims/{id}/complete/ - Mark claim as completed
    - POST /warranty-claims/check-eligibility/ - Check warranty eligibility
    - POST /warranty-claims/redeem-free-cleaning/ - Redeem free cleaning
    - POST /warranty-claims/check-free-cleaning/ - Check free cleaning eligibility
    """

    queryset = WarrantyClaim.objects.all()
    serializer_class = WarrantyClaimSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        drf_filters.SearchFilter,
        drf_filters.OrderingFilter,
    ]
    search_fields = [
        "unit__serial_number",
        "unit__model__name",
        "unit__sale__client__name",
        "issue_description",
    ]
    ordering_fields = ["claim_date", "status", "created_at"]
    filterset_fields = ["status", "claim_type", "unit", "is_valid_claim"]

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'unit',
            'unit__model',
            'unit__model__brand',
            'unit__sale',
            'unit__sale__client',
            'service',
            'reviewed_by'
        )
        return get_role_filtered_queryset(self.request, queryset, stall_field="unit__stall")

    def create(self, request, *args, **kwargs):
        """
        Create a warranty claim.

        Request body:
        {
            "unit_id": 123,
            "claim_type": "repair",
            "issue_description": "Unit not cooling properly",
            "customer_notes": "Started happening last week"
        }
        """
        idempotency_key = request.headers.get("Idempotency-Key")
        if not idempotency_key:
            import hashlib, json
            body_hash = hashlib.sha256(
                json.dumps(request.data, sort_keys=True, default=str).encode()
            ).hexdigest()[:16]
            idempotency_key = f"{request.user.id}:{body_hash}"

        cache_key = f"warranty_claim_create_idempotency:{idempotency_key}"

        if cache.get(cache_key):
            return Response(
                {"detail": "Duplicate request detected. This warranty claim was already submitted."},
                status=status.HTTP_409_CONFLICT,
            )

        cache.set(cache_key, True, timeout=30)

        try:
            serializer = WarrantyClaimCreateSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            claim = serializer.save()

            return Response(
                WarrantyClaimSerializer(claim).data,
                status=status.HTTP_201_CREATED
            )
        except Exception:
            cache.delete(cache_key)
            raise

    @action(detail=True, methods=["post"], url_path="approve")
    def approve_claim(self, request, pk=None):
        """
        Approve a warranty claim and optionally create service.

        Request body:
        {
            "technician_assessment": "Confirmed defective compressor",
            "create_service": true,
            "scheduled_date": "2024-01-20",
            "scheduled_time": "10:00:00"
        }
        """
        claim = self.get_object()

        serializer = WarrantyClaimApproveSerializer(
            data=request.data,
            context={'request': request, 'claim': claim}
        )
        serializer.is_valid(raise_exception=True)
        result = serializer.save()

        response_data = {
            'claim': WarrantyClaimSerializer(result['claim']).data,
        }

        if result.get('service'):
            from services.api.serializers import ServiceSerializer
            response_data['service'] = ServiceSerializer(result['service']).data

        return Response(response_data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="reject")
    def reject_claim(self, request, pk=None):
        """
        Reject a warranty claim.

        Request body:
        {
            "rejection_reason": "Unit damage caused by improper use",
            "is_valid_claim": false
        }
        """
        claim = self.get_object()

        serializer = WarrantyClaimRejectSerializer(
            data=request.data,
            context={'request': request, 'claim': claim}
        )
        serializer.is_valid(raise_exception=True)
        claim = serializer.save()

        return Response(
            WarrantyClaimSerializer(claim).data,
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=["post"], url_path="cancel")
    def cancel_claim(self, request, pk=None):
        """
        Cancel a warranty claim.

        Request body:
        {
            "cancellation_reason": "Customer no longer needs service"
        }
        """
        claim = self.get_object()

        serializer = WarrantyClaimCancelSerializer(
            data=request.data,
            context={'claim': claim}
        )
        serializer.is_valid(raise_exception=True)
        claim = serializer.save()

        return Response(
            WarrantyClaimSerializer(claim).data,
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=["post"], url_path="complete")
    def complete_claim(self, request, pk=None):
        """
        Mark a warranty claim as completed.
        """
        from installations.business_logic import WarrantyClaimManager

        claim = self.get_object()
        claim = WarrantyClaimManager.complete_claim(claim)

        return Response(
            WarrantyClaimSerializer(claim).data,
            status=status.HTTP_200_OK
        )

    @action(detail=False, methods=["post"], url_path="check-eligibility")
    def check_warranty_eligibility(self, request):
        """
        Check if a unit is eligible for warranty claim.

        Request body:
        {
            "unit_id": 123
        }

        Response:
        {
            "eligible": true,
            "reason": "Unit is under warranty",
            "warranty_days_left": 180,
            "warranty_end_date": "2024-07-15"
        }
        """
        serializer = WarrantyEligibilitySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        result = serializer.check()

        return Response(result, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path="redeem-free-cleaning")
    def redeem_free_cleaning(self, request):
        """
        Redeem free cleaning for an aircon unit.

        Request body:
        {
            "unit_id": 123,
            "scheduled_date": "2024-01-20",
            "scheduled_time": "14:00:00"
        }

        Response:
        {
            "service": {...},
            "unit": {...}
        }
        """
        serializer = FreeCleaningRedemptionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        result = serializer.save()

        from services.api.serializers import ServiceSerializer

        return Response({
            'service': ServiceSerializer(result['service']).data,
            'unit': AirconUnitSerializer(result['unit']).data,
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["post"], url_path="redeem-free-cleaning-batch")
    def redeem_free_cleaning_batch(self, request):
        """
        Redeem free cleaning for multiple aircon units under one client.
        Creates a single cleaning service with all units as appliances.

        Request body:
        {
            "client_id": 1,
            "unit_ids": [1, 2, 3],
            "scheduled_date": "2024-01-20",
            "scheduled_time": "14:00:00"
        }

        Response:
        {
            "service": {...},
            "units": [...]
        }
        """
        from installations.api.serializers import FreeCleaningBatchRedemptionSerializer

        serializer = FreeCleaningBatchRedemptionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        result = serializer.save()

        from services.api.serializers import ServiceSerializer

        return Response({
            'service': ServiceSerializer(result['service']).data,
            'units': AirconUnitSerializer(result['units'], many=True).data,
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["post"], url_path="check-free-cleaning")
    def check_free_cleaning_eligibility(self, request):
        """
        Check if a unit is eligible for free cleaning redemption.

        Request body:
        {
            "unit_id": 123
        }

        Response:
        {
            "eligible": true,
            "reason": "Unit is eligible for free cleaning",
            "warranty_days_left": 180
        }
        """
        serializer = FreeCleaningEligibilitySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        result = serializer.check()

        return Response(result, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="filters")
    def get_filters(self, request):
        filters_config = {
            "status": {
                "options": lambda: [
                    {"label": label, "value": value}
                    for value, label in WarrantyClaim.ClaimStatus.choices
                ]
            },
            "claim_type": {
                "options": lambda: [
                    {"label": label, "value": value}
                    for value, label in WarrantyClaim.ClaimType.choices
                ]
            },
            "is_valid_claim": {
                "options": lambda: [
                    {"label": "Valid", "value": "true"},
                    {"label": "Invalid", "value": "false"},
                ]
            },
        }
        ordering_config = [
            {"label": "Claim Date", "value": "claim_date"},
            {"label": "Status", "value": "status"},
            {"label": "Created At", "value": "created_at"},
        ]
        return get_role_based_filter_response(request, filters_config, ordering_config)
